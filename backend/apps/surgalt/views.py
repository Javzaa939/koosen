import os
# import logging
import json
import ast
import traceback
from openpyxl import load_workbook
from core.fns import WithChoices
from rest_framework import mixins
from rest_framework import generics
from rest_framework.views import APIView
from rest_framework.filters import SearchFilter
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import permission_classes

from main.utils.function.pagination import CustomPagination
from main.utils.function.utils import override_get_queryset, save_data_with_signals

from main.utils.function.utils import has_permission, get_domain_url, _filter_queries, get_teacher_queryset, pearson_corel, str2bool
from main.utils.file import save_file
from main.utils.file import remove_folder

from django.db import transaction
from django.db.models import Sum, Count, Q, Subquery, OuterRef,  Value, CharField, F, Case, When, IntegerField, FloatField, Max, Exists
from django.db.models.functions import Concat

from django.shortcuts import get_object_or_404
from django.conf import settings

from elselt.models import ElseltUser
from elselt.models import MentalUser
from elselt.models import AdmissionUserProfession

from lms.models import (
    Group,
    Student,
    TimeTable,
    SubOrgs,
    Salbars,
    Exam_repeat,
    LearningPlan,
    ScoreRegister,
    ExamTimeTable,
    Exam_to_group,
    LessonStandart,
    Lesson_title_plan,
    Lesson_to_teacher,
    ProfessionalDegree,
    ProfessionDefinition,
    AdmissionBottomScore,
    Profession_SongonKredit,
    Teachers,
    Challenge,
    QuestionChoices,
    Lesson_materials,
    Lesson_assignment,
    ChallengeStudents,
    TimeTable_to_group,
    ChallengeQuestions,
    TimeTable_to_student,
    Lesson_material_file,
    Lesson_assignment_student,
    Lesson_assignment_student_file,
    PsychologicalTestQuestions,
    PsychologicalQuestionChoices,
    PsychologicalQuestionTitle,
    PsychologicalTest,
    AdmissionRegisterProfession,
    TeacherScore,
    QuestionTitle,
    Lesson_teacher_scoretype,
    QuestionTitle,
    Score,
    CalculatedGpaOfDiploma,
    QuestionMainTitle,
    QuestioSubTitle
)

from core.models import (
    User,
    Employee,
)


from lms.models import get_image_path
from lms.models import get_choice_image_path

from elselt.serializer import ElseltApproveSerializer, MentalUserSerializer

from .serializers import ChallengeGroupsSerializer, ChallengeProfessionsSerializer, ChallengeReport2StudentsDetailSerializer, ChallengeReport2StudentsSerializer, LessonStandartSerializer, ChallengeQuestionsAnswersSerializer, ChallengeReport4Serializer
from .serializers import LessonTitlePlanSerializer
from .serializers import LessonStandartListSerializer
from .serializers import LessonStandartSerialzier
from .serializers import ProfessionDefinitionSerializer
from .serializers import LearningPlanSerializer
from .serializers import LearningPlanListSerializer
from .serializers import ProfessionDefinitionListSerializer
from .serializers import ProfessionLearningPlanSerializer
from .serializers import LearningPlanPrintSerializer
from .serializers import GroupSerializer
from .serializers import AdmissionBottomScoreSerializer
from .serializers import AdmissionBottomScoreListSerializer
from .serializers import ChallengeSerializer
from .serializers import ChallengeListSerializer
from .serializers import ChallengeQuestionListSerializer
from .serializers import GroupListSerializer
from .serializers import LessonAssignmentAssigmentListSerializer
from .serializers import LessonAssignmentAssigmentSerializer
from .serializers import LessonTitleSerializer
from .serializers import LessonMaterialSerializer
from .serializers import LessonMaterialListSerializer
from .serializers import LessonTeacherSerializer
from .serializers import ProfessionDefinitionJustProfessionSerializer
from .serializers import PsychologicalTestSerializer
from .serializers import PsychologicalTestQuestionsSerializer
from .serializers import PsychologicalTestScopeSerializer
from .serializers import TeachersSerializer
from .serializers import StudentSerializer
from .serializers import ElsegchSerializer
from .serializers import PsychologicalTestResultSerializer
from .serializers import PsychologicalTestParticipantsSerializer
from .serializers import TeacherExamTimeTableSerializer
from core.serializers import TeacherNameSerializer
from .serializers import QuestionTitleSerializer
from .serializers import LessonStandartCreateSerializer
from .serializers import ChallengeDetailSerializer,ChallengeStudentsSerializer,StudentChallengeSerializer,ChallengeDetailTableStudentsSerializer

from main.utils.function.utils import remove_key_from_dict, fix_format_date, get_domain_url
from main.utils.function.utils import null_to_none, get_lesson_choice_student, get_active_year_season, json_load
from main.utils.function.serializer import dynamic_serializer

@permission_classes([IsAuthenticated])
class LessonStandartAPIView(
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    generics.GenericAPIView
):
    """ Хичээлийн стандарт """

    queryset = LessonStandart.objects.all().order_by('-updated_at')
    serializer_class = LessonStandartSerializer
    pagination_class = CustomPagination

    filter_backends = [SearchFilter]
    search_fields = [
        'code',
        'name',
        'name_eng',
        'name_uig'
    ]

    @has_permission(must_permissions=['lms-study-lessonstandart-read'])
    def get(self, request, pk=None):
        " Хичээлийн стандартын жагсаалт "

        self.serializer_class = LessonStandartListSerializer

        department = request.query_params.get('department')
        category = request.query_params.get('category')

        if department:
            self.queryset = self.queryset.filter(department=department)

        # Хичээлийн Ангилал сонголтоор хайх
        if category:
            self.queryset = self.queryset.filter(category=category)

        if pk:
            # self.serializer_class = LessonStandartCreateSerializer
            standart = self.retrieve(request, pk).data
            return request.send_data(standart)

        less_standart_list = self.list(request).data

        return request.send_data(less_standart_list)

    @has_permission(must_permissions=['lms-study-lessonstandart-create'])
    @transaction.atomic
    def post(self, request):
        " хичээлийн стандартын шинээр үүсгэх "

        request_data = request.data
        self.serializer_class = LessonStandartCreateSerializer
        serializer = self.get_serializer(data=request_data)

        # transaction savepoint зарлах нь хэрэв алдаа гарвад roll back хийнэ
        sid = transaction.savepoint()

        try:
            serializer = self.serializer_class(data=request_data, many=False)
            if not serializer.is_valid():
                transaction.savepoint_rollback(sid)
                return request.send_error_valid(serializer.errors)

            serializer.save()

        except Exception as e:
            print(e)
            return request.send_error("ERR_002")

        return request.send_info("INF_001")

    @has_permission(must_permissions=['lms-study-lessonstandart-update'])
    @transaction.atomic
    def put(self, request, pk=None):
        " хичээлийн стандартын шинээр үүсгэх "

        request_data = request.data
        teachers_data = request.data.get("teachers")
        self.serializer_class = LessonStandartCreateSerializer

        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request_data, partial=True)

        if not serializer.is_valid(raise_exception=False):
            return request.send_error_valid(serializer.errors)

        self.perform_update(serializer)
        try:
            if teachers_data:
                teacher_ids = [teacher.get('id') for teacher in teachers_data]
                old_teacher_ids = Lesson_to_teacher.objects.filter(lesson=pk).values_list('teacher', flat=True)

                # Шинээр нэмэгдсэн багш
                add_ids = [item for item in teacher_ids if item not in old_teacher_ids]

                # Хуучин багшийг хасах
                remove_ids = [item for item in old_teacher_ids if item not in teacher_ids]

                old_teacher_qs = Lesson_to_teacher.objects.filter(lesson=pk)
                if remove_ids or add_ids:
                    if len(remove_ids) > 0:
                        teacher_qs = old_teacher_qs.filter(teacher__in=remove_ids)
                        # Дүнгийн төрөл
                        score_types = Lesson_teacher_scoretype.objects.filter(lesson_teacher__in=teacher_qs)
                        if len(score_types) > 0:
                            score_types.delete()

                        old_teacher_qs = old_teacher_qs.filter(teacher__in=remove_ids).delete()

                    # Шинээр нэмж байгаа бүрээр нь нэмэх
                    for teacher_id in add_ids:
                        Lesson_to_teacher.objects.update_or_create(
                            lesson_id=pk,
                            teacher_id=teacher_id
                        )
        except Exception as e:
            print(e)
            return request.send_info("INF_002", 'Багшийн системд дүн оруулсан учраас устгах боломжгүй')

        return request.send_info("INF_002")

    @has_permission(must_permissions=['lms-study-lessonstandart-delete'])
    def delete(self, request, pk=None):
        " хичээлийн стандартыг устгах "

        learn_plan = LearningPlan.objects.filter(Q(lesson=pk) | Q(previous_lesson=pk))

        timetable = TimeTable.objects.filter(lesson=pk)

        examtimetable = ExamTimeTable.objects.filter(lesson=pk)

        examrepeat = Exam_repeat.objects.filter(lesson=pk)

        score = ScoreRegister.objects.filter(lesson=pk)

        lesson_to_teacher = Lesson_to_teacher.objects.filter(lesson=pk)
        diplom_lesson = CalculatedGpaOfDiploma.objects.filter(lesson=pk)

        if learn_plan:
            learn_plan.delete()

        if timetable:
            timetable.delete()

        if examtimetable:
            examtimetable.delete()

        if examrepeat:
            examrepeat.delete()

        if len(lesson_to_teacher) > 0:
            return request.send_error("ERR_002", "Тухайн хичээлд холбогдсон багшийн мэдээлэлтэй холбогдосон устгах боломжгүй байна.")

        if len(score) > 0:
            return request.send_error("ERR_002", "Тухайн хичээлд холбогдсон дүнгийн мэдээлэл байгаа устгах боломжгүй байна.")

        if len(diplom_lesson) > 0:
            return request.send_error("ERR_002", "Тухайн хичээл нь төгсөлтийн ажилтай холбогдсон учраас устгах боломжгүй байна.")
        self.destroy(request, pk)
        return request.send_info("INF_003")

@permission_classes([IsAuthenticated])
class LessonTitlePlanAPIView(
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    generics.GenericAPIView
):
    """ Хичээлийн сэдэвчилсэн төлөвлөгөө """

    queryset = Lesson_title_plan.objects.all()
    serializer_class = LessonTitlePlanSerializer

    def get_queryset(self):
        return override_get_queryset(self)

    @has_permission(must_permissions=['lms-study-lessonstandart-read'])
    def get(self, request, pk=None, lessonID=None):
        " хичээлийн сэдэвчилсэн төлөвлөгөө жагсаалт "

        # Хичээлийн хайх
        if lessonID:
            self.queryset = self.queryset.filter(lesson_id=lessonID).order_by('week')


        if pk:
            standart = self.retrieve(request, pk).data
            return request.send_data(standart)

        less_standart_list = self.list(request).data
        return request.send_data(less_standart_list)

    @has_permission(must_permissions=['lms-study-lessonstandart-create'])
    def post(self, request):
        " хичээлийн сэдэвчилсэн төлөвлөгөө шинээр үүсгэх "

        request_data = request.data
        delete_plan_ids = request_data.get('delete_plan_ids')
        all_datas = request_data.get('all_datas')
        lesson_id = request_data.get('lesson_id')

        with transaction.atomic():
            try:
                if delete_plan_ids:
                    for delete_plan_id in delete_plan_ids:
                        title_plan = Lesson_title_plan.objects.filter(id=delete_plan_id)
                        if title_plan:
                            title_plan.delete()

                if all_datas:
                    for datas in all_datas:
                        plan_id = datas.get('id')
                        lesson = datas.get('lesson')
                        week = datas.get('week')
                        title = datas.get('title')
                        content = datas.get('content')
                        lesson_type = datas.get('lesson_type')
                        if lesson:
                            lesson=lesson_id

                        if plan_id:
                            Lesson_title_plan.objects.filter(id=plan_id).update(
                                week=week,
                                lesson_type=lesson_type,
                                title=title,
                                content=content,
                                lesson=lesson,
                            )
                        else:
                            Lesson_title_plan.objects.create(
                                week=week,
                                lesson_type=lesson_type,
                                title=title,
                                content=content,
                                lesson_id=lesson_id,
                            )

            except Exception as e:
                print(e.__str__())
                return request.send_error("ERR_002")

        return request.send_info("INF_013")
    @has_permission(must_permissions=['lms-study-lessonstandart-update'])
    def put(self, request, pk=None):
        " хичээлийн сэдэвчилсэн төлөвлөгөө шинээр үүсгэх "

        request_data = request.data
        teachers_data = request.data.get("teachers")
        instance = self.get_object()

        serializer = self.get_serializer(instance, data=request_data)

        if serializer.is_valid(raise_exception=False):
            is_success = False
            with transaction.atomic():
                try:
                    self.update(request).data
                    is_success = True
                except Exception:
                    raise
            if is_success:
                return request.send_info("INF_001")

            return request.send_error("ERR_002")
        else:
            error_obj = []
            for key in serializer.errors:
                msg = "Хоосон байна"
                if key == 'code':
                    msg = "Код давхцаж байна"

                return_error = {
                    "field": key,
                    "msg": msg
                }

                error_obj.append(return_error)

            if len(error_obj) > 0:
                return request.send_error("ERR_003", error_obj)

            return request.send_error("ERR_002")

    @has_permission(must_permissions=['lms-study-lessonstandart-delete'])
    def delete(self, request, pk=None):
        " хичээлийн сэдэвчилсэн төлөвлөгөө устгах "

        self.destroy(request, pk)
        return request.send_info("INF_003")


@permission_classes([IsAuthenticated])
class LessonStandartListAPIView(
    mixins.ListModelMixin,
    generics.GenericAPIView
):
    """ Хичээлийн стандартын жагсаалт """

    queryset = LessonStandart.objects.all()
    serializer_class = LessonStandartSerialzier

    filter_backends = [SearchFilter]
    search_fields = [
        'code',
        'name',
        'name_eng',
        'name_uig',
        'kredit',
        'category__category_name',
        'definition',
        'department__name'
    ]

    def get(self, request):

        school = request.query_params.get('school')
        # department = request.query_params.get('department')
        profession = request.query_params.get('profession')

        if school:
            self.queryset = self.queryset.filter(school=school)

        # if department:
        #     self.queryset = self.queryset.filter(department=department)

        if profession:
            lesson_ids = LearningPlan.objects.filter(profession=profession).values_list('lesson', flat=True)
            self.queryset = self.queryset.filter(id__in=lesson_ids)
        less_standart_list = self.list(request).data

        return request.send_data(less_standart_list)



@permission_classes([IsAuthenticated])
class ProfessionDefinitionAPIView(
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    generics.GenericAPIView
):
    """  Мэргэжлийн тодорхойлолт """

    queryset = ProfessionDefinition.objects.all()
    serializer_class = ProfessionDefinitionSerializer
    pagination_class = CustomPagination

    filter_backends = [SearchFilter]
    search_fields = ['name', 'code', 'profession_code', 'degree__degree_name']

    @has_permission(must_permissions=['lms-study-profession-read'])
    def get(self, request, pk=None):
        "  Мэргэжлийн тодорхойлолт жагсаалт "

        self.serializer_class = ProfessionDefinitionListSerializer

        degree = self.request.query_params.get('degree')
        department = self.request.query_params.get('department')
        school = self.request.query_params.get('schoolId')

        if school:
            self.queryset = self.queryset.filter(school=school)

        if degree:
            self.queryset = self.queryset.filter(degree=degree)

        if department:
            self.queryset = self.queryset.filter(department=department)

        self.queryset = (
            self.queryset
            .annotate(
                general_base=Subquery(Profession_SongonKredit.objects.filter(profession=OuterRef('pk'), lesson_level=LearningPlan.BASIC).values('songon_kredit')[:1]),
                professional_base=Subquery(Profession_SongonKredit.objects.filter(profession=OuterRef('pk'), lesson_level=LearningPlan.PROF_BASIC).values('songon_kredit')[:1]),
                professional_lesson=Subquery(Profession_SongonKredit.objects.filter(profession=OuterRef('pk'), lesson_level=LearningPlan.PROFESSION).values('songon_kredit')[:1]),
            )
        )

        if pk:
            plan = self.retrieve(request, pk).data
            return request.send_data(plan)

        learn_plan_list = self.list(request).data
        return request.send_data(learn_plan_list)

    @has_permission(must_permissions=['lms-study-profession-create'])
    @transaction.atomic
    def post(self, request):
        "  Мэргэжлийн тодорхойлолт шинээр үүсгэх "

        request_data = request.data
        # with_start = '001'
        # profession_code = 1

        # profession_qs = (
        #     ProfessionDefinition.objects.order_by('profession_code')
        # ).first()

        # if profession_qs:
        #     check_code = profession_qs.profession_code
        #     if check_code:
        #         profession_code = int(check_code) + 1

        # new_profession_code = f'{int(profession_code):0{len(with_start)}d}'
        # request_data['profession_code'] = new_profession_code

        # transaction savepoint зарлах нь хэрэв алдаа гарвад roll back хийнэ
        sid = transaction.savepoint()

        try:
            serializer = self.serializer_class(data=request_data, many=False)
            if not serializer.is_valid():
                transaction.savepoint_rollback(sid)
                return request.send_error_valid(serializer.errors)

            serializer.save()

        except Exception:
            return request.send_error("ERR_002")

        return request.send_info("INF_001")

    @has_permission(must_permissions=['lms-study-profession-update'])
    @transaction.atomic
    def put(self, request, pk=None):
        "  Мэргэжлийн тодорхойлолт  засах "

        request_data = request.data
        general_base = request_data.get('general_base')
        profession = request_data.get('id')
        professional_base = request_data.get('professional_base')
        professional_lesson = request_data.get('professional_lesson')
        admission_lesson = request_data.get('admission_lesson')

        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request_data, partial=True)

            if not serializer.is_valid(raise_exception=False):
                return request.send_error_valid(serializer.errors)

            if general_base:
                ss = Profession_SongonKredit.objects.update_or_create(
                    profession_id=profession,
                    lesson_level=LearningPlan.BASIC,
                    defaults={
                        'songon_kredit': general_base
                    }
                )
            if professional_base:
                ss = Profession_SongonKredit.objects.update_or_create(
                    profession_id=profession,
                    lesson_level=LearningPlan.PROF_BASIC,
                    defaults={
                        'songon_kredit': professional_base
                    }
                )
            if professional_lesson:
                ss = Profession_SongonKredit.objects.update_or_create(
                    profession_id=profession,
                    lesson_level=LearningPlan.PROFESSION,
                    defaults={
                        'songon_kredit': professional_lesson
                    }
                )


            self.perform_create(serializer)

        except Exception as e:
            print(e)
            return request.send_error("ERR_002")

        return request.send_info("INF_002")

    @has_permission(must_permissions=['lms-study-profession-delete'])
    def delete(self, request, pk=None):
        " Мэргэжлийн тодорхойлолт устгах "

        data = request.data
        degree = data.get("degree")

        lear_qs = LearningPlan.objects.filter(profession=pk)
        prof = ProfessionalDegree.objects.filter(id=degree)
        group = Group.objects.filter(profession=pk)
        prof_songon = Profession_SongonKredit.objects.filter(profession=pk)

        # Анги бүлэг устгах
        if group:
            return request.send_error("ERR_003", "Анги бүлгийн бүртгэлд холбогдсон учраас устгах боломжгүй.")

        with transaction.atomic():
            # Сургалтын төлөвлөгөө
            if lear_qs:
                lear_qs.delete()

            if prof_songon or  prof:
                prof_songon.delete()

            if prof:
                prof.delete()

        self.destroy(request, pk)

        return request.send_info("INF_003")



@permission_classes([IsAuthenticated])
class ProfessionIntroductionFileAPIView(
    generics.GenericAPIView,
):
    @has_permission(must_permissions=['lms-study-profession-update'])
    def post(self, request):
        """ Мэргэжлийн тодорхойлолтын танилцуулга хэсэгт файл нэмэх """

        data = request.data
        file = data.get('file')

        path = save_file(file, 'profession_images', settings.PROFESSION)

        domain = get_domain_url()

        file_path = os.path.join(settings.MEDIA_URL, path)

        return_url = '{domain}{path}'.format(domain=domain, path=file_path)

        return request.send_data(return_url)

    def put(self, request, pk=None):

        data = request.data
        introduction = data.get('introduction')

        obj = get_object_or_404(ProfessionDefinition, pk=pk)

        with transaction.atomic():
            try:
                obj.introduction = introduction
                obj.save()
            except Exception as e:
                print(e)
                return request.send_error('ERR_002')

        return request.send_info('INF_002')


@permission_classes([IsAuthenticated])
class StudentNoticeFileAPIView(
    generics.GenericAPIView,
    mixins.DestroyModelMixin,
    ):
        queryset = ProfessionDefinition.objects.all()

        def delete(self, request, pk=None):

            # NOTE: Файл засаж байгаа үед хуучин файлыг устгана
            remove_file = os.path.join(settings.PROFESSION, str(pk))
            if remove_file:
                remove_folder(remove_file)

            self.destroy(request, pk)

            return request.send_info("INF_003")
@permission_classes([IsAuthenticated])
class ProfessionDefinitionListAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    """ Мэргэжлийн тодорхойлолтын жагсаалт """

    queryset = ProfessionDefinition.objects
    serializer_class = ProfessionDefinitionSerializer

    def get_queryset(self):
        queryset = self.queryset
        degreeId = self.request.query_params.get('degreeId')
        department = self.request.query_params.get('department')
        confirm_year = self.request.query_params.get('confirm_year')
        schoolId = self.request.query_params.get('schoolId')

        if schoolId:
            queryset = queryset.filter(school_id=schoolId)

        # Хөтөлбөрийн баг
        if department:
            queryset = queryset.filter(department_id=department)

        # Боловсролын зэрэг
        if degreeId:
            queryset = queryset.filter(degree=degreeId)

        # Хөтөлбөр батлагдсан он
        if confirm_year:
            queryset = queryset.filter(confirm_year=confirm_year)

        return queryset

    def get(self, request):

        all_list = self.list(request).data
        return request.send_data(all_list)

@permission_classes([IsAuthenticated])
class LearningPlanAPIView(
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    generics.GenericAPIView
):
    """  Сургалтын төлөвлөгөө """

    queryset = LearningPlan.objects.all()
    serializer_class = LearningPlanSerializer

    pagination_class = CustomPagination
    filter_backends = [SearchFilter]
    search_fields = ['profession__name', 'lesson__name', 'lesson_level__level_name', 'lesson_type__type_name', 'season__season_name']

    @has_permission(must_permissions=['lms-study-learningplan-read'])
    def get(self, request, pk=None):
        "  Сургалтын төлөвлөгөөны жагсаалт "

        self.serializer_class = LearningPlanListSerializer

        if pk:
            plan = self.retrieve(request, pk).data
            return request.send_data(plan)

        learn_plan_list = self.list(request).data
        return request.send_data(learn_plan_list)

    @has_permission(must_permissions=['lms-study-learningplan-create'])
    def post(self, request):
        "  Сургалтын төлөвлөгөө шинээр үүсгэх "

        request_data = request.data
        profession = request_data.get('profession')
        lesson = request_data.get('lesson')

        serializer = self.get_serializer(data=request.data)

        if serializer.is_valid(raise_exception=False):
            is_success = False
            with transaction.atomic():
                try:
                    self.create(request).data

                    is_success = True
                except Exception:
                    raise
            if is_success:
                return request.send_info("INF_001")

            return request.send_error("ERR_002")
        else:
            if profession and lesson:
                qs = self.queryset.filter(profession=profession, lesson=lesson)

                if qs:
                    return request.send_error("ERR_002", "Тухайн мэргэжлийн сургалтын төлөвлөгөөн дээр энэ хичээл бүртгэгдсэн байна.")

            error_obj = {
                "error": serializer.errors,
                "msg": "Код давхцаж байна"
            }

            return request.send_error("ERR_003", error_obj)


    @has_permission(must_permissions=['lms-study-learningplan-update'])
    def put(self, request, pk=None):
        "  Сургалтын төлөвлөгөөн засах "

        request_data = request.data
        profession = request_data.get('profession')
        lesson = request_data.get('lesson')

        check_qs = self.queryset.filter(id=pk).first()
        if check_qs:
            old_profession = check_qs.profession.id
            old_lesson = check_qs.lesson.id

            if old_profession != int(profession) or int(lesson) != old_lesson:
                qs = self.queryset.filter(profession=profession, lesson=lesson)

                if qs:
                    return request.send_error("ERR_002", "Тухайн мэргэжлийн сургалтын төлөвлөгөөн дээр энэ хичээл бүртгэгдсэн байна.")

        self.update(request, pk)
        return request.send_info("INF_002")

    @has_permission(must_permissions=['lms-study-learningplan-delete'])
    def delete(self, request, pk=None):
        " Сургалтын төлөвлөгөө устгах "

        self.destroy(request, pk)
        return request.send_info("INF_003")

@permission_classes([IsAuthenticated])
class LearningPlanListAPIView(
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    generics.GenericAPIView
):
    """  Сургалтын төлөвлөгөө """

    queryset = LearningPlan.objects.all()
    serializer_class = LearningPlanSerializer

    pagination_class = CustomPagination
    filter_backends = [SearchFilter]
    search_fields = ['profession__name', 'lesson__name', 'lesson_level__level_name', 'lesson_type__type_name', 'season__season_name']

    def get(self, request, salbar=None, level=None, type=None, profession=None):
        "  Сургалтын төлөвлөгөөны жагсаалт "

        self.serializer_class = LearningPlanListSerializer

        self.queryset = self.queryset.filter(profession=profession, lesson_level=level, department=salbar, lesson_type=type)

        learn_plan_list = self.list(request).data
        return request.send_data(learn_plan_list)


@permission_classes([IsAuthenticated])
class ConfirmYearListAPIView(
    mixins.ListModelMixin,
    generics.GenericAPIView
):
    """  Хөтөлбөр батлагдсан оны жагсаалт авах """
    queryset = ProfessionDefinition.objects.all()

    def get(self, request):

        conf_year = list(self.queryset.filter(confirm_year__isnull=False).values('confirm_year').annotate(cnt=Count('confirm_year')).order_by('confirm_year'))

        return request.send_data(conf_year)


@permission_classes([IsAuthenticated])
class LessonStandartStudentListAPIView(
    mixins.ListModelMixin,
    generics.GenericAPIView
):
    """ Тухайн оюутны төгсөлтийн хичээлийн жагсаалт """

    queryset = LessonStandart.objects
    serializer_class = LessonStandartSerializer

    def get_queryset(self):
        queryset = self.queryset
        student = self.request.query_params.get('student')

        if student:
            qs_student = Student.objects.filter(pk=student).first()
            if qs_student:
                qs_group = Group.objects.filter(id=qs_student.group.id).last()
                if qs_group:
                    lesson_ids = LearningPlan.objects.filter(profession=qs_group.profession, lesson_level=LearningPlan.DIPLOM).values_list('lesson', flat=True)

                    queryset = queryset.filter(id__in=lesson_ids)

        return queryset

    def get(self, request, pk=None):

        lesson_list = self.list(request).data
        return request.send_data(lesson_list)



@permission_classes([IsAuthenticated])
class ProfessionPlanListAPIView(
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    generics.GenericAPIView
):
    """ Тухайн мэргэжлийн сургалтын төлөвлөгөө """

    queryset = LearningPlan.objects
    serializer_class = LearningPlanSerializer

    def get_queryset(self):
        queryset = self.queryset
        # department = self.request.query_params.get('department')
        school = self.request.query_params.get('school')
        profession = self.request.query_params.get('profession')
        lesson_level = self.request.query_params.get('level')
        lesson_type = self.request.query_params.get('type')

        # if school:
        #     queryset = queryset.filter(school=school)

        # if department:
        #     queryset = queryset.filter(department=department)

        if profession:
            queryset = queryset.filter(profession=profession)

        if lesson_level:
            queryset = queryset.filter(lesson_level=lesson_level)

        if lesson_type:
            queryset = queryset.filter(lesson_type=lesson_type)

        queryset.order_by('updated_at')

        return queryset

    def get(self, request, pk=None):

        self.serializer_class = ProfessionLearningPlanSerializer

        lesson_plan_list = self.list(request).data
        return request.send_data(lesson_plan_list)

    @has_permission(must_permissions=['lms-study-learningplan-create'])
    def post(self, request):
        "  Сургалтын төлөвлөгөө шинээр үүсгэх "

        school = request.data.get('school')
        department = request.data.get('department')
        profession_id = request.data.get('profession')
        all_datas = request.data.get('all_datas')

        # datas жагсаалтаас key-нь датаг салгах нь
        def get_datas_by_key(datas):
            plan_id = datas.get('id')
            lesson = datas.get('lesson')
            previous_lesson = datas.get('previous_lesson')
            season = datas.get('season')
            lesson_level = datas.get('lesson_level')
            lesson_type = datas.get('lesson_type')
            group_lesson = datas.get('group_lesson')
            is_check_score = datas.get('is_check_score')

            return plan_id, lesson, previous_lesson, season, lesson_level, lesson_type, group_lesson, is_check_score

        # list ээс key болон value гаар нь тухайн element байгаа эсэхийг шалгаж element-ийн утгыг авах
        def find_data_from_list_by_key(list, key, value):
            return next(filter(lambda x: x[key] == value, list), None)

        with transaction.atomic():
            try:

                # хэрэв all_datas байхгүй бол pass
                if all_datas:
                    pass

                update_list = []
                create_list = []
                update_list_ids = []

                updated_data_list = []
                create_learninPlan_list = []

                # create хийх update хийх датаг all_datas жагсаалтаас ялгах нь
                for datas in all_datas:
                    check_id = datas.get('id')
                    # тухайн дата нь id гэсэн element байвал update хийх дата байна гэж үзээд update_list жагсаалтад нэмж  datas жагсаалтаас устгана
                    # all_datas жагсаалтанд шинээр үүсгэх датануудын жагсаалт үлдэнэ
                    if check_id:
                        update_list.append(datas)
                        update_list_ids.append(check_id)
                    else:
                        create_list.append(datas)

                # Хэрэв update хийх жагсалтын урт нь 0 ээс их бол update хийх үйлдэлийг эхлүүлнэ
                if len(update_list) > 0:
                    update_qs_list = self.queryset.filter(id__in=update_list_ids)

                    for update_qs in update_qs_list:
                        lesson = None
                        previous_lesson = None
                        season = None
                        group_lesson = None

                        update_data = find_data_from_list_by_key(update_list, 'id', update_qs.id)

                        if update_data:
                            plan_id, lesson, previous_lesson, season, lesson_level, lesson_type, group_lesson, is_check_score = get_datas_by_key(update_data)

                        old_profession = update_qs.profession.id
                        old_lesson = update_qs.lesson.id

                        # хичээлийн бүртгэл шалгах бүртгэлтэй бол алдаа буцаах
                        # Хичээл зассан бол хичээлийн давхцалыг шалгана
                        if old_profession != int(profession_id) or (lesson and int(lesson) != old_lesson):
                            qs = self.queryset.filter(profession=profession_id, lesson=lesson).first()

                            if qs:
                                profession_name = qs.profession.name
                                lesson_name = qs.lesson.name
                                lesson_code = qs.lesson.code
                                msg = "`{profession}` мэргэжлийн сургалтын төлөвлөгөөн дээр `{lesson_code} {lesson_name}` хичээл бүртгэгдсэн байна." \
                                    .format(
                                        profession=profession_name,
                                        lesson_name=lesson_name,
                                        lesson_code=lesson_code,
                                    )

                                return request.send_error("ERR_002", msg)

                        if lesson:
                            # тухайн queryset-ийн датаг өөрчилж updated_data_list жагсаалтанд нэмэх нь
                            update_qs.group_lesson = LessonStandart.objects.get(id=group_lesson) if group_lesson else None
                            update_qs.previous_lesson = LessonStandart.objects.get(id=previous_lesson) if previous_lesson else None
                            update_qs.lesson = LessonStandart.objects.get(id=lesson)
                            update_qs.season = season
                            update_qs.is_check_score = is_check_score

                            updated_data_list.append(update_qs)

                if len(create_list) > 0:
                    for create_data in create_list:
                        plan_id, lesson, previous_lesson, season, lesson_level, lesson_type, group_lesson, is_check_score = get_datas_by_key(create_data)

                        # хичээлийн бүртгэл шалгах бүртгэлтэй бол алдаа буцаах
                        if lesson:
                            qs = self.queryset.filter(profession=profession_id, lesson=lesson).first()

                            if qs:
                                profession_name = qs.profession.name
                                lesson_name = qs.lesson.name
                                lesson_code = qs.lesson.code
                                msg = "`{profession}` мэргэжлийн сургалтын төлөвлөгөөн дээр `{lesson_code} {lesson_name}` хичээл бүртгэгдсэн байна." \
                                    .format(
                                        profession=profession_name,
                                        lesson_name=lesson_name,
                                        lesson_code=lesson_code,
                                    )

                                return request.send_error("ERR_002", msg)

                        if lesson:
                            create_learninPlan_list.append(
                                LearningPlan(
                                    school=SubOrgs.objects.get(id=school),
                                    lesson=LessonStandart.objects.get(id=lesson),
                                    department=Salbars.objects.get(id=department) if department else None,
                                    profession=ProfessionDefinition.objects.get(id=profession_id),
                                    previous_lesson=LessonStandart.objects.get(id=previous_lesson) if previous_lesson else None,
                                    group_lesson = LessonStandart.objects.get(id=group_lesson) if group_lesson else None,
                                    season=season if season else None,
                                    lesson_level=lesson_level,
                                    lesson_type=lesson_type,
                                    is_check_score=is_check_score,
                                )
                            )

                # bulk_create, bulk_update
                self.queryset.bulk_create(create_learninPlan_list)
                self.queryset.bulk_update(updated_data_list, ['group_lesson', 'previous_lesson', 'lesson', 'season', 'is_check_score'])

            except Exception as e:
                print(e.__str__())
                return request.send_error("ERR_002")

        return request.send_info("INF_013")

    @has_permission(must_permissions=['lms-study-learningplan-delete'])
    def delete(self, request, pk=None):
        " Сургалтын төлөвлөгөө устгах "

        self.destroy(request, pk)
        return request.send_info("INF_003")


@permission_classes([IsAuthenticated])
class LessonStandartBagtsAPIView(
    generics.GenericAPIView
):

    def get(self, request, pk=None):
        """ Нэг хичээлийн багц цагийн мэдээлэл авах """

        return_values = []
        if pk:
            values = LessonStandart.objects.filter(id=pk).values('lecture_kr', 'seminar_kr', 'laborator_kr', 'practic_kr', 'biedaalt_kr')

            for value in values:
                obj = {}
                lk = value.get('lecture_kr')
                sem = value.get('seminar_kr')
                lab = value.get('laborator_kr')
                pr = value.get('practic_kr')
                bd = value.get('biedaalt_kr')

                if lk:
                    obj['id'] = TimeTable.LECT
                    obj['name'] = 'Лекц'
                    return_values.append(obj)
                    obj = {}

                if sem:
                    obj['id'] = TimeTable.SEM
                    obj['name'] = 'Семинар'
                    return_values.append(obj)
                    obj = {}

                if lab:
                    obj['id'] = TimeTable.LAB
                    obj['name'] = 'Лаборатор'
                    return_values.append(obj)
                    obj = {}

                if pr:
                    obj['id'] = TimeTable.PRACTIC
                    obj['name'] = 'Практик'
                    return_values.append(obj)
                    obj = {}

                if bd:
                    obj['id'] = TimeTable.BIY_DAALT
                    obj['name'] = 'Бие даалт'
                    return_values.append(obj)
                    obj = {}

        return request.send_data(list(return_values))


@permission_classes([IsAuthenticated])
class LearningPlanProfessionDefinitionAPIView(
    generics.GenericAPIView
):
    """ Мэргэжлээс сургалтын төлөвлөгөөний хичээл авах """

    def get(self, request, profession=None):

        all_list = []
        queryset = LearningPlan.objects
        school = request.query_params.get('school')
        profession = request.query_params.get('profession')

        if school:
            queryset = queryset.filter(school=school)

        if profession:
            get_object_or_404(ProfessionDefinition, id=profession)

            lesson_ids = queryset.filter(profession=profession).values_list('lesson', flat=True).distinct('lesson')

            all_list = list(LessonStandart.objects.filter(id__in=lesson_ids).values('id', 'name', 'code', 'kredit'))

            seasons = queryset.filter(lesson__in=lesson_ids, profession=profession).values('lesson', 'season')
            for clist in all_list:

                full_name = ''
                name = clist.get('name')
                code = clist.get('code')
                p_id = clist.get('id')

                for lseason in seasons:
                    if p_id == lseason.get('lesson'):
                        season = lseason.get('season')

                clist['season'] = season if season else ''

                if code:
                    full_name += code
                if name:
                    full_name +=  ' ' + name

                clist['full_name'] = full_name

        return request.send_data(all_list)


@permission_classes([IsAuthenticated])
class ProfessionPrintPlanAPIView(
    generics.GenericAPIView
):
    """ Мэргэжлээс сургалтын төлөвлөгөөний хичээл авах """

    def get(self, request):

        all_data = dict()

        profession = request.query_params.get('profession')
        group = request.query_params.get('group')
        student = request.query_params.get('student')
        season = request.query_params.get('season')

        profession_qs = ProfessionDefinition.objects.filter(id=profession).values('id', 'dep_name', 'name').last()
        group_queryset = Group.objects.filter(profession=profession, is_finish=False).order_by('id')

        # анги байгаа үед л ангиар шүүх
        if group:
            group_queryset = group_queryset.filter(id=group)

        group_data = GroupSerializer(group_queryset, many=True).data

        # student байгаа эсэх шалгах
        if student:
            student_qs = Student.objects.filter(id=student).values('id', 'code', 'first_name', 'last_name', 'register_num').first()

        all_data['group'] = group_data

        all_data['id'] = profession_qs['id']
        all_data['dep_name'] = profession_qs['dep_name']
        all_data['name'] = profession_qs['name']

        # student байгаа үед info авах
        if student:
            all_data['student_id'] = student_qs['id']
            all_data['code'] = student_qs['code']
            all_data['first_name'] = student_qs['first_name']
            all_data['last_name'] = student_qs['last_name']
            all_data['register_num'] = student_qs['register_num']

        request.data['group_queryset'] = group_queryset

        learning_plan_queryset = LearningPlan.objects.filter(profession=profession)

        # хичээлийн улиралаар хайх
        if season:
            # улирал тэгш сондгой эсэхийг шалгах
            year_diff = int(season) % 2

            # тэгш эсвэл сондгой улиралаар хайх
            odd_or_even_seasons = [f"[{season}]" for season in range(1, 9) if season % 2 == year_diff % 2]
            learning_plan_queryset = learning_plan_queryset.filter(season__in=odd_or_even_seasons)

        all_levels_data = list()
        total_studied_credits = 0

        for lesson_level in LearningPlan.LESSON_LEVEL:
            level_data = learning_plan_queryset.filter(lesson_level=lesson_level[0])

            if level_data:
                one_lesson_datas = dict()
                one_lesson_datas['level'] = lesson_level[1]
                one_lesson_datas['count'] = level_data.aggregate(Sum('lesson__kredit')).get('lesson__kredit__sum')
                one_lesson_datas['total_studied_credits'] = 0
                one_type_datas = list()

                for lesson_type in LearningPlan.LESSON_TYPE:

                    type_datas = level_data.filter(lesson_type=lesson_type[0])

                    if type_datas:
                        lesson_datas = dict()
                        type_data = LearningPlanPrintSerializer(type_datas, context={ "request": request }, many=True).data
                        lesson_datas['type'] = lesson_type[1]
                        lesson_datas['lessons'] = type_data
                        lesson_datas['count'] = type_datas.aggregate(Sum('lesson__kredit')).get('lesson__kredit__sum')
                        one_type_datas.append(lesson_datas)

                        # lesson_type болгонд хэдэн хичээл үзсэн тоолох
                        for lesson in type_data:
                            if 'student_study' in lesson['lesson']:
                                for study in lesson['lesson']['student_study']:
                                    # хичээлийг үзсэн бол кредитийг тоолох
                                    if study['value'] == 2:
                                        one_lesson_datas['total_studied_credits'] += lesson['lesson']['kredit']

                one_lesson_datas['data'] = one_type_datas

                all_levels_data.append(one_lesson_datas)
                total_studied_credits += one_lesson_datas['total_studied_credits']

        all_data['lesson'] = all_levels_data
        all_data['total_studied_credits'] = total_studied_credits

        return request.send_data(all_data)

@permission_classes([IsAuthenticated])
class AdmissionBottomScoreAPIView(
    mixins.CreateModelMixin,
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    generics.GenericAPIView
):

    """ Элсэлтийн шалгалтын хичээл ба босго оноо мэргэжлээр """

    queryset = AdmissionBottomScore.objects
    serializer_class = AdmissionBottomScoreSerializer

    def get(self, request, pk=None):

        self.serializer_class = AdmissionBottomScoreListSerializer

        #Өнөөдрөөс хойших датаг авна
        self.queryset = self.queryset.filter(profession=pk)

        return_datas = self.list(request).data

        return request.send_data(return_datas)


    # @has_permission(must_permissions=['lms-role-teacher-score-update'])
    def put(self, request, pk=None):
        """ ЭЕШ оноо засах
            мэргэжлийн id = pk
        """
        datas = request.data
        self.create(request, datas)

        return request.send_info("INF_002")

    def delete(self, request, pk=None):

        try:
            obj = AdmissionBottomScore.objects.filter(id=pk)
            if obj:
                obj.delete()
        except:
            return request.send_error("ERR_002", "Амжилтгүй")

        return request.send_info("INF_003")

@permission_classes([IsAuthenticated])
class LessonStandartTimetableListAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    """ Хичээлийн хуваарьт зориулж хичээлийн лист авах """

    queryset = LessonStandart.objects.all()
    serializer_class = LessonStandartSerialzier

    def get(self, request):

        year = self.request.query_params.get('lesson_year')
        season = self.request.query_params.get('lesson_season')
        school = self.request.query_params.get('school')

        qs_lplan = LearningPlan.objects.all()

        # if school:
        #     # qs_lplan = qs_lplan.filter(school=school)
        #     self.queryset = self.queryset.filter(Q(Q(school=school) | Q(school__org_code=10)))

        even_i = []
        odd_i = []

        for i in range(1,13):
            if i % 2 == 0:
                even_i.append(i)
            else:
                odd_i.append(i)


        # Идэвхтэй улиралд байгаа хичээлүүдийг авах
        # if season == '1':
        #     lookups = [Q(season__contains=str(value)) for value in odd_i]

        #     # Combine the individual lookups using the OR operator (|) for a match on any of the values
        #     filter_query = Q()
        #     for lookup in lookups:
        #         filter_query |= lookup

        #     lesson_ids = qs_lplan.filter(filter_query).values_list('lesson', flat=True).distinct('lesson')
        # else:
        #     lookups = [Q(season__contains=str(value)) for value in even_i]

        #     # Combine the individual lookups using the OR operator (|) for a match on any of the values
        #     filter_query = Q()
        #     for lookup in lookups:
        #         filter_query |= lookup

        #     lesson_ids = qs_lplan.filter(filter_query).values_list('lesson', flat=True).distinct('lesson')

        # self.queryset = self.queryset.filter(id__in=list(lesson_ids))

        all_list = self.list(request).data

        return request.send_data(list(all_list))


@permission_classes([IsAuthenticated])
class LessonStandartDiplomaListAPIView(
    mixins.ListModelMixin,
    generics.GenericAPIView
):
    """ Тухайн оюутны төгсөлтийн хичээлийн жагсаалт """

    queryset = LessonStandart.objects
    serializer_class = LessonStandartSerializer

    def get_queryset(self):
        queryset = self.queryset
        student = self.request.query_params.get('student')

        if student:
            qs_student = Student.objects.filter(pk=student).first()
            if qs_student:
                qs_group = Group.objects.filter(pk=qs_student.group.id).first()
                if qs_group:
                    lesson_ids = LearningPlan.objects.filter(Q(Q(profession=qs_group.profession) & Q(Q(lesson_level=LearningPlan.DIPLOM) | Q(lesson_level=LearningPlan.MAG_DIPLOM)))).values_list('lesson', flat=True)
                    queryset = queryset.filter(id__in=lesson_ids)

        return queryset

    def get(self, request, pk=None):

        lesson_list = self.list(request).data
        return request.send_data(lesson_list)


@permission_classes([IsAuthenticated])
class LessonStandartProfessionListAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    ''' Нэг мэггэжлийн сургалтын төлөвлөгөөнд байгаа хичээлүүд '''

    queryset = LessonStandart.objects.all()
    serializer_class = LessonStandartSerialzier

    def get(self, request, profession=None):
        all_list = []

        if profession:
            learningplan_lesson_ids = LearningPlan.objects.filter(profession=profession).values_list('lesson', flat=True)

            self.queryset = self.queryset.filter(id__in=list(learningplan_lesson_ids))

            all_list = self.list(request).data

        return request.send_data(all_list)

@permission_classes([IsAuthenticated])
class ChallengeAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin
):
    """ Өөрийгөө сорих тест """

    queryset = Challenge.objects.all().order_by("-created_at")
    serializer_class = ChallengeSerializer

    pagination_class = CustomPagination
    filter_backends = [SearchFilter]
    search_fields = ['lesson__name', 'lesson__code', 'title']

    @has_permission(must_permissions=['lms-exam-read'])
    def get(self, request):

        self.serializer_class = ChallengeListSerializer
        lesson = request.query_params.get('lesson')
        time_type = request.query_params.get('type')
        teacher_id = request.query_params.get('teacher')
        school = request.query_params.get('school')
        lesson_year = request.query_params.get('lesson_year')
        lesson_season = request.query_params.get('lesson_season')
        challenge_type = request.query_params.get('challenge_type')

        if lesson_year:
            self.queryset = self.queryset.filter(lesson_year=lesson_year)

        if lesson_season:
            self.queryset = self.queryset.filter(lesson_season=lesson_season)

        if school:
            self.queryset = self.queryset.filter(created_by__sub_org=school)

        if teacher_id:
            self.queryset = self.queryset.filter(created_by=teacher_id)

        if lesson:
            self.queryset = self.queryset.filter(lesson=lesson)

        if time_type:
            state_filters = Challenge.get_state_filter(time_type)

            self.queryset = self.queryset.filter(**state_filters)

        if challenge_type:
            self.queryset = self.queryset.filter(challenge_type=challenge_type)

        datas = self.list(request).data

        return request.send_data(datas)

    @has_permission(must_permissions=['lms-exam-create'])
    def post(self, request):

        lesson_year = request.query_params.get('year')
        lesson_season = request.query_params.get('season')

        general_datas = request.data

        question_ids = general_datas.get('question_ids')

        qs_student = Student.objects.all()

        user = request.user
        teacher = Teachers.objects.filter(user_id=user).first()

        # Анги эсвэл оюутны select хийгдсэн ids
        selected_ids = general_datas.get('selected')

        # Хамрах хүрээ
        scope = general_datas.get('scope')
        lesson_id = general_datas.get('lesson')

        lesson = LessonStandart.objects.get(id=lesson_id)

        general_datas = remove_key_from_dict(general_datas, [ 'scope', 'scopeName', 'selected', 'select', 'lesson'])

        # Өмнөх шалгалтын мэдээллээс сонгож хадгалсан үед дараах serializer-аар  дамжиж ирсэн утгуудыг хасна
        if general_datas.get('id'):
            general_datas = remove_key_from_dict(general_datas, ['group', 'student', 'startAt', 'endAt', 'id'])

        if 'question_ids' in general_datas:
            del general_datas['question_ids']

        if 'questions' in general_datas:
            del general_datas['questions']

        if 'scopeName' in general_datas:
            del general_datas['scopeName']

        general_datas['lesson'] = lesson
        general_datas['deleted_by'] = teacher
        general_datas['created_by'] = teacher

        # Тухайн хичээлийн хуваарь
        timetable_qs = TimeTable.objects.filter(lesson_year=lesson_year, lesson_season=lesson_season, lesson=lesson_id, teacher=teacher)
        timetable_ids = timetable_qs.values_list('id', flat=True)

        # Хичээлийн хуваариас хасалт хийлгэсэн оюутнууд
        exclude_student_ids = TimeTable_to_student.objects.filter(timetable_id__in=timetable_ids, add_flag=False).values_list('student', flat=True)

        kind = Challenge.KIND_LESSON

        # Хамрах хүрээ
        if scope == 'all':
            all_lesson_students = get_lesson_choice_student(lesson_id, teacher.id, '', lesson_year, lesson_season)
            student_ids = qs_student.filter(id__in=all_lesson_students).values_list('id', flat=True)

        elif scope == 'group':
            kind = Challenge.KIND_GROUP

            all_student_ids = qs_student.filter(group_id__in=selected_ids).values_list('id', flat=True)

            student_ids = set(all_student_ids) - set(list(exclude_student_ids))

        elif scope == 'student':
            kind = Challenge.KIND_STUDENT
            student_ids = selected_ids

        general_datas['kind'] = kind

        general_datas = null_to_none(general_datas)

        with transaction.atomic():

            try:
                question_qs = ChallengeQuestions.objects.filter(id__in=question_ids)

                challenge = self.queryset.create(**general_datas)

                # ManytoMany field нэмэх хэсэг
                if challenge:
                    challenge.questions.set(list(question_qs))
                    challenge.student.set(list(student_ids))

            except Exception as e:
                print(e)

                return request.send_error('ERR_002')

        return request.send_info('INF_001')

    @has_permission(must_permissions=['lms-exam-update'])
    def put(self, request, pk):

        general_datas = request.data

        lesson_year = request.query_params.get('year')
        lesson_season = request.query_params.get('season')

        question_ids = general_datas.get('question_ids')

        qs_student = Student.objects.all()

        user = request.user
        teacher = Teachers.objects.filter(user_id=user).first()

        # Анги эсвэл оюутны select хийгдсэн ids
        selected_ids = general_datas.get('selected')

        # Хамрах хүрээ
        scope = general_datas.get('scope')
        lesson_id = general_datas.get('lesson')

        general_datas = remove_key_from_dict(general_datas, ['group', 'student', 'scopeName', 'startAt', 'endAt', 'question_ids', 'questions', 'selected', 'created_by'])

        # Тухайн хичээлийн хуваарь
        timetable_qs = TimeTable.objects.filter(lesson_year=lesson_year, lesson_season=lesson_season, lesson=lesson_id, teacher=teacher)
        timetable_ids = timetable_qs.values_list('id', flat=True)

        # Хичээлийн хуваариас хасалт хийлгэсэн оюутнууд
        exclude_student_ids = TimeTable_to_student.objects.filter(timetable_id__in=timetable_ids, add_flag=False).values_list('student', flat=True)

        kind = Challenge.KIND_LESSON

        # Хамрах хүрээ
        if scope == 'all':
            all_lesson_students = get_lesson_choice_student(lesson_id, teacher.id, '',  lesson_year, lesson_season)
            student_ids = qs_student.filter(id__in=all_lesson_students).values_list('id', flat=True)

        elif scope == 'group':
            kind = Challenge.KIND_GROUP

            all_student_ids = qs_student.filter(group_id__in=selected_ids).values_list('id', flat=True)

            student_ids = set(all_student_ids) - set(list(exclude_student_ids))

        elif scope == 'student':
            kind = Challenge.KIND_STUDENT
            student_ids = selected_ids

        general_datas['kind'] = kind
        general_datas['created_by'] = teacher.id

        with transaction.atomic():

            try:
                self.update(request, pk).data

                challenge = self.queryset.get(id=pk)

                # ManytoMany field нэмэх хэсэг
                if challenge:

                    question_qs = ChallengeQuestions.objects.filter(id__in=question_ids)

                    # ManytoMany field ээ устгаад нэмнэ.
                    challenge.questions.clear()
                    challenge.student.clear()

                    # ManytoMany field ээ шинээр нэмнэ.
                    challenge.questions.set(list(question_qs))
                    challenge.student.set(list(student_ids))

            except Exception as e:
                print(e)

                return request.send_error('ERR_002')

        return request.send_info('INF_002')

    @has_permission(must_permissions=['lms-exam-delete'])
    def delete(self, request, pk=None):

        challenge_obj = self.queryset.get(id=pk)

        try:
            if challenge_obj:
               challenge_obj.questions.clear()
               challenge_obj.student.clear()
               challenge_obj.save()

            # Шалгалтад хариулсан сурагчид
            challenge_students = ChallengeStudents.objects.filter(challenge=challenge_obj)
            if challenge_students:
                challenge_students.delete()

            self.destroy(request, pk)

        except Exception as e:
            print(e)
            return request.send_error("ERR_002")

        return request.send_info("INF_003")

@permission_classes([IsAuthenticated])
class ChallengeAllAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin
):
    """ Өөрийгөө сорих шалгалт бүх жагсаалт """

    queryset = Challenge.objects.all()
    serializer_class = ChallengeListSerializer

    def get(self, request):

        challenge = request.query_params.get('challenge')

        if challenge:
            instance = self.queryset.filter(id=int(challenge)).first()
            serializer_data = self.get_serializer(instance).data

            return request.send_data(serializer_data)

        datas = self.list(request).data

        return request.send_data(datas)

@permission_classes([IsAuthenticated])
class ChallengeSelectAPIView(
    generics.GenericAPIView
):
    def get(self, request):

        lesson = ''
        all_list = []

        year = request.query_params.get('year')
        season = request.query_params.get('season')
        lesson = request.query_params.get('lesson')

        ctype = request.query_params.get('type')

        user = request.user
        teacher = get_object_or_404(Teachers, user_id=user, action_status=Teachers.APPROVED)

        quesyset = TimeTable.objects.all()

        if year and season:
            quesyset = quesyset.filter(lesson_year=year, lesson_season=season)

        if lesson:
            quesyset = quesyset.filter(lesson=lesson)

        quesyset = quesyset.filter(teacher=teacher)

        timetable_ids = quesyset.values_list('id', flat=True)

        if ctype == 'group':

            group_ids = TimeTable_to_group.objects.filter(timetable_id__in=timetable_ids).values_list('group', flat=True)

            qroup_qs = Group.objects.filter(id__in=group_ids)

            all_list = GroupListSerializer(qroup_qs, many=True).data

        else:

            all_student = get_lesson_choice_student(lesson, teacher, '', year, season)

            student_list = Student.objects.filter(id__in=all_student).values_list('id', flat=True)

            for student_id in student_list:
                obj = {}
                obj['id'] = student_id

                student = Student.objects.filter(id=student_id).first()

                obj['full_name'] = student.full_name + " " + student.code

                all_list.append(obj)

        return request.send_data(list(all_list))


@permission_classes([IsAuthenticated])
class QuestionsAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    mixins.DestroyModelMixin
):
    """
        Асуултын хэсэг
    """

    queryset = ChallengeQuestions.objects.all().order_by('-created_at')
    serializer_class = ChallengeQuestionListSerializer

    pagination_class = CustomPagination

    filter_backends = [SearchFilter]
    search_fields = ['question', 'subject__title']


    @has_permission(must_permissions=['lms-exam-question-read'])
    def get(self, request, pk=None):

        lesson = request.query_params.get('lesson')
        subject = request.query_params.get('subject')

        user = request.user
        teacher = get_object_or_404(Teachers, user_id=user, action_status=Teachers.APPROVED)

        self.queryset = self.queryset.filter(created_by=teacher)

        if lesson:
            self.queryset = self.queryset.filter(subject__lesson=lesson)

        if subject:
            self.queryset = self.queryset.filter(subject=subject)

        if pk:
            row_data = self.retrieve(request, pk).data
            return request.send_data(row_data)

        all_list = self.list(request).data

        return request.send_data(all_list)


    # @has_permission(must_permissions=['lms-exam-question-update'])
    # def put(self, request, pk):

    #     datas = request.data.dict()
    #     subject_id = datas.get('subject')

    #     subject = Lesson_title_plan.objects.filter(id=subject_id).first()

    #     quesion_imgs = request.FILES.getlist('questionImg')
    #     choice_imgs = request.FILES.getlist('choiceImg')

    #     questions = request.POST.getlist('question')

    #     user = request.user
    #     teacher = Teachers.objects.filter(user_id=user).first()

    #     with transaction.atomic():
    #         sid = transaction.savepoint()
    #         try:
    #             # Асуултыг хадгалах хэсэг
    #             for question in questions:
    #                 question = json_load(question)

    #                 qkind = question.get("kind")
    #                 image_name = question.get('imageName')

    #                 score = question.get('score') # Асуултын оноо

    #                 question['created_by'] = teacher
    #                 question['subject'] = subject

    #                 question_img = None

    #                 # Асуултын сонголтууд
    #                 choices = question.get('choices')

    #                 # Асуултын зураг хадгалах хэсэг
    #                 for img in quesion_imgs:
    #                     if image_name == img.name:
    #                         question_img = img
    #                         break

    #                 question = remove_key_from_dict(question, [ 'image', 'choices'])

    #                 if 'imageName' in question:
    #                     del question['imageName']

    #                 if 'imageUrl' in question:
    #                     del question['imageUrl']

    #                 if 'kind_name' in question:
    #                     del question['kind_name']

    #                 if not question.get('max_choice_count'):
    #                     question['max_choice_count'] = 0

    #                 if not question.get('rating_max_count'):
    #                     question['rating_max_count'] = 0
    #                 else:
    #                     question['rating_max_count'] = 5

    #                 question = null_to_none(question)

    #                 question_obj, created = ChallengeQuestions.objects.update_or_create(
    #                     id=pk,
    #                     defaults={
    #                         **question
    #                     }
    #                 )

    #                 # Асуултанд зураг байвал хадгалах хэсэг
    #                 if question_img:
    #                     question_img_path = get_image_path(question_obj)

    #                     file_path = save_file(question_img, question_img_path)[0]

    #                     question_obj.image = file_path
    #                     question_obj.save()
    #                 else:
    #                     old_image = question_obj.image

    #                     # Хуучин зураг засах үедээ устгасан бол файл устгана.
    #                     if old_image and not image_name:
    #                         remove_folder(str(old_image))

    #                         question_obj.image = None
    #                         question_obj.save()

    #                 choice_ids = list()

    #                 # Асуултын сонголтуудыг үүсгэх нь
    #                 if int(qkind) in [ChallengeQuestions.KIND_MULTI_CHOICE, ChallengeQuestions.KIND_ONE_CHOICE]:

    #                     # Олон сонголттой үед асуултын оноог хувааж тавина
    #                     if int(qkind) == ChallengeQuestions.KIND_MULTI_CHOICE:
    #                         max_choice_count = int(question.get('max_choice_count'))

    #                         score = float(score) / max_choice_count

    #                     for choice in choices:
    #                         choice['created_by'] = teacher
    #                         checked = choice.get('checked')

    #                         choice['score'] = score if checked else 0

    #                         img_name = choice.get('imageName')

    #                         choice_img = None

    #                         # Хариултын зураг хадгалах хэсэг
    #                         for cimg in choice_imgs:
    #                             if img_name == cimg.name:
    #                                 choice_img = cimg
    #                                 break

    #                         choice = remove_key_from_dict(choice, ['image', 'checked'])

    #                         if 'imageName' in choice:
    #                             del choice['imageName']

    #                         if 'imageUrl' in choice:
    #                             del choice['imageUrl']

    #                         choice_obj, created = QuestionChoices.objects.update_or_create(
    #                             id=choice.get('id'),
    #                             defaults={
    #                                 **choice
    #                             }
    #                         )

    #                         # Асуултанд зураг байвал хадгалах хэсэг
    #                         if choice_img:
    #                             choice_img_path = get_choice_image_path(choice_obj)

    #                             file_path = save_file(choice_img, choice_img_path)[0]

    #                             choice_obj.image = file_path
    #                             choice_obj.save()
    #                         else:
    #                             choice_old_image = choice_obj.image

    #                             # Хуучин зураг засах үедээ устгасан бол файл устгана.
    #                             if choice_old_image and not img_name:
    #                                 remove_folder(str(choice_old_image))

    #                                 choice_obj.image = None
    #                                 choice_obj.save()

    #                         choice_ids.append(choice_obj.id)

    #                 question_obj.choices.set(choice_ids)

    #         except Exception as e:
    #             print(e)
    #             transaction.savepoint_rollback(sid)

    #             return request.send_error('ERR_002')

    #         return request.send_info('INF_002')
    def put(self, request, pk):

        request_data = request.data.dict()
        type = request.query_params.get('type')

        if type == "question":
            question_img = request_data['image']
            request_data = remove_key_from_dict(request_data, [ 'image'])
            with transaction.atomic():
                question_obj = ChallengeQuestions.objects.filter(id=pk).first()

                if isinstance(question_img, str) != True:
                    question_img_path = get_image_path(question_obj)

                    file_path = save_file(question_img, question_img_path)[0]

                    old_image = question_obj.image
                    question_obj.image = file_path
                    question_obj.save()
                    if old_image:
                        remove_folder(str(old_image))


                if isinstance(question_img, str) == True and question_img == '':
                    old_image = question_obj.image
                    question_img_path = get_image_path(question_obj)
                    # Хуучин зураг засах үедээ устгасан бол файл устгана.
                    question_obj.image = None
                    question_obj.save()
                    if old_image:
                        remove_folder(str(old_image))

                updated_question_rows = ChallengeQuestions.objects.filter(id=pk).update(
                    **request_data
                )
                data = None
                if updated_question_rows > 0:
                    updated_question = ChallengeQuestions.objects.filter(id=pk).first()
                    ser = dynamic_serializer(ChallengeQuestions, "__all__", 1)
                    data = ser(updated_question).data
                return request.send_info('INF_002', data)

        else:
            answer_img = request_data["image"]
            answer_id = request_data.get('id')
            request_data = remove_key_from_dict(request_data, ['image', 'id'])
            with transaction.atomic():
                answer_obj = QuestionChoices.objects.filter(id=answer_id).first()
                question_obj = ChallengeQuestions.objects.filter(id=pk).first()
                if isinstance(answer_img, str) != True:
                    answer_img_path = get_choice_image_path(answer_obj)
                    file_path = save_file(answer_img, answer_img_path)[0]
                    old_image = answer_obj.image
                    answer_obj.image = file_path
                    answer_obj.save()
                    if old_image:
                        remove_folder(str(old_image))

                # Delete image
                if isinstance(answer_img, str) == True and answer_img == '':
                    old_image = answer_obj.image
                    answer_img_path = get_choice_image_path(answer_obj)
                    answer_obj.image = None
                    answer_obj.save()
                    if old_image:
                        remove_folder(str(old_image))

                updated_rows = QuestionChoices.objects.filter(id=answer_id).update(**request_data)
                data = None
                if updated_rows > 0:
                    updated_answer = QuestionChoices.objects.filter(id=answer_id).first()
                    ser = dynamic_serializer(QuestionChoices, "__all__")
                    data = ser(updated_answer).data
                    print(data)
                return request.send_info('INF_002', data)

    @has_permission(must_permissions=['lms-exam-question-create'])
    def post(self, request):

        datas = request.data.dict()
        subject_id = datas.get('subject')

        subject = Lesson_title_plan.objects.filter(id=subject_id).first()

        quesion_imgs = request.FILES.getlist('questionImg')
        choice_imgs = request.FILES.getlist('choiceImg')

        questions = request.POST.getlist('question')

        user = request.user
        teacher = Teachers.objects.filter(user_id=user).first()

        with transaction.atomic():
            sid = transaction.savepoint()
            try:
                # Асуултыг хадгалах хэсэг
                for question in questions:
                    question = json_load(question)

                    qkind = question.get("kind")
                    image_name = question.get('imageName')
                    yes_or_no = question.get('yes_or_no')

                    score = question.get('score') # Асуултын оноо

                    question['created_by'] = teacher
                    question['subject'] = subject

                    if not yes_or_no:
                        question['yes_or_no'] = None

                    question_img = None

                    # Асуултын сонголтууд
                    choices = question.get('choices')

                    # Асуултын зураг хадгалах хэсэг
                    for img in quesion_imgs:
                        if image_name == img.name:
                            question_img = img
                            break

                    question = remove_key_from_dict(question, [ 'image', 'choices'])

                    if 'imageName' in question:
                        del question['imageName']

                    if 'imageUrl' in question:
                        del question['imageUrl']


                    if not question.get('max_choice_count'):
                        question['max_choice_count'] = 0

                    if not question.get('rating_max_count'):
                        question['rating_max_count'] = 0
                    else:
                        question['rating_max_count'] = 5

                    question = null_to_none(question)

                    question_obj = ChallengeQuestions.objects.create(
                        **question
                    )

                    # Асуултанд зураг байвал хадгалах хэсэг
                    if question_img:
                        question_img_path = get_image_path(question_obj)

                        file_path = save_file(question_img, question_img_path)[0]

                        question_obj.image = file_path
                        question_obj.save()

                    choice_ids = list()

                    # Асуултын сонголтуудыг үүсгэх нь
                    if int(qkind) in [ChallengeQuestions.KIND_MULTI_CHOICE, ChallengeQuestions.KIND_ONE_CHOICE]:

                        # Олон сонголттой үед асуултын оноог хувааж тавина
                        if int(qkind) == ChallengeQuestions.KIND_MULTI_CHOICE:
                            max_choice_count = int(question.get('max_choice_count'))

                            score = float(score) / max_choice_count

                        for choice in choices:

                            choice['created_by'] = teacher

                            checked = choice.get('checked')

                            choice['score'] = score if checked else 0

                            img_name = choice.get('imageName')

                            choice_img = None

                            # Хариултын зураг хадгалах хэсэг
                            for cimg in choice_imgs:
                                if img_name == cimg.name:
                                    choice_img = cimg
                                    break

                            choice = remove_key_from_dict(choice, ['image', 'checked'])

                            if 'imageName' in choice:
                                del choice['imageName']

                            if 'imageUrl' in choice:
                                del choice['imageUrl']

                            choice_obj = QuestionChoices.objects.create(
                                **choice
                            )

                            # Асуултанд зураг байвал хадгалах хэсэг
                            if choice_img:
                                choice_img_path = get_choice_image_path(choice_obj)

                                file_path = save_file(choice_img, choice_img_path)[0]

                                choice_obj.image = file_path
                                choice_obj.save()

                            choice_ids.append(choice_obj.id)

                    question_obj.choices.set(choice_ids)

            except Exception as e:
                print(e)
                transaction.savepoint_rollback(sid)

                return request.send_error('ERR_002')

            return request.send_info('INF_001')

    @has_permission(must_permissions=['lms-exam-question-delete'])
    def delete(self, request):

        delete_ids = request.query_params.getlist('delete')

        questions = self.queryset.filter(id__in=delete_ids)

        with transaction.atomic():
            try:
                for question in questions:

                    # Хэрвээ асуултанд зураг байвал устгана
                    question_img = question.image

                    if question_img:
                        remove_folder(str(question_img))

                        if question:
                            choices = question.choices.all()

                            for choice in choices:
                                img = choice.image

                                # Хэрвээ хариултын хэсэгт зураг байвал устгана
                                if img:
                                    remove_folder(str(img))

                                choice.delete()

                            # Хариултын хэсгийг устгах хэсэг
                            question.choices.clear()

                    question.delete()

            except Exception as e:
                print(e)
                return request.send_error("ERR_002")

        return request.send_info("INF_003")


class PsychologicalTestQuestionsAPIView(
    generics.GenericAPIView,
    APIView,
    mixins.ListModelMixin,
    mixins.DestroyModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin
):
    queryset = PsychologicalTestQuestions.objects.all().order_by('id')
    serializer_class = dynamic_serializer(PsychologicalTestQuestions, "__all__")

    def get(self, request):
        datas = self.list(request).data
        return request.send_data(datas)

    def post(self, request):
        questions = request.POST.getlist('questions')
        files = request.FILES.getlist('files')

        user = request.user

        with transaction.atomic():
            try:
                for question in questions:
                    question = json.loads(question)
                    qkind = question.get("kind")
                    question_img = question.get('image')
                    score = question.get('score')
                    answers = question.get('answers')
                    question['created_by'] = user
                    question['yes_or_no'] = None
                    question['max_choice_count'] = 0
                    question['rating_max_count'] = score

                    if 'score' in question:
                        score = question.get('score')
                        question['score'] = int(score) if score else None

                    for image in files:
                        if hasattr(image, "name") and question['image'] == image.name:
                            question_img = image

                    for ans in answers:
                        if ans['is_correct'] and ans['value'] == 'Тийм':
                            question['yes_or_no'] = True
                        if ans['is_correct'] and ans['value'] == 'Үгүй':
                            question['yes_or_no'] = False
                        if ans['is_correct']:
                            question['max_choice_count'] += 1

                    choices = answers
                    question['has_score'] = question['level'] == 1

                    keys_to_remove = ['image', 'answers', 'level']
                    question = remove_key_from_dict(question, keys_to_remove)
                    question = null_to_none(question)

                    question_obj = PsychologicalTestQuestions.objects.create(**question)

                    if question_img:
                        question_img_path = get_image_path(question_obj)
                        file_path = save_file(question_img, question_img_path)
                        question_obj.image = file_path
                        question_obj.save()

                    choice_ids = []

                    if int(qkind) in [PsychologicalTestQuestions.KIND_MULTI_CHOICE, PsychologicalTestQuestions.KIND_ONE_CHOICE, PsychologicalTestQuestions.KIND_BOOLEAN]:
                        if int(qkind) == PsychologicalTestQuestions.KIND_MULTI_CHOICE and score:
                            max_choice_count = int(question.get('max_choice_count'))
                            score = float(score) / max_choice_count

                        for choice in choices:
                            choice['created_by'] = user

                            choice_img = None
                            for image in files:
                                if hasattr(image, "name") and choice['image'] == image.name:
                                    choice_img = image

                            choice = remove_key_from_dict(choice, ['image'])
                            choice_obj = PsychologicalQuestionChoices.objects.create(**choice)

                            if choice_img:
                                choice_img_path = get_choice_image_path(choice_obj)
                                file_path = save_file(choice_img, choice_img_path)
                                choice_obj.image = file_path
                                choice_obj.save()

                            choice_ids.append(choice_obj.id)

                    question_obj.choices.set(choice_ids)

            except Exception as e:
                print(e)
                return request.send_error('ERR_002')

        return request.send_info('INF_001')

    def put(self, request, pk):

        request_data = request.data.dict()
        type = request.query_params.get('type')

        # Асуултийг edit хийх үед
        if type == "question":
            question_img = request_data['image']

            answers = request_data['answer']

            # Олон хариу ирвэл салгаж авна
            answers_list = answers.split(',') if answers else []
            request_data = remove_key_from_dict(request_data, ['image', 'answer'])

            with transaction.atomic():
                question_obj = PsychologicalTestQuestions.objects.filter(id=pk).first()

                # Хариулт нь хоосон биш мөн оноотой үед ажилна
                if answers_list and request_data['score'] != 'null' and answers != 'undefined':

                    # Зөв хариултуудыг авна
                    correct_choice_ids = set()

                    # Орж ирсэн хариулт болгоныг авч өөрчлөн
                    for answer in answers_list:
                        answer_object = PsychologicalQuestionChoices.objects.filter(id=answer).first()

                        if answer_object:
                            answer_object.is_correct = True
                            answer_object.save()
                            correct_choice_ids.add(answer_object.id)

                    # Үлдсэн хариуг худал болгох
                    question_obj.choices.exclude(id__in=correct_choice_ids).update(is_correct=False)

                # Хэрэв оноо орж ирсэн тохиолдолд тухайн object-ийн has_score field-ийг true болгоно
                if 'score' in request_data:
                    score = request_data.get('score')
                    if score and score != 'null':
                        request_data['score'] = int(score)
                        request_data['has_score'] = True

                    # Хэрэв оноогүй тохиолдолд тухайн object-ийн хариултуудыг is_correct-STATE False болгож зөв хариултгүй болгон
                    else:
                        request_data['score'] = None
                        request_data['has_score'] = False
                        other_choices = question_obj.choices.all()
                        other_choices.update(is_correct=False)

                request_data = remove_key_from_dict(request_data, ['level'])

                # Тухайн data-г base дээр update хийнэ
                updated_question_rows = PsychologicalTestQuestions.objects.filter(id=pk).update(
                    **request_data
                )

                if isinstance(question_img, str) != True:
                    question_img_path = get_image_path(question_obj)

                    file_path = save_file(question_img, question_img_path)

                    old_image = question_obj.image
                    question_obj.image = file_path
                    question_obj.save()
                    if old_image:
                        remove_folder(str(old_image))

                if isinstance(question_img, str) == True and question_img == '':
                    old_image = question_obj.image
                    question_img_path = get_image_path(question_obj)
                    # Хуучин зураг засах үедээ устгасан бол файл устгана.
                    question_obj.image = None
                    question_obj.save()
                    if old_image:
                        remove_folder(str(old_image))

                data = None
                if updated_question_rows > 0:
                    updated_question = PsychologicalTestQuestions.objects.filter(id=pk).first()
                    ser = dynamic_serializer(PsychologicalTestQuestions, "__all__", 1)
                    data = ser(updated_question).data
                return request.send_info('INF_002', data)

        # Хариултыг өөрчлөх үед
        else:
            answer_img = request_data["image"]
            answer_id = request_data["id"]
            updated_value = request_data['value']

            if 'score' in request_data:
                score = request_data.get('score')
                if score and (score != 'null' and score != 'undefined'):
                    request_data["is_correct"] = True
                else:
                    request_data['score'] = None

            request_data = remove_key_from_dict(request_data, ['image', 'id', 'score'])
            with transaction.atomic():
                answer_obj = PsychologicalQuestionChoices.objects.filter(id=answer_id).first()
                question_obj = PsychologicalTestQuestions.objects.filter(id=pk).first()
                updated_rows = PsychologicalQuestionChoices.objects.filter(id=answer_id).update(**request_data)
                if isinstance(answer_img, str) != True:
                    answer_img_path = get_choice_image_path(answer_obj)
                    file_path = save_file(answer_img, answer_img_path)
                    old_image = answer_obj.image
                    answer_obj.image = file_path
                    answer_obj.save()
                    if old_image:
                        remove_folder(str(old_image))

                if isinstance(answer_img, str) == True and answer_img == '':
                    old_image = answer_obj.image
                    answer_img_path = get_choice_image_path(answer_obj)
                    answer_obj.image = None
                    answer_obj.save()
                    if old_image:
                        remove_folder(str(old_image))

                data = None
                if updated_rows > 0:
                    updated_answer = PsychologicalQuestionChoices.objects.filter(id=answer_id).first()
                    ser = dynamic_serializer(PsychologicalQuestionChoices, "__all__")
                    updated_answer.value = updated_value
                    updated_answer.save()
                    data = ser(updated_answer).data
                return request.send_info('INF_002', data)


    def delete(self, request):

        delete_ids = request.query_params.getlist('delete')
        questions = self.queryset.filter(id__in=delete_ids)

        with transaction.atomic():
            try:
                for question in questions:

                    # Хэрвээ асуултанд зураг байвал устгана
                    question_img = question.image

                    if question_img:
                        remove_folder(str(question_img))

                        if question:
                            choices = question.choices.all()

                            for choice in choices:
                                img = choice.image

                                # Хэрвээ хариултын хэсэгт зураг байвал устгана
                                if img:
                                    remove_folder(str(img))

                                choice.delete()

                            # Хариултын хэсгийг устгах хэсэг
                            question.choices.clear()

                    question.delete()

            except Exception as e:
                print(e)
                return request.send_error("ERR_002")

        return request.send_info("INF_003")


class PsychologicalQuestionTitleListAPIView(APIView):

    """ Шалгалтын Гарчиг"""

    def get_queryset(self, user):
        if user.is_superuser:
            question_titles = PsychologicalTestQuestions.objects.all().values_list("title", flat=True)
        else:
            question_titles = PsychologicalTestQuestions.objects.filter(created_by=user).values_list("title", flat=True)
        return PsychologicalQuestionTitle.objects.filter(id__in=question_titles)

    def get(self, request):
        user = request.user
        data = self.get_queryset(user).values("id", "name")
        return request.send_data(list(data))


class PsychologicalQuestionTitleAPIView(
    generics.GenericAPIView,
    APIView,
    mixins.ListModelMixin,
    mixins.DestroyModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin

):
    """ Шалгалтын асуултууд Гарчигаар
    """
    queryset = PsychologicalQuestionTitle.objects.all()
    serializer_class = dynamic_serializer(PsychologicalQuestionTitle, "__all__")

    def get(self, request, pk=None):
        title_id = request.query_params.get('titleId')
        title_id = int(title_id)
        user = request.user.id
        user_obj = User.objects.filter(pk=user).first()
        if user_obj.is_superuser:
            challenge_qs = PsychologicalTestQuestions.objects.all().order_by('question_number')
            if title_id:
                challenge_qs = challenge_qs.filter(title=title_id)

            ser = dynamic_serializer(PsychologicalTestQuestions, "__all__", 1)
            data = ser(challenge_qs, many=True)
            count = challenge_qs.count()

            result = {
                "count": count,
                "results": data.data
            }
            return request.send_data(result)

        if pk:
            data = self.retrieve(request, pk).data
            questions = PsychologicalTestQuestions.objects.filter(title=pk, created_by=user)
            other_questions = PsychologicalTestQuestions.objects.filter(created_by=user).exclude(id__in=questions.values_list('id', flat=True)).values("id", "question", "title__name").order_by('question_number')

            questions = questions.values("id", "question", "title__name")
            return request.send_data({"title": data, "questions": list(questions), "other_questions": list(other_questions)})

        # 0  Бүх асуулт
        if title_id == 0 and not user_obj.is_superuser:
            challenge_qs = PsychologicalTestQuestions.objects.filter(created_by=user)
        # -1  Сэдэвгүй асуултууд
        elif title_id == -1:
            challenge_qs = PsychologicalTestQuestions.objects.filter(Q(created_by=user) & Q(title__isnull=True))
        else:
            challenge_qs = PsychologicalTestQuestions.objects.filter(created_by=user, title=title_id)

        ser = dynamic_serializer(PsychologicalTestQuestions, "__all__", 1)
        challenge_qs = challenge_qs.order_by('question_number')
        data = ser(challenge_qs, many=True)
        count = challenge_qs.count()

        result = {
            "count": count,
            "results": data.data
        }
        return request.send_data(result)


    def post(self, request):
        request_data = request.data
        question_ids = request.data.pop("questions")
        serializer = self.get_serializer(data=request_data)
        if serializer.is_valid(raise_exception=True):
            saved_obj =  serializer.save()
            questions_to_update = PsychologicalTestQuestions.objects.filter(id__in=question_ids)
            for question in questions_to_update:
                question.title.add(saved_obj)
            data = self.serializer_class(saved_obj).data
            return request.send_info("INF_001", data)
        else:
            print(serializer.errors)
        return request.send_info("ERR_001")


    def put(self, request, pk=None):
        request_data = request.data
        question_ids = request.data.pop("questions")
        other_question_ids = request.data.pop("other_questions")
        qs = self.queryset.filter(id=pk).get()
        serializer = self.get_serializer(qs, data=request_data, partial=True)
        if serializer.is_valid(raise_exception=False):
            self.perform_update(serializer)
            questions_to_update = PsychologicalTestQuestions.objects.filter(id__in=question_ids)
            other_questions = PsychologicalTestQuestions.objects.filter(id__in=other_question_ids)
            for question in questions_to_update:
                question.title.add(qs)
            for question in other_questions:
                question.title.remove(qs)
            return request.send_info("INF_002")

        return request.send_info("ERR_001")


    def delete(self, request, pk=None):
        self.destroy(request, pk)
        return request.send_info("INF_003")


class PsychologicalTestOptionsAPIView(
    mixins.ListModelMixin,
    generics.GenericAPIView
):
    def get(self, request):

        # PsychologicalTest model-ийн scope_kind-ийн val-ийг авна
        set_of_scope = PsychologicalTest.SCOPE_CHOICES
        # Ирсэн set data-г [{}] болгон хөрвүүлнэ
        scope_options = [{'id':id, 'name':name} for id, name in set_of_scope]

        # type_option-ийг гараараа бичээд явуулчихлаа
        type_options = [{'id':1, 'name':'Оноогүй'},{'id':2, 'name':'Асуултаас шалтгаалах'}]

        return_datas ={
            'scope_options':scope_options,
            'type_options':type_options
        }
        return request.send_data(return_datas)


@permission_classes([IsAuthenticated])
class PsychologicalTestAPIView(
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    generics.GenericAPIView
):

    """ Сэтгэл зүйн шалгалт сорил """

    queryset = PsychologicalTest.objects.all().order_by('id')
    serializer_class = PsychologicalTestSerializer

    pagination_class = CustomPagination

    filter_backends = [SearchFilter]
    search_fields = ['title', 'description']

    @has_permission(must_permissions=['lms-psychologicaltesting-exam-read'])
    def get(self, request, pk=None):

        scope = self.request.query_params.get('scope')

        if scope:
            self.queryset = self.queryset.filter(scope_kind=scope)

        if pk:
            data = self.retrieve(request, pk).data
            return request.send_data(data)

        datas = self.list(request).data
        return request.send_data(datas)

    def post(self, request):
        with transaction.atomic():
            sid = transaction.savepoint()
            try:
                # Үүсгэсэн хэрэглэгч нь employee байна
                datas = request.data.copy()
                datas['created_by'] = Employee.objects.get(user=request.user.id).id

                # Бүх датагаа serializer-даад хадгална
                serializer = self.get_serializer(data=datas)
                if serializer.is_valid(raise_exception=True):
                    serializer.save()
                transaction.savepoint_commit(sid)
                return request.send_info("INF_001")
            except Exception as e:
                print(e)
                return request.send_error('ERR_002')

    def put(self, request, pk=None):
        # Өөрчлөлт хийх сорилын instance
        instance = self.queryset.filter(id=pk).first()

        with transaction.atomic():
            sid = transaction.savepoint()
            try:
                data = request.data.copy()
                data['created_by'] = Employee.objects.get(user=request.user.id).id

                # description болон duration 2 null ирж болно
                if data['duration'] == 'null':
                    data['duration'] = None
                if data['description'] == 'null':
                    data['description'] = None

                serializer = self.get_serializer(instance, data=data)
                # Тэгээд шууд serializer ашигланаа
                if serializer.is_valid(raise_exception=True):
                    serializer.save()
                transaction.savepoint_commit(sid)
                return request.send_info("INF_002")
            except Exception as e:
                transaction.savepoint_rollback(sid)
                print(e)
                return request.send_error('ERR_002')

    def delete(self, request, pk):
        self.destroy(request, pk)
        return request.send_info("INF_003")


class PsychologicalTestOneAPIView(
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    generics.GenericAPIView
):
    """ Сэтгэл зүйн сорилд хамаарагдах багц асуултууд """

    queryset = PsychologicalTestQuestions.objects.all().order_by('id')
    serializer_class = PsychologicalTestQuestionsSerializer

    pagination_class = CustomPagination

    def get(self, request):
        # Асуулт нь ямар сорилд хамааралтайг мэдэхийн тулд сорилын id-г авна
        test_id = self.request.query_params.get('test_id')

        # Тухайн сорилын асуултуудын id
        question_ids = PsychologicalTest.objects.filter(id=test_id).values_list('questions', flat=True)

        # Асуултуудынхаа id-г ашиглан асуултуудаа аваад дараа нь serializer ашиглан буцаана
        if question_ids:
            self.queryset = self.queryset.filter(id__in=question_ids)

        datas = self.list(request).data
        return request.send_data(datas)

    def post(self, request, pk):
        data = request.data

        # Тухайн сорилд нэмэх асуултын багцуудын  id-г хадгална
        question_title_ids = list()
        queryset = PsychologicalTest.objects.filter(id=pk).first()

        # Хэрэглэгчийн сонгосон асуултын багцуудын id-г авна
        for value in data:
            question_title_ids.append(value['id'])

        # Тус багцад хамаарах асуултуудыг, багцын id-г ашиглан хадгална
        question_ids = self.queryset.filter(title__in=question_title_ids).values_list('id', flat=True)

        with transaction.atomic():
            try:
                if queryset:
                    # Тухайн сорилдоо асуултуудаа нэмээд хадгална
                    queryset.questions.add(*question_ids)
                    queryset.save()
                    return request.send_info("INF_001")
                else:
                    return request.send_error('ERR_002', "Тухайн багцад хамаарах асуултууд олдсонгүй")
            except Exception as e:
                print(e)
                return request.send_error('ERR_002')

    def delete(self, request, pk):
        # Сорилын id
        test_id = self.request.query_params.get('test_id')
        try:
            # Сорилын id-гаар сорилын qs-ийг аваад
            test = PsychologicalTest.objects.filter(id=test_id).first()
            if test:
                # Тухайн сорилын pk-р илэрхийлэгдэх асуултыг хасаад хадгална
                test.questions.remove(pk)
                test.save()
            else:
                return request.send_error('ERR_002', "Сорил олдсонгүй")
        except Exception as e:
            print(e)
            return request.send_error('ERR_002')
        return request.send_info("INF_003")


class PsychologicalTestScopesAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
):
    queryset = PsychologicalTest.objects.all()

    def get(self, request):
        datas = []

        # Parametr-үүд
        test_id = request.query_params.get('test_id')
        search_value = request.query_params.get('search')

        # Тухайн ёорилоо авна
        test_instance = self.queryset.get(id=test_id)

        # Cорилын хамрах хүрээний төрлөөс шалтгаалан хамрах хүрээг хаанаас авхаа тодорхойлно
        scope = test_instance.scope_kind
        participants = test_instance.participants

        # Хуудаслалт
        self.pagination_class = CustomPagination

        # Хамрах хүрээний боломжит утгууд
        scope_to_model_serializer = {
            1: (Teachers, TeachersSerializer, ['first_name', 'last_name', 'register']),
            2: (ElseltUser, ElsegchSerializer, ['first_name', 'last_name', 'code']),
            3: (Student, StudentSerializer, ['first_name', 'last_name', 'code'])
        }

        # scope-өөс шалтгаалан ашиглах model, serializer өөр, өөр байна
        # None, none гэсэн нь шууд байгаа утгыг нь авна
        model_class, serializer_class, search_fields = scope_to_model_serializer.get(scope, (None, None, None))

        # scope-д тохирсон model байвал
        if model_class is not None:
            # Оролцогчдоороо filter-ээд
            self.queryset = model_class.objects.filter(id__in=participants)
            # Serializer-г нь заагаад
            self.serializer_class = serializer_class
            # Хайх утгуудын өгнө
            if search_value:
                self.queryset = _filter_queries(self.queryset, search_value, search_fields)

            datas = self.list(request).data
        return request.send_data(datas)

@permission_classes([IsAuthenticated])
class PsychologicalTestScopeOptionsAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
):
    """ Сэтгэл зүйн сорилын хамрах хүрээг сонгох """

    def get(self, request):
        scope = self.request.query_params.get('scope')
        school = self.request.query_params.get('school')
        mode = self.request.query_params.get('mode')
        state = self.request.query_params.get('state')
        select1 = self.request.query_params.get('select1')

        # only for participants filtering arguments
        if mode == 'participants':
            select2 = self.request.query_params.get('select2')
            search_query = self.request.query_params.get('search')

        # if school selected filters all by school
        if school:
            school = int(school)

        # to avoid not assigned errors on return
        return_data = {}
        admission_options = department_options = group_options = teacher_options = elselt_user_options = student_options = []

        # partial loading of participants on startup and ondemand
        if mode in ['start'] or mode in ['participants'] and not search_query:

            # scroll lazy loading inside select input
            if state == '2':
                qs_start = (int(state) - 2) * 10
                qs_filter = int(state) * 10
            else:
                qs_start = (int(state) - 1) * 10
                qs_filter = int(state) * 10

        # Teachers
        if mode in ['start'] or mode in ['participants'] and scope == '1':
            teacher_options = Teachers.objects.order_by('id')
            if school:
                teacher_options = teacher_options.filter(sub_org=school)

        # filter teachers by select input
        if mode in ['participants'] and scope == '1':
            if select1:
                select1 = [int(item) for item in select1.split(',')]
                teacher_options = teacher_options.filter(salbar__in=select1)

            if search_query:
                search_vector = Q()

                # List of searchable fields
                for field in ['register', 'first_name']:
                    search_vector |= Q(**{f"{field}__icontains": search_query})
                teacher_options = teacher_options.filter(search_vector)

        # load by scroll if search by string is not performed
        if mode in ['start'] or mode in ['participants'] and scope == '1' and not search_query:
            teacher_options = teacher_options[qs_start:qs_filter]

        # building output data fields
        if mode in ['start'] or mode in ['participants'] and scope == '1':
            teacher_options = TeachersSerializer(teacher_options, many=True).data

            # mapping of original fields to custom keys and getting only them
            only_mapped_fields = {
                'id': 'id',
                'register': 'code',
                'full_name': 'full_name'
            }
            teacher_options = [{only_mapped_fields.get(key, key): item[key] for key in item if key in only_mapped_fields} for item in teacher_options]

        # Elselt users
        if mode in ['start'] or mode in ['participants'] and scope == '2':
            # Бие бялдар тэнцсэн элсэгч
            elselt_user_options = ElseltUser.objects.filter(physqueuser__state=AdmissionUserProfession.STATE_APPROVE, admissionuserprofession__user=F('id')).order_by('id')

        # filter elselt users by select inputs
        if mode in ['participants'] and scope == '2':
            if select1:
                elselt_user_options = elselt_user_options.filter(admissionuserprofession__profession__admission=select1)

            if select2:
                select2 = [int(item) for item in select2.split(',')]
                elselt_user_options = elselt_user_options.filter(admissionuserprofession__profession__profession__in=select2)

            if search_query:
                search_vector = Q()

                # List of searchable fields
                for field in ['code', 'register', 'first_name']:
                    search_vector |= Q(**{f"{field}__icontains": search_query})
                elselt_user_options = elselt_user_options.filter(search_vector)

        # load by scroll if search by string is not performed
        if mode in ['start'] or mode in ['participants'] and scope == '2' and not search_query:
            elselt_user_options = elselt_user_options[qs_start:qs_filter]

        # building output data fields
        if mode in ['start'] or mode in ['participants'] and scope == '2':
            elselt_user_options = ElsegchSerializer(elselt_user_options, many=True).data

            # make "code" key from register or code fields because some records contains emails instead of real code
            elselt_user_options_temp = []
            for ordered_dict in elselt_user_options:
                filtered_dict = {}
                filtered_dict['id'] = ordered_dict['id']
                filtered_dict['code'] = ordered_dict['code'] if ordered_dict['code'] and '@' not in ordered_dict['code'] else ordered_dict['register'] if '@' not in ordered_dict['register'] else ''
                filtered_dict['full_name'] = ordered_dict['full_name']

                # Append the new OrderedDict to the new_data list
                elselt_user_options_temp.append(filtered_dict)
            elselt_user_options = elselt_user_options_temp
            del elselt_user_options_temp

        # Students
        if mode in ['start'] or mode in ['participants'] and scope == '3':
            student_options = Student.objects.order_by('id')
            if school:
                student_options = student_options.filter(school=school)

        # filter students by select inputs
        if mode in ['participants'] and scope == '3':
            if select1:
                student_options = student_options.filter(department=select1)

            if select2:
                select2 = [int(item) for item in select2.split(',')]
                student_options = student_options.filter(group__in=select2)

            if search_query:
                search_vector = Q()

                # List of searchable fields
                for field in ['code', 'first_name']:
                    search_vector |= Q(**{f"{field}__icontains": search_query})
                student_options = student_options.filter(search_vector)

        # load by scroll if search by string is not performed
        if mode in ['start'] or mode in ['participants'] and scope == '3' and not search_query:
            student_options = student_options[qs_start:qs_filter]

        # building output data fields
        if mode in ['start'] or mode in ['participants'] and scope == '3':
            student_options = StudentSerializer(student_options, many=True).data

            # getting only specified keys
            only_mapped_fields = {
                'id': 'id',
                'code': 'code',
                'full_name': 'full_name'
            }
            student_options = [{only_mapped_fields.get(key, key): item[key] for key in item if key in only_mapped_fields} for item in student_options]

        if mode in ['start']:
            # Teachers and students departments
            department_options = Salbars.objects.all()
            if school:
                department_options = department_options.filter(sub_orgs=school)

            # get only required fields after all filters
            department_options =  department_options.values('id', 'name')

            # Elselt users admissions
            admission_options = AdmissionRegisterProfession.objects.annotate(
                admission_name = F('admission__name'),
            ).values('admission_name', 'admission').distinct('admission')

        # Students groups
        if mode in ['start', 'group']:
            group_options = Group.objects.all()
            if school:
                group_options = group_options.filter(school=school)

        # to filter students groups by selected departments
        if mode in ['group'] and select1:
            group_options = group_options.filter(department=select1)

        # get only required fields after all filters
        if mode in ['start', 'group']:
            group_options = group_options.values('id', 'name')

        # Тэгээд  select-д харуулхын тулд буцаана
        return_data = {
            'scope_kind': scope,
            'admission_options': list(admission_options),
            'department_options': list(department_options),
            'group_options': list(group_options),
            'teacher_options': teacher_options,
            'elselt_user_options': elselt_user_options,
            'student_options': student_options
        }

        return request.send_data(return_data)

    def put(self, request, pk):
        # Data болон scope, оролцогчдоо авна
        datas = request.data
        scope = datas.get('scope')
        school_id = datas.get('school_id')

        # teachers select inputs
        department_teacher = datas.get('department_teacher')
        teacher = datas.get('teacher')

        # elselt users select inputs
        admission = datas.get('admission')
        profession = datas.get('profession')
        elselt_user = datas.get('elselt_user')

        # students select inputs
        department_student = datas.get('department_student')
        group = datas.get('group')
        student = datas.get('student')

        # pk болон scope 2-ийн утга 2-уулаа байх үед
        if pk and scope is not None:
            # Оролцогчдын төрлөө өөрчлөөд
            PsychologicalTest.objects.filter(id=pk).update(scope_kind=scope)

            # if atleast 1 filter is not selected then do not put anything
            if (not department_teacher and not teacher and
                not admission and not profession and not elselt_user and
                not department_student and not group and not student):
                return request.send_info("INF_002")

            # Хэрвээ багшаа сорил авах бол
            if scope == 1:
                participant_ids = Teachers.objects.all()

                if teacher:
                    teacher = [item['id'] for item in teacher]
                    participant_ids = participant_ids.filter(id__in=teacher)

                if department_teacher:
                    department_teacher = [item['id'] for item in department_teacher]
                    participant_ids = participant_ids.filter(salbar__in=department_teacher)

                if school_id:
                    participant_ids = participant_ids.filter(sub_org=school_id)

            # Хэрвээ элсэгчдээс сорил авах бол
            elif scope == 2:
                # Бие бялдар тэнцсэн элсэгч
                participant_ids = ElseltUser.objects.filter(physqueuser__state=AdmissionUserProfession.STATE_APPROVE, admissionuserprofession__user=F('id'))

                if elselt_user:
                    elselt_user = [item['id'] for item in elselt_user]
                    participant_ids = participant_ids.filter(id__in=elselt_user)

                if profession:
                    profession = [item['prof_id'] for item in profession]
                    participant_ids = participant_ids.filter(admissionuserprofession__profession__profession__in=profession)

                if admission:
                    participant_ids = participant_ids.filter(admissionuserprofession__profession__admission=admission['admission'])

            # Хэрвээ оюутанаас сорил авах бол
            elif scope == 3:
                participant_ids = Student.objects.all()

                if student:
                    student = [item['id'] for item in student]
                    participant_ids = participant_ids.filter(id__in=student)

                if group:
                    group = [item['id'] for item in group]
                    participant_ids = participant_ids.filter(group__in=group)

                if department_student:
                    participant_ids = participant_ids.filter(department=department_student['id'])

                if school_id:
                    participant_ids = participant_ids.filter(school=school_id)

            # Хуучин хүүхдүүд
            old_participants = PsychologicalTest.objects.get(id=pk).participants or []

            # Convert lists to sets and find the difference
            set_parts = set(participant_ids.values_list('id', flat=True))
            set_old_participants = set(old_participants)

            # Өмнө нэмэгдсэн хэрэглэгчийг олох
            common_elements = set_parts & set_old_participants

            # Огт нэмэгдээгүй хэрэглэгчид
            unique_parts = set_parts - common_elements

            # Convert sets back to lists
            unique_parts = list(unique_parts)

            # Хуучин сэтгэлзүй өгөх хэрэглэгч дээрээ шинээр нэмэгдэж байгаагаа нэмж update хийнэ
            old_new_participants = old_participants + unique_parts

            # Тэгээд эцэст нь бааздаа хадгална
            PsychologicalTest.objects.filter(id=pk).update(participants=list(old_new_participants))

        return request.send_info("INF_002")

    def delete(self, request, pk):
        # Сорилын id
        test_id = self.request.query_params.get('test_id')
        try:
            # Сорилын id-гаар сорилын qs-ийг аваад
            test = PsychologicalTest.objects.filter(id=test_id).first()
            if test:
                # Тухайн сорилын pk-р илэрхийлэгдэх залууг хасаад хадгална
                test.participants.remove(pk)
                test.save()
            else:
                return request.send_error('ERR_002', "Хэрэглэгч олдсонгүй")
        except Exception as e:
            print(e)
            return request.send_error('ERR_002')
        return request.send_info("INF_003")


class PsychologicalTestResultAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    """ Сэтгэлзүйн сорилын үр дүн """

    queryset = PsychologicalTest.objects.all()
    serializer_class = PsychologicalTestResultSerializer

    pagination_class = CustomPagination

    filter_backends = [SearchFilter]
    search_fields = ['title', 'start_date', 'end_date', 'description', 'duration']

    def get(self, request):
        scope = self.request.query_params.get('scope')

        if scope:
            self.queryset = self.queryset.filter(scope_kind=scope)

        data = self.list(request).data
        return request.send_data(data)


class PsychologicalTestResultParticipantsAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    """ Сэтгэлзүйн сорилд оролцогчид """

    queryset = PsychologicalTest.objects.all()

    def get(self, request):
        datas = []

        # Parametr-үүд
        test_id = request.query_params.get('test_id')
        search_value = request.query_params.get('search')

        # Тухайн сорилоо авна
        test_instance = self.queryset.get(id=test_id)

        # Cорилын хамрах хүрээний төрлөөс шалтгаалан хамрах хүрээг хаанаас авхаа тодорхойлно
        scope = test_instance.scope_kind
        participants = test_instance.participants

        if participants is None:
            return request.send_data(datas)

        # Хуудаслалт
        self.pagination_class = CustomPagination

        # Хамрах хүрээний боломжит утгууд
        scope_to_model_serializer = {
            1: (Teachers, TeachersSerializer, ['first_name', 'last_name', 'register']),
            2: (MentalUser, MentalUserSerializer, ['user__first_name', 'user__last_name', 'user__code']),
            3: (Student, StudentSerializer, ['first_name', 'last_name', 'code'])
        }

        # scope-өөс шалтгаалан ашиглах model, serializer өөр, өөр байна
        # None, none гэсэн нь шууд байгаа утгыг нь авна
        model_class, serializer_class, search_fields = scope_to_model_serializer.get(scope, (None, None, None))

        # scope-д тохирсон model байвал
        if model_class is not None:
            if model_class == MentalUser:
                self.queryset = model_class.objects.filter(user__in=participants, challenge=test_id).annotate(
                    first_name = F('user__first_name'),
                    last_name = F('user__last_name'),
                    code = F('user__code'),
                )
            else:
                # Оролцогчдоороо filter-ээд
                self.queryset = model_class.objects.filter(id__in=participants, challenge=test_id)
            # Serializer-г нь заагаад
            self.serializer_class = serializer_class
            # Хайх утгуудын өгнө
            if search_value:
                self.queryset = _filter_queries(self.queryset, search_value, search_fields)

            datas = self.list(request).data
        return request.send_data(datas)

class PsychologicalTestResultShowAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    """Сорилийн оноо асуулт хариултыг харах API"""

    queryset = PsychologicalTestQuestions.objects.all()
    serializer_class = PsychologicalTestQuestionsSerializer

    def post(self,request):

        datas = request.data
        question_ids = []
        chosen_choices = []
        big_data = []
        return_data = []
        totalscore = 0

        # Орж ирж буй датан {} хаалттай орж ирж байгаа учир авна
        value = str(datas)[1:-1]

        # таслалаар нь салгах
        pairs = [pair.strip() for pair in value.split(',')]

        # 176:250 ийм датаг урд талын question_ids-д авч хойд тал нь сонгосон хариултын id
        for pair in pairs:
            question_id, choice_id = pair.split(':')
            question_ids.append(question_id.strip().strip("'"))
            chosen_choices.append(choice_id.strip().strip("'"))

        def convert_to_int(value):
            if value == 'True':
                return 1
            elif value == 'False':
                return 0
            else:
                return int(value)

        question_ids = list(map(int, question_ids))
        chosen_choices = list(map(convert_to_int, chosen_choices))

        # Тухайн 2 датаг нийлүүлж асуултын id гаар нь бүх мэдээлэлийн PsychologicalTestQuestions model-оос авчирна
        for question_id, choice_id in zip(question_ids, chosen_choices):
            queryset = PsychologicalTestQuestions.objects.filter(id=question_id).first()
            if queryset:

                serializer = self.serializer_class(queryset)
                data = serializer.data

                if choice_id != 1 and choice_id != 0:
                    # big_data шалгуулагчийн сонгосон хариултын id буцаан
                    data['chosen_choice'] = int(choice_id)
                else:
                    data['chosen_choice'] = 'Тийм' if choice_id == 1 else 'Үгүй'

                # Тухайн өгсөн шалгалтийн нийт оноо
                if(data['has_score']):
                    totalscore += data['score']

                big_data.append(data)

        return_data ={
            'question':big_data,
            'total_score':totalscore
        }
        return request.send_data(return_data)


class PsychologicalTestResultExcelAPIView(
    generics.GenericAPIView
):
    """ Сэтгэлзүйн шалгалтын үр дүн тайлан excel """

    # Оноо бүрд if нөхцөл ашиглахгүйн тулд function ашигласан
    def classify_score(self, score, thresholds, labels):
        # threshold нь үүнээс бага гэсэн нөхцлийг зааж өгнө
        for idx, threshold in enumerate(thresholds):
            # Оноо нь thres дотор байвал label-ийн ижил idx-тэй value-г буцаана
            if score <= threshold:
                return labels[idx]
            # Бусад үед маш хүчтэй гэсэн value-г буцаана
        return labels[-1]

    # Асуулга ялгаж өгөх function
    def find_question_type(self, all_questions, type):
        # type-аас шалтгаалан Асуулга 3 эсвэл 4 гэсэн үг асуултын нэр дотор байвал тэр шүүсэн асуултуудаа аваад
        question_type = PsychologicalTestQuestions.objects.filter(question__icontains=f'Асуулга-{type}')
        # Хэрэглэгчийн хариулсан бүх асуулт дунд шүүсэн асуултуудтай ижил id-тай асуулт байвал тэднийг dict дотор багцлаад буцаана
        return {str(question.id): all_questions[str(question.id)] for question in question_type if str(question.id) in all_questions}

    # DASS21 нэг бүрчилсэн өгөгдөл
    def dass21(self, user, questions):
        # Хариулсан асуултуудын id-ууд
        question_ids = list(questions.keys())
        datas = {}

        # Нэг dict дотор key-нь асуултын id value-нь тухайн асуултын obj байхаар авна
        questions_qs_dass21 = PsychologicalTestQuestions.objects.filter(id__in=question_ids).in_bulk(field_name='id')

        # Асуултууд дотроо loop гүйлгээд ---> {1:QuestionObj(1)}
        for key, value in questions_qs_dass21.items():
            # Хэрэглэгчийн хариулсан асуултуудын id дунд questions_qs_dass21-ийн key-нь байвал
            if str(key) in question_ids:
                # Тухайн асуултанд хариулсан хариултын id-аар PsychologicalQuestionChoices-модел-с шүүд obj-ийш авна
                choice = PsychologicalQuestionChoices.objects.filter(id=questions[f'{key}']).first()
                # Дараа нь буцаах дата дотроо тухайн асуултын question_number-ийг key болгоод
                # value-д нь сонгосон хариултын value-г өгнө
                datas[f'{value.question_number}'] = int(choice.value)

        # Сүүлд нь хэрэглэгчийнхээ мэдээллийг өгөөд
        datas['first_name'] = user.first_name
        datas['last_name'] = user.last_name
        datas['register'] = user.register

        return datas # Буцаана

    # Сурах сэдэл нэг бүрчилсэн өгөгдөл
    def motivation(self, user, questions, all_score_questions):
        # Хариулсан асуултуудын id-ууд
        question_ids = list(questions.keys())
        datas = {}

        # Нэг dict дотор key-нь асуултын id value-нь тухайн асуултын obj байхаар авна
        question_qs_dict = PsychologicalTestQuestions.objects.filter(id__in=question_ids).in_bulk(field_name='id')

        # Ийм question_number-тай асуултанд тийм гэж хариулсан бол 1 оноо
        group_true = {4, 17, 26, 9, 31, 33, 43, 48, 49, 24, 35, 38, 44}
        # Ийм question_number-тай асуултанд үгүй гэж хариулсан бол 1 оноо
        group_false = {28, 42, 11}

        # Асуултууд дотроо loop гүйлгээд ---> {1:QuestionObj(1)}
        for key, value in question_qs_dict.items():
            question_number = value.question_number

            # key нь асуултын id, type-ийн тааруулхын тулд str болгосон
            key_str = str(key)

            # Хэрэглэгчийн хариулсан асуултуудын id дунд key_str байвал
            if key_str in question_ids:
                # response-д тухайн асуултанд хэрэглэгч юу гэж хариулсан хариултыг авна
                # шууд авдаг нь асуултын төрөл тийм үгүй асуулт учираас question_choice-ийн id биш шууд True False ирнэ
                response = questions[key_str]

                # Хариулт True бол datas дотроо key-д нь шууд тухайн асуултын question_number-ийг өгөөд
                if response is True:
                    # value-д Тийм гээд бичээд өгсөн
                    datas[question_number] = 'Тийм'
                # Хариулт False бол
                elif response is False:
                    # value-д Үгүй гээд бичээд өгсөн
                    datas[question_number] = 'Үгүй'

                # Тухайн for дотор давтаж байгаа асуултын question_number нь оноо өгөх асуултууд дунд байвал
                if question_number in list(all_score_questions.keys()):
                    # Тухайн асуултанд оноогдох оноог аваад
                    score = all_score_questions[question_number]

                    # Асуултанд хариулсан байдлаас шалтгаалан оноог өгнө
                    if question_number in group_true:
                        # group_true дотор байгаа асуултан Тийм гэж хариулсан бол 1 оноо өгнө гэх мэт
                        datas[f'score_{question_number}'] = score if response is True else 0
                    elif question_number in group_false:
                        datas[f'score_{question_number}'] = score if response is False else 0
                    else: # function-ээ дуудах үед зөв датагаа өгсөн болохоор else-рүү ерөнхийдөө орохгүй
                        pass
        # Сүүлд нь хэрэглэгчийнхээ мэдээллийг өгөөд
        datas['first_name'] = user.first_name
        datas['last_name'] = user.last_name
        datas['register'] = user.register

        return datas # Буцаана

    # Прогноз нэг бүрчилсэн өгөгдөл
    def prognoz(self, user, questions):
        # Хариулсан асуултуудын id-ууд
        question_ids = list(questions.keys())
        datas = {}

        # Нэг dict дотор key-нь асуултын id value-нь тухайн асуултын obj байхаар авна
        questions_qs_prognoz = PsychologicalTestQuestions.objects.filter(id__in=question_ids).in_bulk(field_name='id')

        # Ийм question_number-тай асуултууд яаж хариулснаас шалтгаалан оноотой байна
        mental_true = [
            2, 3, 5, 7, 9, 11, 13, 14, 16, 18, 20, 22, 23, 25, 27, 28,
            29, 31, 32, 33, 34, 36, 37, 39, 40, 42, 43, 45, 47, 48, 51,
            53, 54, 56, 57, 59, 60, 62, 63, 65, 66, 67, 68, 69, 70, 71,
            72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 84, 85, 86
        ]
        mental_false = [4, 8, 17, 24, 30, 35, 41, 46, 50, 55, 64]
        # Энэ асуултанд Үгүй гэж хариулсан бол оноо өгнө, нөгөө 2 нь хувьсагчийн нэрээрээ тодорхой
        true_false = [1, 6, 10, 12, 15, 19, 21, 26, 33, 38, 44, 49, 52, 58, 61]

        # Асуултууд дотроо loop гүйлгээд ---> {1:QuestionObj(1)}
        for key, value in questions_qs_prognoz.items():
            # Тухайн асуултын question_number-ийг авна
            question_number = value.question_number
            # key нь асуултын id, type-ийн тааруулхын тулд str болгосон
            key_str = str(key)

            # Хэрэглэгчийн хариулсан асуултуудын id дунд key_str байвал
            if key_str in question_ids:
                # response-д тухайн асуултанд хэрэглэгч юу гэж хариулсан хариултыг авна
                # шууд авдаг нь асуултын төрөл тийм үгүй асуулт учираас question_choice-ийн id биш шууд True False ирнэ
                response = questions[key_str]

                # Хариулт True бол datas дотроо key-д нь шууд тухайн асуултын question_number-ийг өгөөд
                if response is True:
                    # value-д Тийм гээд бичээд өгсөн
                    datas[question_number] = 'Тийм'
                # Хариулт False бол
                elif response is False:
                    # value-д Үгүй гээд бичээд өгсөн
                    datas[question_number] = 'Үгүй'

                # Асуултанд хариулсан байдлаас шалтгаалан оноог өгнө
                if question_number in mental_false or question_number in true_false:
                    # Өмнөх function-уудтай ижилээр буцаах дата дотроо тухайн асуултын question_number-ийг key болгоод
                    # value-д нь хариултын эсрэг утгыг өгнө. Яагаад гэвэл энэ group-лсэн асуултууд нь False үед л оноо өгнө
                    datas[f'score_{question_number}'] = int(not response)
                    # boolean утгыг int болгож өгвөл 1, 0 гэсэг утга буцаана
                elif question_number in mental_true:
                    # Тийм үед оноо өгөх асуултуудыг яг response-оор нь л value-д утгыг нь оноож өгнө
                    datas[f'score_{question_number}'] = int(response)


        # Сүүлд нь хэрэглэгчийнхээ мэдээллийг өгөөд
        datas['first_name'] = user.first_name
        datas['last_name'] = user.last_name
        datas['register'] = user.register

        return datas # Буцаана

    def get(self, request):
        adm = request.query_params.get('adm') # Элсэлтээр хайх

        # Нийт буцаах датануудыг store хийх variable-ууд
        datas = list()
        dass21_datas = list()
        motivation_datas = list()
        prognoz_datas = list()

        user_ids = AdmissionUserProfession.objects.all().values_list('user', flat=True)
        if adm:
            prod_ids = AdmissionRegisterProfession.objects.filter(admission=adm).values_list('id', flat=True)
            user_ids = AdmissionUserProfession.objects.filter(profession__in=prod_ids).values_list('user', flat=True)

        queryset = MentalUser.objects.filter(user__in=user_ids)

        # Шалгалт өгсөн элсэлгчдийн шалгалтын хариу
        mental_users = queryset.filter(challenge__title__icontains='Сэтгэлзүйн сорил').select_related('user')

        # Сэтгэл гутралын асуултууд
        depression_questions = set(PsychologicalTest.objects.filter(
            title__icontains='Сэтгэлзүйн сорил',
            questions__question_number__in=[3, 5, 10, 13, 16, 17, 21]
        ).values_list('questions__id', flat=True))

        # Түгшүүрийн асуултууд
        anxiety_questions = set(PsychologicalTest.objects.filter(
            title__icontains='Сэтгэлзүйн сорил',
            questions__question_number__in=[2, 4, 7, 9, 15, 19, 20]
        ).values_list('questions__id', flat=True))

        # Стрессийн асуултууд
        stress_questions = set(PsychologicalTest.objects.filter(
            title__icontains='Сэтгэлзүйн сорил',
            questions__question_number__in=[1, 6, 8, 11, 12, 14, 18]
        ).values_list('questions__id', flat=True))

        # Бүх асуултын хариултуудыг нэг dict дотор key-д нь id-г нь өгөөд value-д нь obj-ын өгөөд авна
        # Энэ нь асуултуудаа ялгаж авхад хялбар бас хурдан
        question_choices = PsychologicalQuestionChoices.objects.in_bulk(field_name='id')

        # user-үүдээ бас өмнөхтэй ижил format-аар авсан
        elselt_users = {user.id: user for user in ElseltUser.objects.filter(id__in=mental_users.values_list('user_id', flat=True))}

        # Нийт сорил өгсөн хэрэглэгчид дотроо loop гүйлгээд
        for user in mental_users:
            # user_data-д нэг хэрэглэгчийн сорилын онооны нийт мэдээллийг store хийнэ
            user_data = dict()
            # Хэрэглэгчийн тухай мэдээллүүд
            user_obj = elselt_users[user.user.id]
            user_data['first_name'] = user_obj.first_name
            user_data['last_name'] = user_obj.last_name
            user_data['register'] = user_obj.register

            # Хэрэглэгчийн хариултууд
            if user.answer:
                # Хэрэглэгчийн тест өгсөн асуулт болон хариултуудыг аваад
                answer = ast.literal_eval(user.answer)

                # Тэр асуултуудыг асуулга, асуулгаар нь ялгаж өнгө
                type_question_3 = self.find_question_type(answer, '3')
                type_question_4 = self.find_question_type(answer, '4')
                # dass21-ийн асуултууд function ашиглах шаардлагагүй. Хариултууд нь boolean утгатай биш л байвал тэр нь
                # dass21-т харьялагдах асуултууд гэсэн үг
                dass21_answers = {key: value for key, value in answer.items() if not isinstance(value, bool)}

                # Үүнээс доошоо END хүртэл excel-ийн нийт гэсэн sheet-д хамаарагдах өгөгдлийн зохицуулалтууд явна
                # Асуулга-2 буюу DASS21 -------------------------------------------------------------------------------------------------------------------->
                # dass21-д харьялагдах асуултууд дунд түүний key-нь дээр ялгаж авсан асуултуудын id-тай таарвал тэр асуултуудын value-гийн нийлбэрийг авна
                total_depression_score = sum(int(question_choices[value].value) for key, value in dass21_answers.items() if int(key) in depression_questions)
                total_anxiety_score = sum(int(question_choices[value].value) for key, value in dass21_answers.items() if int(key) in anxiety_questions)
                total_stress_score = sum(int(question_choices[value].value) for key, value in dass21_answers.items() if int(key) in stress_questions)

                # Олж авсан онооны нийлбэрүүдээ user_data-д store хийнэ
                user_data['depression_score'] = total_depression_score
                user_data['anxiety_score'] = total_anxiety_score
                user_data['stress_score'] = total_stress_score

                # Тухайн оноонуудаас шалтгаалан хэрэглэгч ямар шинж тэмдэгтэй байгаад
                depression_thresholds = [4, 6, 10, 13]
                anxiety_thresholds = [3, 5, 7, 9]
                stress_thresholds = [7, 9, 12, 16]
                labels = ['Хэвийн', 'Хөнгөн', 'Дунд зэрэг', 'Хүчтэй', 'Маш хүчтэй']

                # classify_score function-ийг ашиглан мэдэж аваад user_data-д нэмнэ
                user_data['depression'] = self.classify_score(total_depression_score, depression_thresholds, labels)
                user_data['anxiety'] = self.classify_score(total_anxiety_score, anxiety_thresholds, labels)
                user_data['stress'] = self.classify_score(total_stress_score, stress_thresholds, labels)


                # Асуулга-3
                if type_question_3:
                    # Хэрэглэгчийн хариулсан асуулга 3-ийн асуултуудын id
                    type_question_3_keys = list(type_question_3.keys())

                    # Ийм dict-үүдийн key-тэй ижил question_number-тай асуултууд л оноо авна
                    # Оноо нь value-ууд
                    knowledge_score_map = {4: 3.6, 17: 3.6, 26: 2.4, 28: 1.2, 42: 1.2}
                    skill_score_map = {9: 1, 31: 2, 33: 2, 43: 3, 48: 1, 49: 1}
                    diplom_score_map = {24: 2.5, 35: 1.5, 38: 1.5, 44: 1, 11: 3.5}

                    # Дээрх 3-н асуултуудын question_number-ийг нэгтгэнэ
                    all_question_numbers = {**knowledge_score_map, **skill_score_map, **diplom_score_map}.keys()
                    # Энэ харин бүх юмтай нь нэгтгэнэ
                    all_score_questions = {**knowledge_score_map, **skill_score_map, **diplom_score_map}

                    # Бааз дээр хадгалагдаж байгаа оноо бүх Асуулга-3 ийн асуултууд
                    all_questions_type_3 = PsychologicalTestQuestions.objects.filter(
                        question__icontains='Асуулга-3',
                        question_number__in=all_question_numbers
                    ).in_bulk(field_name='id')

                    # Оноотой асуултуудын question_number-ууд dict дотор байвал тохирох оноонуудын нийлбэрийг олно
                    knowledge_score = sum(
                        knowledge_score_map[question.question_number]
                        for key, question in all_questions_type_3.items()
                        if str(key) in type_question_3_keys and (
                            (question.question_number in {4, 17, 26} and type_question_3[f'{key}']) or
                            (question.question_number in {28, 42} and not type_question_3[f'{key}'])
                        )
                    )
                    # skill_score-ийн оноотой байх нөхцөл нь оноотой асуултууд нь бүгд Тийм байх үед учираас
                    # нөгөө 2 оноо шиг дугааруудыг нь заавал заах шаардлагүй
                    skill_score = sum(
                        skill_score_map[question.question_number]
                        for key, question in all_questions_type_3.items()
                        if str(key) in type_question_3_keys and question.question_number in skill_score_map and type_question_3[f'{key}']
                    )
                    diplom_score = sum(
                        diplom_score_map[question.question_number]
                        for key, question in all_questions_type_3.items()
                        if str(key) in type_question_3_keys and (
                            (question.question_number in {24, 35, 38, 44} and type_question_3[f'{key}']) or
                            (question.question_number == 11 and not type_question_3[f'{key}'])
                        )
                    )

                    # Тэгээд user_data-д нэмээд өгнө
                    user_data['knowledge_score'] = knowledge_score
                    user_data['skill_score'] = skill_score
                    user_data['diplom_score'] = diplom_score

                # Асуулга-4
                if type_question_4:
                    # Хэрэглэгчийн хариулсан асуулга 4-ийн асуултуудын id
                    type_question_4_keys = list(type_question_4.keys())

                    # Оноо авах question_number-уудтай асуултууд
                    # Нөхцөл нь л таарвар бүгд 1 оноотой учир dict байх шаардлагагүй
                    true_false = [1, 6, 10, 12, 15, 19, 21, 26, 33, 38, 44, 49, 52, 58, 61]
                    mental_true = [
                        2, 3, 5, 7, 9, 11, 13, 14, 16, 18, 20, 22, 23, 25, 27, 28,
                        29, 31, 32, 33, 34, 36, 37, 39, 40, 42, 43, 45, 47, 48, 51,
                        53, 54, 56, 57, 59, 60, 62, 63, 65, 66, 67, 68, 69, 70, 71,
                        72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 84, 85, 86
                    ]
                    mental_false = [4, 8, 17, 24, 30, 35, 41, 46, 50, 55, 64]

                    # Дээрх 3-н асуултуудын question_number-ийг нэгтгэнэ
                    all_question_numbers_type_4 = mental_true + mental_false + true_false

                    # Нэгтгэсэн question_number-уудаа ашиглан бааз дээр байгаа Асуулга-4-ийн бүх
                    # obj-уудыг id-тай нь харгалзуулcан нэг dict авна
                    all_questions_type_4 = PsychologicalTestQuestions.objects.filter(
                        question__icontains='Асуулга-4',
                        question_number__in=all_question_numbers_type_4
                    ).in_bulk(field_name='id')

                    # Оноотой асуултуудын question_number-ууд dict дотор байвал тохирох оноонуудын нийлбэрийг олно
                    true_false_score = sum(
                        1 for key, question in all_questions_type_4.items()
                        if str(key) in type_question_4_keys and question.question_number in true_false and not type_question_4[f'{key}']
                    )
                    # Бүгдэнд нь сайхан 1 гэсэн оноо өгөөл sum-ийн олчино
                    mental_score = sum(
                        1 for key, question in all_questions_type_4.items()
                        if str(key) in type_question_4_keys and (
                            (question.question_number in mental_true and type_question_4[f'{key}']) or
                            (question.question_number in mental_false and not type_question_4[f'{key}'])
                        )
                    )

                    # Тэгээд user_data-д нэмээд өгнө
                    user_data['mental_score'] = mental_score
                    user_data['true_false_score'] = true_false_score

                    # classify_score function-ийг ашиглан шинж тэмдгийг мэдэж авна
                    overall_review_thresholds = [6, 13, 28, 16]
                    overall_labels = ['өндөр', 'сайн', 'дунд', 'муу']
                    user_data['overall_review'] = self.classify_score(mental_score, overall_review_thresholds, overall_labels)
                # END ------------------------------------------------------------------------------------------------------------------------>

                # Нэг бүрчилсэн өгөгдлүүдийг function-ууд шийднэ өгнө
                # DASS21 -------------------------------------------------------------------------------------------------------------------->
                dass21_data = self.dass21(user_obj, dass21_answers)
                dass21_data['depression_score'] = total_depression_score
                dass21_data['anxiety_score'] = total_anxiety_score
                dass21_data['stress_score'] = total_stress_score
                dass21_data['depression'] = user_data['depression']
                dass21_data['anxiety'] = user_data['anxiety']
                dass21_data['stress'] = user_data['stress']

                # Сурах сэдэл
                motivation_data = self.motivation(user_obj, type_question_3, all_score_questions)
                motivation_data['knowledge_score'] = knowledge_score
                motivation_data['skill_score'] = skill_score
                motivation_data['diplom_score'] = diplom_score

                # Прогноз
                prognoz_data = self.prognoz(user_obj, type_question_4)
                prognoz_data['mental_score'] = mental_score
                prognoz_data['true_false_score'] = true_false_score
                prognoz_data['overall_review'] = user_data['overall_review']
                # END ------------------------------------------------------------------------------------------------------------------------->

                # Хэрэглэгчийнхээ бүх мэдээллийг үндсэн буцаах list-үүддээ нэмнэ
                datas.append(user_data)
                dass21_datas.append(dass21_data)
                motivation_datas.append(motivation_data)
                prognoz_datas.append(prognoz_data)

        # Иймэрдүү хэлбэртэй датаг буцаана
        return_datas = {
            'overall_datas':datas,
            'dass21_datas':dass21_datas,
            'motivation_datas':motivation_datas,
            'prognoz_datas':prognoz_datas
        }
        return request.send_data(return_datas)

class IQTestResultExcelAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    serializer_class = PsychologicalTestQuestionsSerializer
    """ IQ Test үр дүн тайлан excel """

    def get(self, request):
            big_data = []
            adm = request.query_params.get('adm')

            # IQ test Нийт асуултын тоо авах
            question = PsychologicalTest.objects.filter(id=3).values('questions').count()
            user_ids = AdmissionUserProfession.objects.values_list('user', flat=True)
            if adm:
                prof_ids = AdmissionRegisterProfession.objects.filter(admission=adm).values_list('id', flat=True)
                user_ids = AdmissionUserProfession.objects.filter(profession__in=prof_ids).values_list('user', flat=True)

            queryser_mental = MentalUser.objects.filter(user__in=user_ids)
            mental_users = queryser_mental.filter(challenge__title__icontains='IQ Test', answer__isnull=False).select_related('user')
            elselt_users = {user.id: user for user in ElseltUser.objects.filter(id__in=mental_users.values_list('user_id', flat=True))}

            # Шалгалт өгсөн хүн бүр
            for user in mental_users:
                user_data = {
                    'first_name': '',
                    'last_name': '',
                    'register': '',
                    'scores': [],
                    'total_score': 0
                }
                user_obj = elselt_users[user.user.id]
                user_data['first_name'] = user_obj.first_name
                user_data['last_name'] = user_obj.last_name
                user_data['register'] = user_obj.register

                # Хариулттай үед
                answer = ast.literal_eval(user.answer)
                value = str(answer)[1:-1]
                pairs = [pair.strip() for pair in value.split(',')]

                question_ids = []
                chosen_choices = []
                for pair in pairs:
                    question_id, choice_id = pair.split(':')
                    question_ids.append(question_id.strip().strip("'"))
                    chosen_choices.append(choice_id.strip().strip("'"))

                def convert_to_int(value):
                    if value == 'True':
                        return 1
                    elif value == 'False':
                        return 0
                    else:
                        return int(value)

                question_ids = list(map(int, question_ids))
                chosen_choices = list(map(convert_to_int, chosen_choices))

                # TODO Энэ давталт доторх код л удаж байгаа шалтгаан байх
                for question_id, choice_id in zip(question_ids, chosen_choices):
                    # TODO Хүүхэд бүрээр сорилын асуултыг авах шаардлагагүй байх 1 сорилын асуултууд хүүхэд бүр дээр өөрчлөгдөхгүй учраас
                    queryset = PsychologicalTestQuestions.objects.filter(id=question_id).first()
                    if queryset:

                        # Хариулт зөв үгүйг шалгах
                        # TODO Шалгалтын асуулт хариулт бүрийг нэг л авчихвал энэ for дотор хүүхэд бүрийг тоогоор бааз руу ачааллах хурд сайжрах байх
                        choice = PsychologicalQuestionChoices.objects.filter(id=choice_id).first()
                        if choice:
                            score = queryset.score if choice.is_correct else 0

                            # Зөв бол оноо шалгах
                            user_data['scores'].append(int(score))
                            user_data['total_score'] += score

                big_data.append(user_data)

            return_data = {
                'question':question,
                'user_data':big_data
            }

            return request.send_data(return_data)

@permission_classes([IsAuthenticated])
class QuestionsListAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    """ Шалгалт үүсгэхдээ асуулт сонгох хэсэг """

    queryset = ChallengeQuestions.objects.all()
    serializer_class = ChallengeQuestionListSerializer

    def get(self, request):

        datas = []

        queryset = self.queryset

        checked_ids = self.request.query_params.getlist('checked')
        count = self.request.query_params.get('count')
        qtype = self.request.query_params.get('type')

        for check_id in checked_ids:
            obj = {}
            subject = Lesson_title_plan.objects.get(id=check_id)

            obj['id'] = int(check_id)
            obj['name'] = subject.title

            querysets = queryset.filter(subject_id=check_id)

            if count and qtype:
                scount = int(count)

                if qtype == 'Эхний':
                    querysets = querysets.order_by('created_at')
                else:
                    querysets = querysets.order_by('-created_at')

                querysets = querysets[0:scount]

            serializer = self.get_serializer(querysets, many=True).data

            obj['data'] = list(serializer)

            datas.append(obj)

        return request.send_data(datas)


@permission_classes([IsAuthenticated])
class ChallengeSendAPIView(
    generics.GenericAPIView
):
    """ Шалгалтын материал батлуулах хүсэлт илгээх """

    def get(self, request, pk=None):

        challlenge = Challenge.objects.get(id=pk)

        with transaction.atomic():
            challlenge.send_type = Challenge.SEND
            challlenge.save()

        return request.send_info('INF_019')



@permission_classes([IsAuthenticated])
class ChallengeApprovePIView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    """ ХБА батлах шалгалтын хүсэлтүүд """

    queryset = Challenge.objects.all()
    serializer_class = ChallengeListSerializer

    pagination_class = CustomPagination

    filter_backends = [SearchFilter]
    search_fields = ['title', 'start_date', 'end_date', 'lesson__name', 'lesson__code']

    def get(self, request):

        user_id = request.user

        ctype = request.query_params.get('type')
        search_teacher = request.query_params.get('teacher')
        lesson = request.query_params.get('lesson')

        # Шалгалтын төрлөөр хайх
        if ctype:
            self.queryset = self.queryset.filter(send_type=int(ctype))

        teacher = Teachers.objects.get(user=user_id)

        department_id = teacher.salbar.id

        self.queryset = self.queryset.exclude(challenge_type=Challenge.SELF_TEST)

        self.queryset = self.queryset.filter(created_by__salbar__id=department_id)

        teacher_ids = self.queryset.values_list('created_by', flat=True)

        self.queryset = self.queryset.filter(created_by__in=teacher_ids)

        if search_teacher:
            self.queryset = self.queryset.filter(created_by=search_teacher)

        if lesson:
            self.queryset = self.queryset.filter(lesson=lesson)

        all_list = self.list(request).data

        return request.send_data(all_list)

    def post(self, request):
        data = request.data

        challenge_id = data.get('id')
        is_confirm = data.get('is_confirm')
        comment = data.get('comment')

        send_type = Challenge.APPROVE

        if not is_confirm:
            send_type = Challenge.REJECT

        qs = self.queryset.filter(id=challenge_id)

        with transaction.atomic():

            qs.update(
                send_type=send_type,
                comment=comment
            )

        return request.send_info('INF_018')


@permission_classes([IsAuthenticated])
class StudentHomeworkListAPIView(
    mixins.ListModelMixin,
    mixins.DestroyModelMixin,
    mixins.UpdateModelMixin,
    generics.GenericAPIView,
    mixins.RetrieveModelMixin
):

    queryset = Lesson_assignment_student.objects.all().order_by('created_at')
    serializer_class = LessonAssignmentAssigmentListSerializer

    def get(self, request):
        student = request.query_params.get('student')
        assignment = request.query_params.get('assignment')

        user = request.user
        teacher = Teachers.objects.get(user=user)

        lesson_year, lesson_season = get_active_year_season()

        self.queryset = self.queryset.filter(assignment=assignment)

        if student:
            self.queryset = self.queryset.filter(student=student)

        assignment_list = self.list(request).data

        return request.send_data(assignment_list)

    def put(self, request, pk=None):
        data = request.data
        instance = self.queryset.filter(id=pk).first()

        self.serializer_class = LessonAssignmentAssigmentSerializer

        if 'score' in data:
            data['score'] = float(data['score'])

        data['status'] = Lesson_assignment_student.CHECKED

        try:
            """ Нэгээр дүгнэхэд """
            serializer = self.get_serializer(instance, data=data)

            if serializer.is_valid(raise_exception=False):
                self.perform_update(serializer)
            else:
                print(serializer.errors)

                return request.send_error("ERR_003", 'Дүн оруулахад алдаа гарлаа')
        except Exception as e:
            print(e)
            return request.send_error("ERR_003", 'Дүн оруулахад алдаа гарлаа')

        return request.send_info("INF_002")



@permission_classes([IsAuthenticated])
class StudentHomeworkMultiEditAPIView(
    mixins.UpdateModelMixin,
    generics.GenericAPIView,
):

    queryset = Lesson_assignment_student.objects.all()
    serializer_class = LessonAssignmentAssigmentSerializer

    def put(self, request):
        """ Даалгаварт олноор үнэлгээ өгөхөд """

        data = request.data

        score = float(data['score'])
        students = data.get('students')
        assignment = data.get('assignment')

        all_update_datas = list()

        try:
            if students and len(students) > 0:
                """ Олноор дүгнэхэд """
                student_assignment_qs = Lesson_assignment_student.objects.filter(student__in=students, assignment=assignment)

                if student_assignment_qs:
                    for student_assignment in student_assignment_qs:
                        student_assignment.score = score
                        student_assignment.status = Lesson_assignment_student.CHECKED
                        all_update_datas.append(student_assignment)

                if len(all_update_datas) > 0:
                    Lesson_assignment_student.objects.bulk_update(all_update_datas, ['score', 'status'])

        except Exception as e:
            print(e)
            return request.send_error("ERR_003", 'Дүн оруулахад алдаа гарлаа')

        return request.send_info("INF_002")


@permission_classes([IsAuthenticated])
class HomeworkStudentsListAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
):

    queryset = Lesson_assignment_student.objects.all().order_by('created_at')
    serializer_class = LessonAssignmentAssigmentListSerializer

    def get(self, request):
        lesson = request.query_params.get('lesson')
        assignment = request.query_params.get('assignment')

        user = request.user
        teacher = Teachers.objects.get(user=user, action_status=Teachers.APPROVED)
        filters = dict()

        lesson_year, lesson_season = get_active_year_season()

        self.queryset = self.queryset.filter(
            assignment__lesson_material__lesson=lesson,
            assignment__lesson_material__teacher=teacher.id
        )

        lesson_data = get_lesson_choice_student(lesson, teacher.id, '', lesson_year, lesson_season)

        if assignment:
            filters['assignment'] = assignment

        students = (
            Student
            .objects
            .filter(
                id__in=lesson_data
            )
            .annotate(
                unelgee=Subquery(
                    Lesson_assignment_student.objects.filter(student=OuterRef('id'), assignment__lesson_material__lesson=lesson, **filters).values('score')[:1]
                ),
                homework_status=Subquery(
                    Lesson_assignment_student.objects.filter(student=OuterRef('id'), assignment__lesson_material__lesson=lesson, **filters).values('status')[:1]
                ),
            )
            .values()
        )

        assignment_list = self.list(request).data

        return request.send_data({
            'students': list(students),
            'assignment': assignment_list
        })


@permission_classes([IsAuthenticated])
class LessonsTeacher(
    generics.GenericAPIView,
):
    ''' Тухайн багшийн зааж байгаа хичээл  '''

    def get(self, request, pk=None):
        stype = request.query_params.get('stype')

        user = request.user

        teacher = get_object_or_404(Teachers, user_id=user, action_status=Teachers.APPROVED)

        year, season = get_active_year_season()

        lesson_ids = TimeTable.objects.filter(lesson_year=year, lesson_season=season, teacher=teacher).values_list('lesson', flat=True).distinct('lesson')

        if stype == 'one':
            sort_list = LessonStandart.objects.filter(id__in=lesson_ids).values('id', 'name').order_by('name')
        else:
            all_list = LessonStandart.objects.filter(id__in=lesson_ids).values('id', 'name', 'code', 'category__category_name', 'definition', 'updated_at', 'kredit')

            for lesson in all_list:
                is_active = False
                is_active_num = 0
                file = None

                lesson_id = lesson.get('id')
                updated = lesson.get('updated_at')

                lesson_teacher = Lesson_to_teacher.objects.filter(lesson=lesson_id, teacher=teacher).first()

                if lesson_teacher:
                    file = lesson_teacher.file

                    media_path = settings.MEDIA_URL + str(file)
                    domain = get_domain_url()

                    file_path = domain + media_path

                fixed_date = fix_format_date(updated)

                timetable_qs_count = TimeTable.objects.filter(lesson_year=year, lesson_season=season, lesson=lesson_id).count()

                if timetable_qs_count >= 1:
                    is_active = True
                    is_active_num = 1

                lesson_materials_qs = Lesson_materials.objects.filter(lesson=lesson_id, teacher=teacher, material_type=Lesson_materials.PPTX)
                material_count = lesson_materials_qs.count()

                lesson_title_plan_qs = Lesson_title_plan.objects.filter(lesson=lesson_id, lesson_type=Lesson_title_plan.LECT)
                lesson_title_count = lesson_title_plan_qs.count()

                lesson['active'] = is_active
                lesson['is_active_num'] = is_active_num
                lesson['updated_at'] = fixed_date
                lesson['file'] = file_path if file else ''
                lesson['count'] = material_count + lesson_title_count

            sort_list = sorted(list(all_list), key=lambda item: (-item['is_active_num']) )

        return request.send_data(list(sort_list))


@permission_classes([IsAuthenticated])
class LessonOneApiView(
    generics.GenericAPIView,
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin
):
    """ Хичээлийн стандарт мэдээлэл """

    queryset = LessonStandart.objects.all()
    serializer_class = LessonStandartSerializer

    def get(self, request, pk=None):

        school = request.query_params.get('school')
        if school:
            self.queryset = self.queryset.filter(school=school)

        if pk:
            datas = self.retrieve(request, pk).data
        else:
            datas = self.list(request).data

        return request.send_data(datas)



@permission_classes([IsAuthenticated])
class LessonKreditApiView(
    generics.GenericAPIView
):

    def get(self, request, pk=None):
        """ Нэг хичээлийн багц цагийн мэдээлэл авах """

        return_values = []
        qs_title_plan = Lesson_title_plan.objects.all()
        if pk:
            values = LessonStandart.objects.filter(id=pk).values('lecture_kr', 'seminar_kr', 'laborator_kr', 'practic_kr', 'biedaalt_kr')

            for value in values:
                obj = {}

                lk = value.get('lecture_kr')
                sem = value.get('seminar_kr')
                lab = value.get('laborator_kr')
                pr = value.get('practic_kr')
                bd = value.get('biedaalt_kr')

                if lk:
                    sedev_data = qs_title_plan.filter(lesson=pk, lesson_type=Lesson_title_plan.LECT).values()
                    obj['id'] = Lesson_title_plan.LECT
                    obj['name'] = 'Лекц'
                    obj['data'] = list(sedev_data)

                    return_values.append(obj)
                    obj = {}

                if sem:
                    sedev_data = qs_title_plan.filter(lesson=pk, lesson_type=Lesson_title_plan.SEM).values()
                    obj['id'] = Lesson_title_plan.SEM
                    obj['name'] = 'Семинар'
                    obj['data'] = list(sedev_data)

                    return_values.append(obj)
                    obj = {}

                if lab:
                    sedev_data = qs_title_plan.filter(lesson=pk, lesson_type=Lesson_title_plan.LAB).values()
                    obj['id'] = Lesson_title_plan.LAB
                    obj['name'] = 'Лаборатор'
                    obj['data'] = list(sedev_data)

                    return_values.append(obj)
                    obj = {}

                if pr:
                    sedev_data = qs_title_plan.filter(lesson=pk, lesson_type=Lesson_title_plan.PRACTIC).values()
                    obj['id'] = Lesson_title_plan.PRACTIC
                    obj['name'] = 'Практик'
                    obj['data'] = list(sedev_data)

                    return_values.append(obj)
                    obj = {}

                if bd:
                    sedev_data = qs_title_plan.filter(lesson=pk, lesson_type=Lesson_title_plan.BIY_DAALT).values()
                    obj['id'] = Lesson_title_plan.BIY_DAALT
                    obj['name'] = 'Бие даалт'
                    obj['data'] = list(sedev_data)

                    return_values.append(obj)
                    obj = {}

        return request.send_data(list(return_values))



@permission_classes([IsAuthenticated])
class LessonSedevApiView(
    generics.GenericAPIView,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.ListModelMixin
):
    """ Хичээлийн сэдэвчилсэн төлөвлөгөө хадгалах """

    queryset = Lesson_title_plan.objects.all()
    serializer_class = LessonTitleSerializer

    def get(self, request, pk=None):

        if pk:
            self.queryset = self.queryset.filter(lesson=pk)

        all_list = self.list(request).data

        return request.send_data(all_list)

    def post(self, request, pk=None):

        errors = []
        datas = request.data
        lesson_type = request.query_params.get('type')

        is_created = False
        created_weeks = []
        for data in datas:
            lesson_type = data.get('lesson_type')
            data['lesson'] = pk

            week_obj = Lesson_title_plan.objects.filter(lesson=pk, week=data.get('week'), lesson_type=lesson_type).first()

            if week_obj:
                is_created = True

            if is_created:
                serializer = self.get_serializer(week_obj, data=data)
            else:
                serializer = self.get_serializer(data=data)

            if serializer.is_valid(raise_exception=False):

                with transaction.atomic():
                    try:
                       serializer.save()
                       created_weeks.append(data.get('week'))
                    except Exception as e:
                        print(e)

                        return request.send_error('ERR_002')
            else:
                for key in serializer.errors:

                    return_error = {
                        "field": key,
                        "msg": serializer.errors[key]
                    }

                    errors.append(return_error)

                return request.send_error("ERR_003", errors)

        # Хадгалсан датанаас устгах
        Lesson_title_plan.objects.filter(lesson=pk, lesson_type=int(lesson_type)).exclude(week__in=created_weeks).delete()
        return request.send_info('INF_013')


@permission_classes([IsAuthenticated])
class LessonAllApiView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    """ Тухайн тэнхимийн хичээлийг авах """

    queryset = LessonStandart.objects.all().order_by('name')
    serializer_class = LessonStandartSerializer

    def get(self, request):
        user_id = request.user
        teacher = Teachers.objects.get(user=user_id)
        department_id = teacher.salbar.id if teacher.salbar else None

        if teacher.salbar and department_id:
            self.queryset = self.queryset.filter(department=department_id)

        search_teacher = request.query_params.get('teacher')
        if search_teacher:
            lesson_teacher_ids = Lesson_to_teacher.objects.filter(teacher=search_teacher).values_list('lesson', flat=True)
            self.queryset = self.queryset.filter(id__in=lesson_teacher_ids)

        all_list = self.list(request).data

        return request.send_data(all_list)


@permission_classes([IsAuthenticated])
class LessonMaterialApiView(
    generics.GenericAPIView,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.ListModelMixin,
    mixins.DestroyModelMixin
):
    """ Хичээлийн материал хадгалах хэсэг """

    queryset = Lesson_materials.objects.all()
    serializer_class = LessonMaterialSerializer

    def get(self, request, pk):

        self.serializer_class = LessonMaterialListSerializer
        lesson_type = request.query_params.get('lesson_type')
        week = request.query_params.get('week')

        user = request.user
        teacher = get_object_or_404(Teachers, user_id=user, action_status=Teachers.APPROVED)

        lesson = LessonStandart.objects.filter(id=pk).first()

        if pk:
            self.queryset = self.queryset.filter(lesson=pk, teacher=teacher)

        if lesson_type:
            self.queryset = self.queryset.filter(lesson_type=lesson_type)

        if week:
            self.queryset = self.queryset.filter(week=int(week))

        all_list = self.list(request).data

        datas = {
            'datas': all_list,
            'name': lesson.code_name if lesson else ''
        }

        return request.send_data(datas)

    def post(self, request, pk=None):
        """ pk: Хичээлийн ID
            Файлтай хэсгүүдийг хадгалах хэсэг
        """

        datas = request.data.dict()

        files = request.FILES.getlist('files')

        user = request.user
        teacher = get_object_or_404(Teachers, user_id=user, action_status=Teachers.APPROVED)

        # Файл мэдээлэл устгах
        if datas.get('files'):
            datas = remove_key_from_dict(datas, 'files')

        material_type = datas.get('material_type')

        datas['teacher'] = teacher.id
        datas['material_type'] = int(material_type)

        datas = null_to_none(datas)

        with transaction.atomic():
            sid = transaction.savepoint()
            try:

                serializer =  self.get_serializer(data=datas)

                if serializer.is_valid(raise_exception=True):
                    serializer.save()

                    datas = serializer.data

                    # Материал ID
                    lesson_material_id = datas.get('id')

                    for file in files:
                        obj_datas = {}
                        obj_datas['material_id'] = lesson_material_id
                        obj_datas['file'] = file

                        # Оруулсан файл бүрийг хадгалах
                        Lesson_material_file.objects.create(**obj_datas)

            except Exception as e:

                transaction.savepoint_rollback(sid)
                print(e)
                return request.send_error('ERR_002')

        return request.send_info('INF_001')


    def put(self, request, pk=None):
        # Энд засах үйлдэл хийгдэнэ ирж байгааг датаг харж байгаад хийе
        return request.send_info('INF_002')

    def delete(self, request, pk=None):
        """ Хичээлийн материал устгах """

        with transaction.atomic():
            try:
                files = Lesson_material_file.objects.filter(material=pk).values_list('file', flat=True)

                # Файлуудыг устгах
                for file in files:

                    full_path = os.path.join(settings.MEDIA_ROOT, str(file))

                    if os.path.exists(full_path):
                        os.remove(full_path)

                # Хичээлийн материал устгах
                Lesson_material_file.objects.filter(material=pk).delete()

                # Даалгаврыг устгах
                lesson_assignment = Lesson_assignment.objects.filter(lesson_material=pk).values_list('id', flat=True)

                # Тухайн хичээл дээр холбоотой бүх даалгаврын хариуг устгана
                lesson_assignmend_students = Lesson_assignment_student.objects.filter(assignment__in=lesson_assignment).values_list('id', flat=True)

                # Тухайн хичээл дээр багш руу илгээсэн файлууд
                student_files = Lesson_assignment_student_file.objects.filter(student_assignment__in=lesson_assignmend_students).values_list('file', flat=True)
                for file in student_files:

                    full_path = os.path.join(settings.MEDIA_ROOT, str(file))

                    if os.path.exists(full_path):
                        os.remove(full_path)

                # Хичээлийн материал устгах
                Lesson_assignment_student.objects.filter(id__in=lesson_assignmend_students).delete()

                self.destroy(request, pk)

            except Exception as e:
                print(e)
                return request.send_error('ERR_002')

            return request.send_info('INF_003')


@permission_classes([IsAuthenticated])
class LessonMaterialGeneralApiView(
    generics.GenericAPIView,
    mixins.CreateModelMixin
):
    """ Хичээлийн ерөнхий мэдээлэл """

    queryset = Lesson_materials.objects.all()
    serializer_class = LessonMaterialSerializer

    def post(self, request, pk=None):

        data = request.data.dict()

        user = request.user
        teacher = get_object_or_404(Teachers, user_id=user, action_status=Teachers.APPROVED)

        data['teacher'] = teacher.id

        data = null_to_none(data)

        serializer = self.get_serializer(data=data)

        if serializer.is_valid(raise_exception=True):
            with transaction.atomic():
                self.perform_create(serializer)
        else:
            errors = []
            for key in serializer.errors:

                return_error = {
                    "field": key,
                    "msg": serializer.errors[key]
                }

                errors.append(return_error)

            if len(errors) > 0:
                return request.send_error("ERR_003", errors)

        return request.send_info("INF_001")



@permission_classes([IsAuthenticated])
class LessonEditorImage(
    generics.GenericAPIView
):
    """ Editor зураг хадгалах """

    def post(self, request, pk):

        data = request.data
        file = data.get('file')

        path = save_file(file, 'news_images', pk)[0]

        domain = get_domain_url()

        file_path = os.path.join(settings.MEDIA_URL, path)

        return_url = '{domain}{path}'.format(domain=domain, path=file_path)

        return request.send_data(return_url)



@permission_classes([IsAuthenticated])
class LessonMaterialAssignmentApiView(
    generics.GenericAPIView
):
    """ Даалгавар үүсгэх хэсэг"""

    queryset = Lesson_materials.objects.all()
    serializer_class = LessonMaterialSerializer

    def post(self, request, pk):
        """
            pk: Хичээлийн ID
        """

        datas = request.data.dict()

        # Assignment datas
        score = datas.get('score')
        start_date = datas.get('start_date')
        finish_date = datas.get('finish_date')

        files = request.FILES.getlist('files')

        user = request.user
        teacher = get_object_or_404(Teachers, user_id=user, action_status=Teachers.APPROVED)

        # Файл мэдээлэл устгах
        if len(files) > 0:
            datas = remove_key_from_dict(datas, 'files')

        material_type = datas.get('material_type')

        datas['teacher'] = teacher.id
        datas['material_type'] = int(material_type)

        datas = null_to_none(datas)

        with transaction.atomic():
            sid = transaction.savepoint()
            try:

                serializer =  self.get_serializer(data=datas)

                if serializer.is_valid(raise_exception=True):
                    serializer.save()

                    datas = serializer.data

                    # Материал ID
                    lesson_material_id = datas.get('id')

                    for file in files:
                        obj_datas = {}
                        obj_datas['material_id'] = lesson_material_id
                        obj_datas['file'] = file

                        # Оруулсан файл бүрийг хадгалах
                        Lesson_material_file.objects.create(**obj_datas)

                    # Хичээлийн даалгавар үүсгэх хэсэг
                    Lesson_assignment.objects.create(
                        lesson_material_id =  lesson_material_id,
                        score =  int(score),
                        start_date = start_date,
                        finish_date = finish_date
                    )

            except Exception as e:

                transaction.savepoint_rollback(sid)

                print(e)
                return request.send_error('ERR_002')

        return request.send_info('INF_001')

@permission_classes([IsAuthenticated])
class LessonImage(
    generics.GenericAPIView
):
    """ Тухайн багш хичээлдээ зураг оруулах """

    queryset = Lesson_to_teacher.objects.all()
    serializer_class = LessonTeacherSerializer

    def post(self, request, pk=None):

        data = request.data.dict()

        user = request.user
        teacher = get_object_or_404(Teachers, user_id=user, action_status=Teachers.APPROVED)

        instance = self.queryset.filter(lesson=pk, teacher=teacher).first()

        data['teacher'] = teacher.id
        data['lesson'] = pk

        serializer = self.get_serializer(instance=instance, data=data)

        if serializer.is_valid(raise_exception=True):
            serializer.save()

        else:
            error_obj = []
            for key in serializer.errors:

                return_error = {
                    "field": key,
                    "msg": serializer.errors[key]
                }

                error_obj.append(return_error)

            if len(error_obj) > 0:
                return request.send_error("ERR_003", error_obj)

        return request.send_info("INF_002")



@permission_classes([IsAuthenticated])
class LessonMaterialSendApiView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    ''' Хичээлийн материал ХБА руу илгээх '''

    queryset = Lesson_materials.objects.all().order_by('created_at', 'week')
    serializer_class = LessonMaterialListSerializer

    def get(self, request, lesson=None):
        """ Хичээлийн бүр лекцийн материал авах """

        self.queryset = self.queryset.filter(lesson=lesson, material_type=Lesson_materials.PPTX)

        all_list = self.list(request).data

        return request.send_data(all_list)

    def post(self, request, lesson=None):

        # Checked хийсэн ids
        checked_ids = request.data

        queryset = Lesson_materials.objects.filter(lesson=lesson)

        send_ids = queryset.filter(send_type=Lesson_materials.SEND).values_list('id', flat=True)

        remove_sends =  list(set(send_ids) - set(checked_ids))

        qs = queryset.filter(id__in=checked_ids)
        remove_qs = queryset.filter(id__in=remove_sends)

        with transaction.atomic():
            try:
                qs.update(
                    send_type = Lesson_materials.SEND
                )

                remove_qs.update(
                    send_type = None
                )

            except Exception as e:
                print(e)
                return request.send_error("ERR_002")

        return request.send_info("INF_019")


@permission_classes([IsAuthenticated])
class LessonMaterialApproveApiView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    """ ХБА батлах хичээлийн материал """

    queryset = Lesson_materials.objects.filter(material_type=Lesson_materials.PPTX).order_by('created_at', 'teacher', 'week')
    serializer_class = LessonMaterialListSerializer

    pagination_class = CustomPagination

    filter_backends = [SearchFilter]
    search_fields = ['title', 'teacher__first_name', 'end_date', 'teacher__last_name', 'lesson__name', 'lesson__code', 'week']

    def get(self, request):

        user_id = request.user

        ctype = request.query_params.get('type')
        search_teacher = request.query_params.get('teacher')
        lesson = request.query_params.get('lesson')

        # Шалгалтын төрлөөр хайх
        if ctype:
            self.queryset = self.queryset.filter(send_type=int(ctype))

        teacher = Teachers.objects.get(user=user_id)

        department_id = teacher.salbar.id

        self.queryset = self.queryset.filter(teacher__salbar__id=department_id)

        teacher_ids = self.queryset.values_list('teacher', flat=True)

        self.queryset = self.queryset.filter(teacher__in=teacher_ids)

        if search_teacher:
            self.queryset = self.queryset.filter(teacher=search_teacher)

        if lesson:
            self.queryset = self.queryset.filter(lesson=lesson)


        all_list = self.list(request).data

        return request.send_data(all_list)

    def post(self, request):
        data = request.data

        challenge_id = data.get('id')
        is_confirm = data.get('is_confirm')
        comment = data.get('comment')

        send_type = Lesson_materials.APPROVE

        if not is_confirm:
            send_type = Lesson_materials.REJECT

        qs = self.queryset.filter(id=challenge_id)

        with transaction.atomic():

            qs.update(
                send_type=send_type,
                comment=comment
            )

        return request.send_info('INF_018')


@permission_classes([IsAuthenticated])
class LessonStandartGroupAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    """ Тухайн ангийн хичээлүүд """

    queryset = LessonStandart.objects.all()
    serializer_class = LessonStandartSerializer
    def get(self, request, group=None):

        group_obj = get_object_or_404(Group, pk=group)

        lessons = LearningPlan.objects.filter(profession=group_obj.profession).annotate(full_name=Concat("lesson__code", Value("-"), "lesson__name", output_field=CharField())).values('lesson__id', 'full_name').order_by('lesson_level', 'lesson__name')
        students = Student.objects.filter(group=group).annotate(full_name=Concat("last_name", Value(". "), "first_name", output_field=CharField())).values('id', 'code', 'full_name')

        return_datas = {
            'lessons': list(lessons),
            'students': list(students)
        }

        return request.send_data(return_datas)


@permission_classes([IsAuthenticated])
class CopyProfesisonAPIView(
    generics.GenericAPIView,
    mixins.CreateModelMixin
):

    ''' Өгөгдсөн хөтөлбөрийн сургалтын төлөвлөгөөний хичээлүүдийг өөр хөтөлбөрийн сургалтын төлөвлөгөөний хичээл рүү хуулах '''

    queryset = LearningPlan.objects.all()

    # Өгөгдлийн сан дээр үйлдэл хийхэд алдаа гарах үед алдаа гарах хүртэл хийгдсэн үйлдлүүдийг буцаах функц
    @transaction.atomic
    def put(self, request):

        data = request.data
        queryset = self.queryset

        # Хуулж авах хөтөлбөрийн сургалтын төлөвлөгөө
        main_learning_plan = queryset.filter(profession=data['copying_prof'])

        # Хуулж тавих хөтөлбөрийн сургалтын төлөвлөгөөний хичээл устгах
        queryset.filter(profession=data['chosen_prof']).delete()

        # Хуулж авах хөтөлбөрийн мэдээллийг авах
        chosen_prof = ProfessionDefinition.objects.get(id=data['chosen_prof'])

        # Хуулж тавих хөтөлбөрийн сургалтын төлөвлөгөөний хичээлийн объектийг хадгалах хүснэгт
        cloned_lessons = []

        # Хуулж тавих хөтөлбөрийн сургалтын төлөвлөгөөний хичээлийн объектийг үүсгэж хадгалах хүснэгт рүү хийх
        for value in main_learning_plan:
            cloned_lessons.append(
                LearningPlan(
                    lesson = value.lesson,
                    previous_lesson = value.previous_lesson,
                    group_lesson = value.group_lesson,
                    lesson_level = value.lesson_level,
                    lesson_type = value.lesson_type,
                    season = value.season,
                    is_check_score = value.is_check_score,
                    department = value.department,
                    school = value.school,
                    profession = chosen_prof
                )
            )

        # Хүснэгтэд хадгалсан объектуудыг өгөгдлийн санд хадгалах хэсэг
        queryset.bulk_create(cloned_lessons)

        # Хүсэлт амжилттай явагдсан үед амжилттай явагдсан тухай мэдээлэл буцаах хэсэг
        return request.send_info('INF_018')


@permission_classes([IsAuthenticated])
class ProfessionJustProfessionAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    '''
        Сургалтын төлөвлөгөөний хүснэгтэд хадгалагдсан мэргэжлүүдийг авах
    '''

    queryset = ProfessionDefinition.objects.all()
    serializer_class = ProfessionDefinitionJustProfessionSerializer

    def get(self, request):

        # Сургуулиар хайх хэсэг
        school = request.query_params.get('school')
        if school:
            self.queryset = self.queryset.filter(school=school)

        profession_data = self.list(request).data

        return request.send_data(profession_data)


@permission_classes([IsAuthenticated])
class ProfessionPosterFile(
    generics.GenericAPIView,
    mixins.UpdateModelMixin
):
    """ Постер зураг"""
    queryset = ProfessionDefinition.objects.all()
    serializer_class = ProfessionDefinitionSerializer

    def post(self, request):
        datas = request.data
        profession = datas.get('profession')
        file = request.FILES.get('file')

        datas = {
            'poster_image': file
        }

        instance = ProfessionDefinition.objects.get(id=profession)
        with transaction.atomic():
            serializer =  self.get_serializer(instance, datas, partial=True)
            if serializer.is_valid(raise_exception=False):
                self.perform_update(serializer)
            else:
                return request.send_error_valid(serializer.errors)

        return request.send_info('INF_001')

    def delete(self, request, pk=None):
        ''' Poster зураг устгах '''

        with transaction.atomic():
            instance = ProfessionDefinition.objects.get(id=pk)
            instance.poster_image = None
            instance.save()

        return request.send_info('INF_003')

class ProfessionAPIView(
    generics.GenericAPIView
):
    """ Хөтөлбөрийн багаас хамаарч мэргэжлийн жагсаалт авах нь  """

    queryset = ProfessionDefinition.objects.all()

    def get(self, request, depId=None):

        if depId:
            dep_list = self.queryset.filter(department_id=depId).values("id", "name")
            return request.send_data(list(dep_list))

class GroupAPIView(
    generics.GenericAPIView
):
    """ Мэргэжлээс хамаарч ангийн жагсаалт авах нь """

    queryset = Group.objects.all()

    def get(self, request, profId=None):

        if profId:
            group_list = self.queryset.filter(profession_id=profId).values("id", "name")
            return request.send_data(list(group_list))

class SubOrgDepartListAPIView(
    generics.GenericAPIView,
):
    """ Сургуулиас хамаарч хөтөлбөрийн багийн жагсаалт авах нь """

    queryset = Salbars.objects

    def get(self, request, sub_org=None):

        if sub_org:
            list_data = self.queryset.filter(sub_orgs_id=sub_org).values('id', 'name')
            return request.send_data(list(list_data))

# import pandas as pd
# import math

# csv_file = '/home/daria/Downloads/data-1721382723541.csv'

# # Read CSV file into a DataFrame
# df = pd.read_csv(csv_file)

# # Print each row in the DataFrame
# for index, row in df.iterrows():
#     row_data = row.to_dict()
#     print(row_data.get('question_number'))
#     if row_data.get('question_number') and not math.isnan(row_data.get('question_number')):
#         obj = PsychologicalTestQuestions.objects.filter(id=row_data.get('id')).update(question_number=int(row_data.get('question_number')))
#         print(obj)

@permission_classes([IsAuthenticated])
class QuestionsTitleAPIView(
    generics.GenericAPIView,
    APIView,
    mixins.ListModelMixin,
    mixins.DestroyModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin

):
    """ Шалгалтын асуултууд Гарчигаар
    """
    queryset = QuestionTitle.objects.all()
    serializer_class = dynamic_serializer(QuestionTitle, "__all__")

    def get(self, request, pk=None):

        title_id = request.query_params.get('titleId')
        if title_id:
            title_id = int(title_id)

        stype = request.query_params.get('stype')
        level = request.query_params.get('level')
        is_graduate = request.query_params.get('is_graduate')
        is_elselt = request.query_params.get('is_elselt')

        # # 0  Бүх асуулт
        if title_id == 0:
            if is_elselt == 'true':
                challenge_qs = ChallengeQuestions.objects.filter(is_admission=True)
            else:
                challenge_qs = ChallengeQuestions.objects.all()
        # -1  Сэдэвгүй асуултууд
        elif title_id == -1:
            if is_elselt == 'true':
                challenge_qs = ChallengeQuestions.objects.filter(title__isnull=True,is_admission=True)
            else:
                challenge_qs = ChallengeQuestions.objects.filter(Q(title__isnull=True))
        else:
            if is_graduate == 'true':
                challenge_qs = ChallengeQuestions.objects.filter(graduate_title=title_id)
            else:
                if is_elselt == 'true':
                    challenge_qs = ChallengeQuestions.objects.filter(title=title_id,title__is_admission=True)
                else:
                    challenge_qs = ChallengeQuestions.objects.filter(title=title_id)

        if stype:
            challenge_qs = challenge_qs.filter(kind=stype)
        if level:
            challenge_qs = challenge_qs.filter(level=level)

        ser = dynamic_serializer(ChallengeQuestions, "__all__", 1)
        data = ser(challenge_qs, many=True)
        count = challenge_qs.count()

        result = {
            "count": count,
            "results": data.data
        }
        return request.send_data(result)


    def post(self, request):
        user_id = request.user
        teacher = get_object_or_404(Teachers, user_id=user_id, action_status=Teachers.APPROVED)
        question_ids = []
        request_data = request.data

        if 'questions' in request_data:
            question_ids = request_data.pop("questions")

        request_data['created_by'] = teacher.id
        serializer = self.get_serializer(data=request_data)

        if serializer.is_valid(raise_exception=False):
            saved_obj =  serializer.save()

            if question_ids:
                questions_to_update = ChallengeQuestions.objects.filter(id__in=question_ids)

                for question in questions_to_update:
                    question.title.add(saved_obj)

            data = self.serializer_class(saved_obj).data

            return request.send_info("INF_001", data)

        return request.send_info("ERR_001")


    def put(self, request, pk=None):
        request_data = request.data
        question_ids = request.data.pop("questions")
        other_question_ids = request.data.pop("other_questions")
        qs = self.queryset.filter(id=pk).get()
        serializer = self.get_serializer(qs, data=request_data, partial=True)
        if serializer.is_valid(raise_exception=False):
            self.perform_update(serializer)
            questions_to_update = ChallengeQuestions.objects.filter(id__in=question_ids)
            other_questions = ChallengeQuestions.objects.filter(id__in=other_question_ids)
            for question in questions_to_update:
                question.title.add(qs)
            for question in other_questions:
                question.title.remove(qs)
            return request.send_info("INF_002")

        return request.send_info("ERR_001")


    def delete(self, request, pk=None):
        self.destroy(request, pk)
        return request.send_info("INF_003")


@permission_classes([IsAuthenticated])
class QuestionsTitleListAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
):
    """ Шалгалтын Гарчиг
    """
    queryset = QuestionTitle.objects.all()
    serializer_class = dynamic_serializer(QuestionTitle, "__all__")

    def get(self, request,pk):
        teacher = None

        # in admin WEB teacher ID not passed (0 passed means falsy value), so may be it supposed to get titles by all teachers
        if pk and pk != 0:
            teacher = Teachers.objects.filter(id=pk).first()
            self.queryset = self.queryset.filter(created_by=teacher)

        lesson = request.query_params.get('lesson')
        season = request.query_params.get('season')
        challenge_type = request.query_params.get('challengeType')

        if lesson:
            self.queryset = self.queryset.filter(lesson=lesson)

        if season == 'false':
            self.queryset = self.queryset.filter(is_season=False)
        else:
            self.queryset = self.queryset.filter(is_season=True)

        if challenge_type == f'{Challenge.ADMISSION}':
            self.queryset = self.queryset.filter(is_admission=True)

        question_sub = ChallengeQuestions.objects.filter(title=OuterRef('id')).values('title').annotate(count=Count('id')).values('count')
        data = self.queryset.annotate(question_count=Subquery(question_sub)).values("id", "name", 'lesson__name', 'lesson__code', 'lesson__id', 'question_count', 'is_open')

        return request.send_data(list(data))


@permission_classes([IsAuthenticated])
class QuestionsLevelListAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
):
    """ Шалгалтын түвшнээр групп хийв
    """

    queryset = QuestionTitle.objects.all().filter(is_season=True)
    def get(self, request):
        lesson = request.query_params.get('lesson')

        if lesson:
            self.queryset = self.queryset.filter(lesson=lesson)

        question_queryset = ChallengeQuestions.objects.filter(title__in=self.queryset)
        datas = (
            question_queryset
            .values('level')
            .annotate(question_count=Count('id'))
            .annotate(type_name=WithChoices(ChallengeQuestions.DIFFICULTY_LEVELS, 'level'))
            .values('type_name', 'question_count', 'level')
        )

        return request.send_data(list(datas))


class TestQuestionsAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    mixins.DestroyModelMixin
):
    """
        Асуултын хэсэг
    """

    queryset = ChallengeQuestions.objects.all().order_by('-created_at')
    serializer_class = ChallengeQuestionListSerializer

    pagination_class = CustomPagination

    filter_backends = [SearchFilter]
    search_fields = ['question', 'subject__title']

    def get(self, request, pk=None):

        lesson = request.query_params.get('lesson')
        subject = request.query_params.get('subject')

        user = request.user
        teacher = get_object_or_404(Teachers, user_id=user, action_status=Teachers.APPROVED)

        self.queryset = self.queryset.filter(created_by=teacher)

        if lesson:
            self.queryset = self.queryset.filter(subject__lesson=lesson)

        if subject:
            self.queryset = self.queryset.filter(subject=subject)

        if pk:
            row_data = self.retrieve(request, pk).data
            return request.send_data(row_data)

        all_list = self.list(request).data

        return request.send_data(all_list)

    def put(self, request, pk):

        request_data = request.data.dict()
        type = request.query_params.get('type')

        if type == "question":
            question_img = request_data['image']
            request_data = remove_key_from_dict(request_data, [ 'image'])
            with transaction.atomic():
                question_obj = ChallengeQuestions.objects.filter(id=pk).first()

                if isinstance(question_img, str) != True:
                    question_img_path = get_image_path(question_obj)

                    file_path = save_file(question_img, question_img_path)[0]

                    old_image = question_obj.image
                    question_obj.image = file_path
                    question_obj.save()
                    if old_image:
                        remove_folder(str(old_image))


                if isinstance(question_img, str) == True and question_img == '':
                    old_image = question_obj.image
                    question_img_path = get_image_path(question_obj)
                    # Хуучин зураг засах үедээ устгасан бол файл устгана.
                    question_obj.image = None
                    question_obj.save()
                    if old_image:
                        remove_folder(str(old_image))

                updated_question_rows = ChallengeQuestions.objects.filter(id=pk).update(
                    **request_data
                )
                data = None
                if updated_question_rows > 0:
                    updated_question = ChallengeQuestions.objects.filter(id=pk).first()
                    ser = dynamic_serializer(ChallengeQuestions, "__all__", 1)
                    data = ser(updated_question).data
                return request.send_info('INF_002', data)

        else:
            answer_img = request_data["image"]
            answer_id = request_data.get('id')
            request_data = remove_key_from_dict(request_data, ['image', 'id'])
            with transaction.atomic():
                answer_obj = QuestionChoices.objects.filter(id=answer_id).first()
                question_obj = ChallengeQuestions.objects.filter(id=pk).first()
                if isinstance(answer_img, str) != True:
                    answer_img_path = get_choice_image_path(answer_obj)
                    file_path = save_file(answer_img, answer_img_path)[0]
                    old_image = answer_obj.image
                    answer_obj.image = file_path
                    answer_obj.save()
                    if old_image:
                        remove_folder(str(old_image))

                # Delete image
                if isinstance(answer_img, str) == True and answer_img == '':
                    old_image = answer_obj.image
                    answer_img_path = get_choice_image_path(answer_obj)
                    answer_obj.image = None
                    answer_obj.save()
                    if old_image:
                        remove_folder(str(old_image))

                updated_rows = QuestionChoices.objects.filter(id=answer_id).update(**request_data)
                data = None
                if updated_rows > 0:
                    updated_answer = QuestionChoices.objects.filter(id=answer_id).first()
                    ser = dynamic_serializer(QuestionChoices, "__all__")
                    data = ser(updated_answer).data
                    print(data)
                return request.send_info('INF_002', data)


    def post(self, request):
        questions = request.POST.getlist('questions')
        files = request.FILES.getlist('files')
        request_data = request.data.dict()
        title = request_data.get('title')
        graduate_title = request_data.get('main_title')
        is_admission = request_data.get('is_admission')

        user = request.user
        teacher = Teachers.objects.filter(user_id=user).first()

        with transaction.atomic():
            sid = transaction.savepoint()
            try:
                # Асуултыг хадгалах хэсэг
                for question in questions:
                    question = json.loads(question)
                    qkind = question.get("kind")
                    question_img = question.get('image')
                    score = question.get('score') # Асуултын оноо
                    answers = question.get('answers')
                    question['created_by'] = teacher
                    question['yes_or_no'] = None
                    question['max_choice_count'] = 0
                    question['rating_max_count'] = score
                    if is_admission == 'true':
                        question['is_admission'] = True

                    for image in files:
                        if hasattr(image, "name") and question['image'] == image.name:
                            question_img = image

                    for ans in answers:
                        if ans['is_correct'] == True and ans['value'] == 'Тийм':
                            question['yes_or_no'] = True
                        if ans['is_correct'] == True and ans['value'] == 'Үгүй':
                            question['yes_or_no'] = False
                        if ans['is_correct'] == True:
                            question['max_choice_count'] += 1

                    # Асуултын сонголтууд
                    choices = answers

                    question = remove_key_from_dict(question, ['image','answers'])
                    question = null_to_none(question)

                    question_obj = ChallengeQuestions.objects.create(
                        **question
                    )

                    # Улирлын шалгалтынг хувьд
                    if title and title != '-1':
                        question_obj.title.add(QuestionTitle.objects.get(pk=title))

                    if graduate_title:
                        question_obj.graduate_title.add(QuestioSubTitle.objects.get(pk=graduate_title))

                    # Асуултанд зураг байвал хадгалах хэсэг
                    if question_img:
                        question_img_path = get_image_path(question_obj)

                        file_path = save_file(question_img, question_img_path)[0]

                        question_obj.image = file_path
                        question_obj.save()

                    choice_ids = list()

                    # Асуултын сонголтуудыг үүсгэх нь
                    if int(qkind) in [ChallengeQuestions.KIND_MULTI_CHOICE, ChallengeQuestions.KIND_ONE_CHOICE, ChallengeQuestions.KIND_BOOLEAN]:

                        # Олон сонголттой үед асуултын оноог хувааж тавина
                        if int(qkind) == ChallengeQuestions.KIND_MULTI_CHOICE:
                            max_choice_count = int(question.get('max_choice_count'))

                            score = float(score) / max_choice_count

                        for choice in choices:
                            choice['created_by'] = teacher

                            checked = choice.get('is_correct')

                            choice['score'] = score if checked else 0
                            choice_img = None
                            choice['choices'] = choice['value']
                            for image in files:
                                if hasattr(image, "name") and choice['image'] == image.name:
                                    choice_img = image

                            choice = remove_key_from_dict(choice, ['image', 'is_correct', 'value'])


                            choice_obj = QuestionChoices.objects.create(
                                **choice
                            )

                            # Асуултанд зураг байвал хадгалах хэсэг
                            if choice_img:
                                choice_img_path = get_choice_image_path(choice_obj)

                                file_path = save_file(choice_img, choice_img_path)[0]

                                choice_obj.image = file_path
                                choice_obj.save()

                            choice_ids.append(choice_obj.id)

                    question_obj.choices.set(choice_ids)

            except Exception as e:
                print(e)
                transaction.savepoint_rollback(sid)

                return request.send_error('ERR_002')

        return request.send_info('INF_001')


class TestQuestionsListAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    """ Шалгалт үүсгэхдээ асуулт сонгох хэсэг """

    queryset = ChallengeQuestions.objects.all()
    serializer_class = ChallengeQuestionListSerializer

    def get(self, request):
        questions_qs = ChallengeQuestions.objects.all()
        is_admission = request.query_params.get('is_admission')
        if is_admission == 'true':
            questions_qs = questions_qs.filter(is_admission=True)

        created_by = request.user
        if not created_by.is_superuser:
            teacher = Teachers.objects.filter(user=created_by).first()
            questions_qs = questions_qs.filter(created_by=teacher)

        ser = dynamic_serializer(ChallengeQuestions, "__all__", 1)
        data = ser(questions_qs, many=True).data
        return request.send_data(data)

class TeacherExaminationScheduleAPIView(
    mixins.RetrieveModelMixin,
    mixins.ListModelMixin,
    generics.GenericAPIView
):
    " Шалгалтын хуваарь "
    queryset = ExamTimeTable.objects.all()
    serializer_class = TeacherExamTimeTableSerializer

    def get(self, request):

        year = request.query_params.get('year')
        season = request.query_params.get('season')

        if year:
            self.queryset = self.queryset.filter(lesson_year=year)

        if season:
            self.queryset = self.queryset.filter(lesson_season=season)

        datas = self.list(request).data
        return request.send_data(datas)

class QuestionExcelAPIView(generics.GenericAPIView, mixins.ListModelMixin):
    KIND_MAP = {
        'Нэг сонголт': 1,
        'Олон сонголт': 2,
        'Үнэн, Худал сонголт': 3,
        'Үнэлгээ': 4,
        'Эссэ бичих': 5,
        'Богино нөхөх хариулт': 6,
        'Харгалзуулах, жиших': 7,
        'Тооцоолж бодох': 8,
        'Төсөл боловсруулах': 9,
        'Товч хариулт': 10
    }

    LEVEL_MAP = {
        'Түвшин-1': 1,
        'Түвшин-2': 2,
        'Түвшин-3': 3,
        'Түвшин-4': 4,
        'Түвшин-5': 5,
        'Түвшин-6': 6,
    }

    LEVEL_MAP_SEASON = {
        'Хөнгөн': 1,
        'Дунд': 2,
        'Хүнд': 3,
    }

    def post(self, request):
        user = request.user
        teacher = Teachers.objects.filter(user_id=user).first()
        excel_file = request.FILES.get('file')
        title = request.data.get('title')
        graduate_title = request.data.get('main_title')
        is_admission = request.data.get('is_admission')

        if not excel_file:
            return request.send_info("INF_002")

        # Load the workbook and get the active sheet
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active
        data = [row for row in sheet.iter_rows(values_only=True)]
        headers = data[0]
        data_dicts = [dict(zip(headers, row)) for row in data[1:]]

        for entry in data_dicts:
            self.process_entry(entry, teacher, graduate_title, is_admission)

        return request.send_info("INF_001")

    def process_entry(self, entry, teacher, graduate_title, is_admission):
        question = entry.get('Асуулт')

        if question in {"Жишээ 1","Жишээ 2","Жишээ 3", 'Жишээ 4', 'Жишээ 5', 'Жишээ 6', 'Жишээ 7', 'Жишээ 8', 'Жишээ 9'}:
            return

        level_str = entry.get('Асуултын түвшин')
        score = entry.get('Асуултын оноо')
        kind_str = entry.get('Асуултын төрөл')
        correct_answer_index = entry.get('Зөв хариулт')

        kind = self.KIND_MAP.get(kind_str)
        level = self.LEVEL_MAP.get(level_str)
        if graduate_title:
            level = self.LEVEL_MAP.get(level_str)
        else:
            level = self.LEVEL_MAP_SEASON.get(level_str)

        # max_choice зөв хариулт хэд байгаагаас
        max_choice_count = len(correct_answer_index.split()) if kind == 2 else 1

        # Challenge үүсгэх
        challenge = ChallengeQuestions(
            question=question,
            kind=kind,
            score=score,
            level=level,
            max_choice_count=max_choice_count,
            created_by=teacher,
            is_admission=str2bool(is_admission)
        )
        challenge.save()
        # Төгсөлтйин шалгалт бол
        if graduate_title:
            challenge.graduate_title.add(QuestioSubTitle.objects.get(pk=graduate_title))

        choice_keys = ['Хариулт 1', 'Хариулт 2', 'Хариулт 3', 'Хариулт 4', 'Хариулт 5', 'Хариулт 6']
        choice_ids = self.process_choices(entry, teacher, choice_keys, kind, score, correct_answer_index)

        challenge.choices.set(choice_ids)

    def process_choices(self, entry, teacher, choice_keys, kind, score, correct_answer_index):
        choice_ids = []

        if kind in {1, 3}:  # Single choice or true/false
            correct_answer_index = int(correct_answer_index)
            for index, key in enumerate(choice_keys):
                choice_text = entry.get(key)
                if choice_text:
                    is_correct = 1 if (index + 1) == correct_answer_index else 0
                    choice = QuestionChoices(choices=choice_text, created_by=teacher)
                    choice.score = score if is_correct else 0
                    choice.save()
                    choice_ids.append(choice.id)

        elif kind == 2:  # Multiple choice
            correct_answer_indices = {int(x) for x in correct_answer_index.split()}
            for index, key in enumerate(choice_keys):
                choice_text = entry.get(key)
                if choice_text:
                    is_correct = 1 if (index + 1) in correct_answer_indices else 0
                    choice = QuestionChoices(choices=choice_text, created_by=teacher)
                    choice.score = score / len(correct_answer_indices) if is_correct else 0
                    choice.save()
                    choice_ids.append(choice.id)

        return choice_ids

@permission_classes([IsAuthenticated])
class TestTeacherApiView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin
):
    """ Багшийн жагсаалт """

    queryset = Teachers.objects.all()
    serializer_class = TeacherNameSerializer

    pagination_class = CustomPagination

    filter_backends = [SearchFilter]
    search_fields = ['first_name', 'last_name']

    def get_queryset(self):
        "Багшийн мэдээллийг сургууль, Хөтөлбөрийн багаар харуулах "

        queryset = self.queryset
        teacher_queryset = queryset.all().values_list('user', flat=True)
        qs_employee_user = Employee.objects.filter(user_id__in=list(teacher_queryset), state=Employee.STATE_WORKING).values_list('user', flat=True)
        if qs_employee_user:
            queryset = queryset.filter(user_id__in = list(qs_employee_user))

        queryset = get_teacher_queryset()
        sub_org = self.request.query_params.get('sub_org')
        salbar = self.request.query_params.get('salbar')
        position = self.request.query_params.get('position')
        season = self.request.query_params.get('season')
        sorting = self.request.query_params.get('sorting')
        is_season = self.request.query_params.get('season')
        stype = self.request.query_params.get('type')

        # Бүрэлдэхүүн сургууль
        if sub_org:
            queryset = queryset.filter(sub_org=sub_org)

        # салбар, тэнхим
        if salbar:
            queryset = queryset.filter(salbar=salbar)

        # Албан тушаалаар хайх
        if position:
            user_ids = Employee.objects.filter(org_position=position, state=Employee.STATE_WORKING).values_list('user', flat=True)
            queryset = queryset.filter(user_id__in=user_ids)

        # Sort хийх үед ажиллана
        if sorting:
            if not isinstance(sorting, str):
                sorting = str(sorting)

            queryset = queryset.order_by(sorting)

        if is_season == 'true':
            if stype and stype == 'graduate':
                title_teachers = QuestionTitle.objects.filter(is_graduate=True).values_list('created_by', flat=True)
            else:
                title_teachers = QuestionTitle.objects.filter(is_season=True).values_list('created_by', flat=True)

            queryset = queryset.filter(id__in=title_teachers)
        else:
            title_ids = QuestionTitle.objects.filter(is_season=False).values_list('id', flat=True)
            teacher_ids = ChallengeQuestions.objects.filter(title__in=title_ids).values_list('created_by', flat=True).distinct()
            queryset = queryset.filter(id__in=teacher_ids)

        return queryset

    def get(self, request):
        " нийт багшийн жагсаалт"

        teach_info = self.list(request).data
        return request.send_data(teach_info)

@permission_classes([IsAuthenticated])
class TestLessonTeacherApiView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin
):
    """ Багшийн хичээлийн жагсаалт """
    pagination_class = CustomPagination
    filter_backends = [SearchFilter]
    search_fields = ['first_name', 'last_name']

    def get(self, request, pk):
        teacher = Teachers.objects.filter(id=pk).first()
        if not teacher:
            return request.send_error("ERR_002", "Багш олдсонгүй.")

        exam_type = request.query_params.get('exam_type')
        obj = ChallengeQuestions.objects.filter(created_by_id=teacher).prefetch_related('title').filter(title__is_season=exam_type == '1')
        titles = set()
        for q in obj:
            titles.update(q.title.all())
        serialized_data = QuestionTitleSerializer(titles, many=True).data

        return_data = {
            "count":len(serialized_data),
            "name":teacher.full_name,
            "data" :serialized_data
        }

        return request.send_data(return_data)


@permission_classes([IsAuthenticated])
class ChallengeSedevCountAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.DestroyModelMixin
):

    def post(self, request):
        data = request.data
        print(data)
        challenge_id = data.get("challenge")
        title = data.get("subject")
        level = data.get("level_of_question")
        subInfo = data.get("subInfo") # Төгсөлтийн шалгалт
        number_of_questions = data.get('number_questions')
        number_of_questions_percentage = data.get("number_questions_percentage")

        if title is not None and challenge_id is not None:
            if subInfo:
                lesson_title = QuestioSubTitle.objects.filter(id=subInfo).first()
            else:
                lesson_title = QuestionTitle.objects.filter(id=title).first()

            challenge = Challenge.objects.filter(id=challenge_id).first()

            if challenge and lesson_title:
                if subInfo:
                    challenge_questions = ChallengeQuestions.objects.filter(
                        graduate_title=lesson_title,
                        level=level,
                    )
                else:
                    if challenge.challenge_type not in [Challenge.SORIL1,Challenge.ADMISSION]:
                        level = challenge.level
                    challenge_questions = ChallengeQuestions.objects.filter(
                        title=lesson_title,
                        level=level,
                        title__lesson=challenge.lesson,
                        title__is_season=(challenge.challenge_type == Challenge.SEMESTR_EXAM)
                    )

                random_questions = None
                if number_of_questions:
                    if challenge.has_shuffle:
                        random_questions = challenge_questions.order_by('?')[:int(number_of_questions)]
                    else:
                        random_questions = challenge_questions.order_by('id')[:int(number_of_questions)]
                elif number_of_questions_percentage:
                    total_questions = challenge_questions.count()
                    question_count = int(total_questions * (int(number_of_questions_percentage) /  100))

                    if challenge.has_shuffle:
                        random_questions = challenge_questions.order_by('?')[:question_count]
                    else:
                        random_questions = challenge_questions.order_by('id')[:question_count]
                else:
                    if challenge.has_shuffle:
                        random_questions = challenge_questions.order_by('?')
                    else:
                        random_questions = challenge_questions.order_by('id')

                challenge.questions.add(*random_questions)
                challenge.save()

                return request.send_info("INF_001")
        return request.send_error("ERR_003", 'Шалгалтын сэдэв сонгоно уу.')

    def delete(self, request, pk=None):
        try:
            question_ids = request.query_params.get('questions')

            if not isinstance(question_ids,list):
                question_ids = question_ids.split(',')
            challenge = Challenge.objects.get(id=pk)
            question_objs = challenge.questions.filter(id__in=question_ids)
            challenge.questions.remove(*question_objs)
            challenge.save()
        except Exception:
            traceback.print_exc()
            return request.send_error("ERR_002")
        return request.send_info("INF_003")


@permission_classes([IsAuthenticated])
class ChallengeLevelCountAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.DestroyModelMixin
):

    def post(self, request):
        data = request.data
        challenge_id = data.get("challenge")
        level = data.get("subject")

        # Асуулт тоогоор оруулах
        number_of_questions = data.get('number_questions')

        # Асуулт хувиар оруулах
        number_of_questions_percentage = data.get("number_questions_percentage")
        challenge = Challenge.objects.filter(id=challenge_id).first()

        # NOTE Яг яахаа үл мэднэ
        # teacher_ids = TeacherScore.objects.filter(score_type__lesson_teacher__lesson=challenge.lesson, student__in=challenge.student.values_list('id', flat=True)).values_list('score_type__lesson_teacher__teacher', flat=True).distinct()
        if level is not None and challenge_id is not None:
            if challenge and level:
                challenge_questions = ChallengeQuestions.objects.filter(
                    level=level,
                    title__lesson=challenge.lesson,
                    title__is_season=True,
                )

                if number_of_questions:
                    random_questions = challenge_questions.order_by('?')[:int(number_of_questions)]

                elif number_of_questions_percentage:
                    total_questions = challenge.question_count or 0
                    question_count = int(total_questions * (int(number_of_questions_percentage) /  100))

                    # Бодогдож байгаа асуултын тоо 0-ээс бага байх үед
                    if question_count == 0:
                        question_count = total_questions

                    random_questions = challenge_questions.order_by('?')[:question_count]
                else:
                    random_questions = challenge_questions.order_by('?')

                challenge.questions.add(*random_questions)
                challenge.save()

                return request.send_info("INF_001")

        return request.send_error("ERR_003", 'Шалгалтын түвшин сонгоно уу.')

    def delete(self, request, pk=None):

        question = ChallengeQuestions.objects.filter(id=pk).first()
        challenges = Challenge.objects.filter(questions=question).first()
        if challenges:
            challenges.questions.remove(question)
            challenges.save()

        return request.send_info("INF_003")


@permission_classes([IsAuthenticated])
class ChallengeAddAdmissionUserAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin
):
    queryset = ElseltUser.objects.all()
    serializer_class = ElsegchSerializer

    pagination_class = CustomPagination

    filter_backends = [SearchFilter]
    search_fields = ['first_name', 'register', 'email', 'code', 'last_name', 'mobile']

    def get(self, request):
        challenge = request.query_params.get('challenge')
        challenge_student_all = Challenge.objects.get(id=challenge).elselt_user.all().values_list('id', flat=True)
        self.queryset = self.queryset.filter(id__in=list(challenge_student_all))
        datas = self.list(request).data
        return request.send_data(datas)

    @has_permission(must_permissions=['lms-exam-update'])
    def put(self, request):
        data = request.data
        challenge_id = data.get("challenge")
        student_code = data.get("student")

        if student_code is None:
            return request.send_error("ERR_003", 'Элсэгчийн код хоосон байна!')

        try:
            student = ElseltUser.objects.filter(register=student_code).first()

            if student:
                challenge = Challenge.objects.filter(id=challenge_id).first()

                if challenge:
                    check = challenge.elselt_user.filter(id=student.id)

                    if check:
                        return request.send_error("ERR_003",f'"{student_code}" кодтой элсэгч шалгалтанд үүссэн байна.')

                    challenge.elselt_user.add(student)
                    challenge.save()
                    return request.send_info("INF_002")
                else:
                    return request.send_error("ERR_002")
            else:
                return request.send_error("ERR_003",f'"{student_code}" кодтой элсэгч олдсонгүй!')
        except Exception as e:
            print(e)
            return request.send_error("ERR_002")

    @has_permission(must_permissions=['lms-exam-update'])
    def delete(self, request, pk, student):
        try:
            challenge = Challenge.objects.get(id=pk)
            student = ElseltUser.objects.get(pk=student)
            challenge.elselt_user.remove(student)
            challenge.save()

            return request.send_info("INF_003")

        except Exception as e:
            print(e)
            return request.send_error("ERR_002")


@permission_classes([IsAuthenticated])
class AdmissionChallengeAddKindAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
):

    queryset = Challenge.objects.all()
    serializer_class = ChallengeSerializer

    @has_permission(must_permissions=['lms-exam-update'])
    def put(self, request, pk=None):
        datas = request.data
        select_ids = datas.get('admission')

        if not select_ids:
            return request.send_error("ERR_003", 'Элсэлтийг сонгоно уу!')
        qs_admission_user = ElseltUser.objects.filter(
            admissionuserprofession__profession__admission__is_store=False,
            admissionuserprofession__state=AdmissionUserProfession.STATE_SEND
        )

        qs_admission_user = qs_admission_user.filter(admissionuserprofession__profession__admission__in=select_ids)

        try:
            with transaction.atomic():
                challenge = Challenge.objects.get(id=pk)

                for admission_user in qs_admission_user:
                    challenge.elselt_user.add(admission_user)
                challenge.save()
        except Exception:
            traceback.print_exc()
            return request.send_error("ERR_002")
        return request.send_info("INF_002")


@permission_classes([IsAuthenticated])
class ChallengeAddStudentAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin
):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer

    pagination_class = CustomPagination

    filter_backends = [SearchFilter]
    search_fields = ['code', 'first_name', 'last_name']

    def get(self, request):
        challenge = request.query_params.get('challenge')
        challenge_student_all = Challenge.objects.get(id=challenge).student.all().values_list('id', flat=True)

        self.queryset = self.queryset.filter(id__in=list(challenge_student_all))

        datas = self.list(request).data
        return request.send_data(datas)

    def put(self, request):
        data = request.data
        challenge_id = data.get("challenge")
        student_code = data.get("student")

        if student_code is None:
            return request.send_error("ERR_003", 'Оюутны код хоосон байна!')

        try:
            student = Student.objects.filter(code=student_code).first()

            if student:
                challenge = Challenge.objects.filter(id=challenge_id).first()

                if challenge:
                    check = challenge.student.filter(id=student.id)

                    if check:
                        return request.send_error("ERR_003",f'"{student_code}" кодтой оюутан шалгалтанд үүссэн байна.')

                    challenge.student.add(student)
                    challenge.save()
                    return request.send_info("INF_002")
                else:
                    return request.send_error("ERR_002")
            else:
                return request.send_error("ERR_003",f'"{student_code}" кодтой оюутан олдсонгүй!')
        except Exception as e:
            print(e)
            return request.send_error("ERR_002")

    def delete(self, request, pk, student):
        try:
            challenge = Challenge.objects.get(id=pk)
            student = Student.objects.get(pk=student)
            challenge.student.remove(student)
            challenge.save()

            return request.send_info("INF_003")

        except Exception as e:
            print(e)
            return request.send_error("ERR_002")


@permission_classes([IsAuthenticated])
class ChallengeAddStudentDeleteAPIView(
    generics.GenericAPIView,
):
    queryset = Challenge.objects.all()
    serializer_class = ChallengeQuestionListSerializer

    def post(self, request):
        """
            Асуултын олноор устгах хэсэг
        """

        data = request.data
        deleted_ids = data.get('ids')
        challenge_id = data.get('challenge_id')

        challenge_qs = Challenge.objects.get(pk=challenge_id)
        question = ChallengeQuestions.objects.filter(id__in=deleted_ids)
        if challenge_qs:
            challenge_qs.questions.remove(*question)
            challenge_qs.save()

        return request.send_info("INF_003")


@permission_classes([IsAuthenticated])
class ChallengeQuestionsAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
):
    queryset = ChallengeQuestions.objects.all()
    serializer_class = ChallengeQuestionListSerializer

    pagination_class = CustomPagination

    def get(self, request):
        challenge_id = request.query_params.get('id')
        challenge = Challenge.objects.filter(id=challenge_id).first()

        if challenge_id:
            all_data = challenge.questions.all().order_by('id')
            paginator = self.pagination_class()
            paginated_all_data = paginator.paginate_queryset(all_data, request, view=self)
            serializer = self.get_serializer(paginated_all_data, many=True)

            response_data = {
                "question_count": paginator.page.paginator.count,
                'next': paginator.get_next_link(),
                'previous': paginator.get_previous_link(),
                "questions": serializer.data,
                "challenge_question_count": challenge.question_count
            }
        return request.send_data(response_data)


@permission_classes([IsAuthenticated])
class TestQuestionsAllAPIView(
    mixins.RetrieveModelMixin,
    generics.GenericAPIView,
    mixins.ListModelMixin,
):
    "Шалгалтын асуулт"

    queryset = ChallengeQuestions.objects.all()
    serializer_class= ChallengeQuestionListSerializer

    def get(self, request, pk=None):

        if pk:
            test = Challenge.objects.get(id=pk)
            questions= test.questions.order_by('id')

        serializer = self.serializer_class(questions, many=True)
        data = serializer.data
        return request.send_data(data)


@permission_classes([IsAuthenticated])
class TestQuestionsDifficultyLevelsAPIView(
    generics.GenericAPIView,
):
    "Difficulty levels"

    queryset = ChallengeQuestions.objects.all()

    def get(self, request):
        data = [{'value': key, 'label': label} for key, label in self.queryset.model._meta.get_field('level').choices]

        return request.send_data(data)


@permission_classes([IsAuthenticated])
class ChallengeStudentsScoreAPIView(
    generics.GenericAPIView,
):

    @has_permission(must_permissions=['lms-exam-update'])
    def put(self, request):
        data = request.data
        exam = data

        if not exam:

            return request.send_error("ERR_002")

        challenge_students_qs = ChallengeStudents.objects.filter(challenge=exam)
        question_count_actual = ChallengeQuestions.objects.filter(challenge__id=exam).count()
        changed_values = {}

        # to update question count field (i am not sure why this field needed) if questions collection is changed. For example: to calculate proper percents values in reports (if this field is used)
        challenge_qs = Challenge.objects.filter(id=exam).first()

        if challenge_qs.question_count != question_count_actual:
            changed_values['Challenge'] = {
                'question_count': {
                    'old': challenge_qs.question_count,
                    'new': question_count_actual
                }
            }

            challenge_qs.question_count = question_count_actual
            challenge_qs.save()

        changed_values['ChallengeStudents'] = {}

        for challenge_student in challenge_students_qs:
            changed_values['ChallengeStudents'][f'{challenge_student.id}-{challenge_student.student.code}'] = {}

            if challenge_student.take_score != question_count_actual:
                changed_values['ChallengeStudents'][f'{challenge_student.id}-{challenge_student.student.code}']['take_score'] = {
                    'old': challenge_student.take_score,
                    'new': question_count_actual
                }

                # to get total possible score. 1 question counted as 1 point of score (may be summary of values of "QuestionChoices,score" field should be used instead, i am not sure)
                challenge_student.take_score = question_count_actual

            # region score update code
            if not challenge_student.answer:

                continue

            answers = json.loads(challenge_student.answer.replace("'", '"'))
            answers_choices = []

            # to optimize database usage by avoiding sql queries in loop. to do it collect all ids first and make single sql query later
            if answers:
                for value in answers.values():
                    answers_choices.append(value)

            # to detect attempt. attempt counted only if any choice was given (i am not sure is it correct detect way or not)
            if answers_choices:
                if challenge_student.tried != True:
                    changed_values['ChallengeStudents'][f'{challenge_student.id}-{challenge_student.student.code}']['tried'] = {
                        'old': challenge_student.tried,
                        'new': True
                    }

                    challenge_student.tried = True

            # to get correct answers count. any value of "QuestionChoices,score" greater than 0 counted as 1 point of score (may be summary of values of "QuestionChoices,score" field should be used instead, i am not sure)
            score = QuestionChoices.objects.filter(id__in=answers_choices, score__gt=0).count()

            if challenge_student.score != score:
                changed_values['ChallengeStudents'][f'{challenge_student.id}-{challenge_student.student.code}']['old_score'] = {
                    'old': challenge_student.old_score,
                    'new': challenge_student.score
                }

                changed_values['ChallengeStudents'][f'{challenge_student.id}-{challenge_student.student.code}']['score'] = {
                    'old': challenge_student.score,
                    'new': score
                }

                challenge_student.old_score = challenge_student.score
                challenge_student.score = score
            # endregion

            challenge_student.save()

        return request.send_data(changed_values)


@permission_classes([IsAuthenticated])
class ChallengeReportAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
):
    "Challenge report"

    pagination_class = CustomPagination

    filter_backends = [SearchFilter]

    @staticmethod
    def parse_answers(json_data, challenge_id):
        answers = []

        try:
            json_data = json_data.replace("'", '"')

            # for json.loads() to make correct JSON standart (true, false, и null). E.g. True False not standart so error occurs (wrong case)
            json_data = json_data.replace('": True', '": true')
            json_data = json_data.replace('": False', '": false')

            answer_json = json.loads(json_data)
            # print('json_data no error in data: ', json_data)

            # to get last title because title field is manytomany type
            questions = (
                ChallengeQuestions.objects
                .filter(challenge__id=challenge_id)
                .order_by('id')
                .annotate(last_title_id=Max('title__id'))
                .filter(title__id=F('last_title_id'))
                .values('id', 'question', 'title__name')
            )

            for question in questions:
                choice_id = answer_json.get(str(question.get('id')))
                is_right = False

                if choice_id:
                    # to process boolean values. because sometimes answer_json.get(<id>) has boolean value and sometimes it has choice_id (integer) value
                    if choice_id == True:
                        is_right = True

                    else:
                        choice_obj = QuestionChoices.objects.get(id=choice_id)

                        if choice_obj.score > 0:
                            is_right = True

                answers.append({
                    'question_id': question.get('id'),
                    'question_text': question.get('question'),
                    'is_answered_right': is_right,
                    'choice_id': choice_id,
                    'question_title': question.get('title__name'),
                })

        except Exception:
            traceback.print_exc()
            print('json_data error in data: ', json_data)

        return answers

    @staticmethod
    def get_question_stats(exam_results):
        # to collect questions reliability stats
        question_stats = {}

        for res in exam_results:
            question_id = res["question_id"]

            if question_id not in question_stats:
                question_stats[question_id] = {"correct": 0, "total": 0, 'question_text': res['question_text']}

            question_stats[question_id]["total"] += 1

            if res["is_answered_right"]:
                question_stats[question_id]["correct"] += 1

        return question_stats

    def get_queryset(self):
        request = self.request
        report_type = request.query_params.get('report_type')
        school_id = request.query_params.get('school')
        lesson_year = request.query_params.get('lesson_year')
        lesson_season = request.query_params.get('lesson_season')

        if not report_type:
            return None

        queryset = None
        exam_qs = ExamTimeTable.objects.all()
        if lesson_year and lesson_season:
            exam_qs = ExamTimeTable.objects.filter(lesson_year=lesson_year, lesson_season=lesson_season)
        # if school_id:
        #     exam_ids = Exam_to_group.objects.filter(exam__lesson_year=lesson_year, exam__lesson_season=lesson_season, group__school=school_id).values_list('exam', flat=True)
        #     exam_qs = exam_qs.filter(id__in=exam_ids)

        lesson_ids = exam_qs.values_list('lesson', flat=True)
        if report_type in ['students', 'students_detail']:
            queryset = TeacherScore.objects.filter(
                lesson_year=lesson_year,
                lesson_season_id=lesson_season,
                score__gt=0,
                score_type__lesson_teacher__lesson__in=lesson_ids
            )

        if report_type == 'students':
            queryset = (

                # this is "group by" part of sql query
                queryset.values(
                    student_idnum=F('student_id'),
                    student_first_name=F('student__first_name'),
                    student_last_name=F('student__last_name'),
                    student_code=F('student__code'),
                    group_name=F('student__group__name')

                ).annotate(
                    scored_lesson_count = Count('score_type__lesson_teacher__lesson__name', distinct=True),
                    exam_type_scored_lesson_count = Count(
                        Case(
                            When(
                                score_type__score_type=Lesson_teacher_scoretype.SHALGALT_ONOO,
                                then='score_type__lesson_teacher__lesson__name'
                            )
                        ),
                        distinct=True
                    ),
                    success_scored_lesson_count=Count(
                        Case(
                            When(
                                Q(score_type__score_type=Lesson_teacher_scoretype.SHALGALT_ONOO) &
                                Q(score__gte=18),
                                then='score_type__lesson_teacher__lesson__name'
                            )
                        ),
                        distinct=True
                    ),
                    failed_scored_lesson_count=F('scored_lesson_count') - F('success_scored_lesson_count')
                ).order_by('-failed_scored_lesson_count')
            )

        elif report_type == 'students_detail':
            student_id = request.query_params.get('student')
            lesson_ids = ExamTimeTable.objects.filter(lesson_year=lesson_year, lesson_season_id=lesson_season).values_list('lesson', flat=True)

            if student_id:
                queryset = (
                    queryset.filter(student_id=student_id, score_type__lesson_teacher__lesson__in=lesson_ids).values(
                        student_first_name=F('student__first_name'),
                        student_last_name=F('student__last_name'),
                        student_code=F('student__code'),
                        lesson_name=F('score_type__lesson_teacher__lesson__name')
                    ).annotate(
                        exam_score = Max(Case(
                            When(
                                score_type__score_type=Lesson_teacher_scoretype.SHALGALT_ONOO,
                                then='score',
                            ),
                                default=Value(0), output_field=FloatField(),
                            )
                        ),
                        exam_teacher_first_name = Max(Case(
                            When(
                                score_type__score_type=Lesson_teacher_scoretype.SHALGALT_ONOO,
                                then='score_type__lesson_teacher__teacher__first_name'
                            ))
                        ),
                        exam_teacher_last_name = Max(Case(
                            When(
                                score_type__score_type=Lesson_teacher_scoretype.SHALGALT_ONOO,
                                then='score_type__lesson_teacher__teacher__last_name'
                            ))
                        ),
                        teach_score = Max(Case(
                            When(
                                score_type__score_type=Lesson_teacher_scoretype.BUSAD,
                                then='score'
                            ))
                        ),
                        teach_teacher_first_name = Max(Case(
                            When(
                                score_type__score_type=Lesson_teacher_scoretype.BUSAD,
                                then='score_type__lesson_teacher__teacher__first_name'
                            ))
                        ),
                        teach_teacher_last_name = Max(Case(
                            When(
                                score_type__score_type=Lesson_teacher_scoretype.BUSAD,
                                then='score_type__lesson_teacher__teacher__last_name'
                            ))
                        ),
                    ).order_by('-exam_score')
                )

        else:
            queryset = ChallengeStudents.objects.order_by('-score').filter(
                challenge__challenge_type=Challenge.SEMESTR_EXAM
            )

            if lesson_year and lesson_season:
                queryset = queryset.filter(
                    challenge__lesson_year=lesson_year,
                    challenge__lesson_season=lesson_season
                )

            exam = request.query_params.get('exam')

            if exam:
                queryset = queryset.filter(challenge=exam)
            else:
                return None

        group = request.query_params.get('group')

        if group:
            queryset = queryset.filter(student__group=group)

        profession = request.query_params.get('profession')

        if profession:
            queryset = queryset.filter(student__group__profession=profession)

        # report1. for datatable with pagination
        if report_type == 'dt_reliability':
            answers = []

            for obj in queryset:
                if not obj.answer:

                    continue

                answers.extend(self.parse_answers(obj.answer, obj.challenge.id))

            answers_stats = self.get_question_stats(answers)
            questions_qs = ChallengeQuestions.objects.filter(challenge__id=exam)

            for question_obj in questions_qs:
                stats = answers_stats.get(question_obj.id)

                if stats:
                    question_obj.reliability = ((stats['correct'] * 100) / stats['total']) if stats['total'] > 0 else 0

            queryset = questions_qs

        elif (
            report_type == 'groups' or
            report_type == 'professions'
        ):
            assessments = Score.objects.all().values('score_min','score_max','assesment')
            assessment_dict = {}

            for assessment in assessments:
                assesment_value = assessment['assesment']
                score_min = assessment['score_min']
                score_max = assessment['score_max']

                # to get real min max values if assesment letter has duplications
                if assesment_value in assessment_dict:
                    if assessment_dict[assesment_value]['score_min'] > score_min:
                        assessment_dict[assesment_value]['score_min'] = score_min

                    if assessment_dict[assesment_value]['score_max'] < score_max:
                        assessment_dict[assesment_value]['score_max'] = score_max

                else:
                    assessment_dict[assesment_value] = {
                        'score_min': score_min,
                        'score_max': score_max
                    }

            queryset = (
                queryset
                    .order_by() # to remove above sortings because it conflicts with "group by"
                    .select_related('student')
            )

            if report_type == 'groups':
                queryset = (
                    queryset
                        .prefetch_related('student__group')
                        .annotate(
                            group_name=F('student__group__name'),
                            score_percentage=Case(
                                When(take_score__gt=0, then=(F('score') * 100 / F('take_score'))),
                                default=Value(0),
                                output_field=FloatField()
                            ),
                        )
                        .values('group_name') # to group students by group_name
                )

            elif report_type == 'professions':
                queryset = (
                    queryset
                        .prefetch_related('student__group__profession')
                        .annotate(
                            profession_name=F('student__group__profession__name'),
                            score_percentage=Case(
                                When(take_score__gt=0, then=(F('score') * 100 / F('take_score'))),
                                default=Value(0),
                                output_field=FloatField()
                            ),
                        )
                        .values('profession_name') # to group students by profession_name
                )

            queryset = (
                queryset
                    .annotate(
                        student_count=Count('student', distinct=True),
                        A2_count=Count(
                            Case(
                                When(
                                    score_percentage__gte=assessment_dict.get('A+').get('score_min'),
                                    score_percentage__lte=assessment_dict.get('A+').get('score_max'),
                                    then=Value(1)),
                                output_field=IntegerField()
                            )
                        ),
                        A_count=Count(
                            Case(
                                When(
                                    score_percentage__gte=assessment_dict.get('A').get('score_min'),
                                    score_percentage__lte=assessment_dict.get('A').get('score_max'),
                                    then=Value(1)),
                                output_field=IntegerField()
                            )
                        ),
                        B2_count=Count(
                            Case(
                                When(
                                    score_percentage__gte=assessment_dict.get('B+').get('score_min'),
                                    score_percentage__lte=assessment_dict.get('B+').get('score_max'),
                                    then=Value(1)),
                                output_field=IntegerField()
                            )
                        ),
                        B_count=Count(
                            Case(
                                When(
                                    score_percentage__gte=assessment_dict.get('B').get('score_min'),
                                    score_percentage__lte=assessment_dict.get('B').get('score_max'),
                                    then=Value(1)),
                                output_field=IntegerField()
                            )
                        ),
                        C2_count=Count(
                            Case(
                                When(
                                    score_percentage__gte=assessment_dict.get('C+').get('score_min'),
                                    score_percentage__lte=assessment_dict.get('C+').get('score_max'),
                                    then=Value(1)),
                                output_field=IntegerField()
                            )
                        ),
                        C_count=Count(
                            Case(
                                When(
                                    score_percentage__gte=assessment_dict.get('C').get('score_min'),
                                    score_percentage__lte=assessment_dict.get('C').get('score_max'),
                                    then=Value(1)),
                                output_field=IntegerField()
                            )
                        ),
                        D_count=Count(
                            Case(
                                When(
                                    score_percentage__gte=assessment_dict.get('D').get('score_min'),
                                    score_percentage__lte=assessment_dict.get('D').get('score_max'),
                                    then=Value(1)),
                                output_field=IntegerField()
                            )
                        ),
                        F_count=Count(
                            Case(
                                When(
                                    score_percentage__gte=assessment_dict.get('F').get('score_min'),
                                    score_percentage__lte=assessment_dict.get('F').get('score_max'),
                                    then=Value(1)),
                                output_field=IntegerField()
                            )
                        )
                    )
            )

        # for reports where pagination is required
        if report_type in ['students', 'students_detail', 'dt_reliability', 'report4', 'groups', 'professions']:
            sorting = self.request.query_params.get('sorting')

            # Sort хийх үед ажиллана
            if sorting:
                if not isinstance(sorting, str):
                    sorting = str(sorting)

                queryset = queryset.order_by(sorting)

        return queryset

    def get(self, request):
        queryset = self.get_queryset()
        report_type = request.query_params.get('report_type')
        get_result = []

        if not queryset:
            return request.send_data(None)

        # report1. reliability
        if report_type == 'reliability':
            exam_results = []

            for obj in queryset:
                if not obj.answer:

                    continue

                exam_results.extend(self.parse_answers(obj.answer, obj.challenge.id))

            # questions reliability ranges
            questions_reliability_ranges = {
                "Маш хүнд": lambda question_reliability: question_reliability <= 20,
                "Хүндэвтэр": lambda question_reliability: 21 <= question_reliability <= 40,
                "Дунд зэрэг": lambda question_reliability: 41 <= question_reliability <= 60,
                "Хялбар": lambda question_reliability: 61 <= question_reliability <= 80,
                "Маш хялбар": lambda question_reliability: question_reliability >= 81,
            }

            question_stats = self.get_question_stats(exam_results)

            # to calculate question reliability
            questions_reliability = []

            for question_id, stats in question_stats.items():
                if stats["total"] > 0:
                    question_reliability = (stats["correct"] / stats["total"]) * 100

                    questions_reliability.append({
                        'question_id': question_id,
                        'question_reliability': question_reliability,
                        'question_text': stats['question_text']
                    })

            # to group questions and their reliabilities by reliability ranges
            grouped_questions = {key: [] for key in questions_reliability_ranges}

            for item in questions_reliability:
                question_id = item.get('question_id')
                question_reliability = item.get('question_reliability')
                question_text = item.get('question_text')

                for range_name, condition in questions_reliability_ranges.items():
                    if condition(question_reliability):
                        grouped_questions[range_name].append({"question_id": question_id, 'question_text': question_text, "question_reliability": question_reliability})

                        break

            rechart_data = []
            total_questions_count = len(question_stats)

            for key, questions in grouped_questions.items():
                # to build dict for recharts format
                rechart_data.append({
                    "questions_reliability_name": key,
                    "questions_count_percent": (len(questions) * 100 / total_questions_count) if total_questions_count else 0
                })

            get_result = rechart_data

        # region for datatable with pagination
        # report1
        elif report_type == 'dt_reliability':
            self.serializer_class = ChallengeQuestionsAnswersSerializer
            self.search_fields = ['question']

        # region report2 students
        elif report_type == 'students':
            self.serializer_class = ChallengeReport2StudentsSerializer
            self.search_fields = ['student__code', 'student__first_name']

        elif report_type == 'students_detail':
            self.serializer_class = ChallengeReport2StudentsDetailSerializer
            self.search_fields = ['lesson__code', 'lesson__name']
        # endregion report2 students

        # report2 groups
        elif report_type == 'groups':
            self.serializer_class = ChallengeGroupsSerializer

        # report2 professions
        elif report_type == 'professions':
            self.serializer_class = ChallengeProfessionsSerializer

        elif report_type == 'report4':
            self.serializer_class = ChallengeReport4Serializer

        elif report_type == 'report4-1':
            answers = []
            first_student_answers = []

            for index, obj in enumerate(queryset):
                if not obj.answer:

                    continue

                answers.extend(self.parse_answers(obj.answer, obj.challenge.id))

                if index == 0:
                    # because we need only one row data for header
                    first_student_answers = answers.copy()

            question_stats = self.get_question_stats(answers)

            get_result = {
                'questions': first_student_answers,
                'questions_summary': question_stats
            }

        if report_type in ['students', 'students_detail', 'dt_reliability', 'report4', 'groups', 'professions']:
            get_result = self.list(request).data
        # endregion for datatable with pagination

        return request.send_data(get_result)


@permission_classes([IsAuthenticated])
class ChallengeDetailApiView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
):
    queryset = ChallengeStudents.objects.all().order_by('score')
    serializer_class = ChallengeStudentsSerializer

    pagination_class = CustomPagination

    filter_backends = [SearchFilter]
    search_fields = ['code', 'first_name', 'last_name']

    def get(self, request):

        #Тухайн шалгалтын id
        test_id = request.query_params.get('test_id')
        #Энэ шалгалтыг өгсөн оюутануудын id-г student_list-д хадгална
        student_list = self.queryset.filter(challenge=test_id).values_list('student__id', flat=True)
        #Нийт оюутануудаас шалгалт өгсөн оюутануудыг авна
        students_qs = Student.objects.filter(id__in=student_list)
        students_qs = self.filter_queryset(students_qs)
        #Энийг мэддэг болоод бичнэ. Гэхдээ queryset-ээ serializer руу дамжуулаад энэ qs-ээ тохирсон дата-г аваад байх шиг байна.
        #Бас context-оор test_id-г дамжуулж байна.
        student_data = StudentChallengeSerializer(students_qs, many=True, context={'test_id': test_id}).data

        for student in student_data:
            challenge_student = ChallengeStudents.objects.filter(student__id=student['id'], challenge=test_id).first()
            if challenge_student:
                student['is_not_exam_failed'] = (challenge_student.score or 0) >= 18

        return request.send_data(student_data)


@permission_classes([IsAuthenticated])
class ChallengeTestDetailApiView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin
):
    queryset = Challenge.objects.all()
    serializer_class = ChallengeDetailSerializer

    def get(self, request, pk):

        test_info = self.retrieve(request, pk).data
        return request.send_data(test_info)


# @permission_classes([IsAuthenticated])
class TestResultShowAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    """Шалгалтын оноо асуулт хариултыг харах API"""

    queryset = ChallengeQuestions.objects.all()
    serializer_class = ChallengeQuestionListSerializer

    def post(self, request):
        datas = request.data
        question_ids = []
        chosen_choices = []
        big_data = []
        return_data = []
        totalscore = 0
        if isinstance(datas, str):
            datas = ast.literal_eval(datas)

        if datas:
            for question_id, choice_id in datas.items():
                question_ids.append(int(question_id))

                if isinstance(choice_id, list):  # It's a list
                    chosen_choices.append([int(id) for id in choice_id if isinstance(id, int)])
                else:
                    chosen_choices.append(int(choice_id))

            # Process each question and choice combination
            for question_id, choice_ids in zip(question_ids, chosen_choices):
                queryset = ChallengeQuestions.objects.filter(id=question_id).first()
                if queryset:
                    serializer = self.serializer_class(queryset)
                    data = serializer.data

                    if isinstance(choice_ids, list):
                        # Multiple choices, handle each choice ID
                        data['chosen_choice'] = [int(id) for id in choice_ids]
                    else:
                        # Single choice
                        data['chosen_choice'] = int(choice_ids) if choice_ids not in [0, 1] else ('Тийм' if choice_ids == 1 else 'Үгүй')
                    if data['score']:
                        totalscore += data['score']

                    big_data.append(data)

        return_data = {
            'question': big_data,
            'total_score': totalscore
        }
        return request.send_data(return_data)


@permission_classes([IsAuthenticated])
class ChallengeAddInformationAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    mixins.RetrieveModelMixin,
):
    queryset = Challenge.objects.all().order_by("-created_at")
    serializer_class = ChallengeSerializer

    def get(self, request, pk=None):
        data = self.retrieve(request, pk).data
        return request.send_data(data)

    def post(self, request):
        data = request.data
        user = request.user
        season = request.query_params.get('season')
        lesson_year, lesson_season = get_active_year_season()
        teacher = get_object_or_404(Teachers, user_id=user, action_status=Teachers.APPROVED)
        data['created_by'] = teacher.id if teacher else None

        if 'is_repeat' in data:
            if data['is_repeat'] == 'true':
                data['is_repeat'] = True
            else:
                data['is_repeat'] = False

        students = data.get('students')
        if 'students' in data:
            del data['students']

        if season:
            data['challenge_type'] = Challenge.SEMESTR_EXAM

        data['lesson_season'] = lesson_season
        data['lesson_year'] = lesson_year
        try:
            with transaction.atomic():
                saved_data, exception, serializer_errors = save_data_with_signals(self.queryset.model, self.serializer_class, True, None, data=data)
                if serializer_errors:
                    return request.send_error_valid(serializer_errors)

                # Үүссэн шалгалт
                challenge = saved_data[0]

                if students and len(students) > 0:
                    student_datas = Student.objects.filter(id__in=students)
                    challenge.student.set(student_datas)
                    challenge.save()

        except Exception:
            traceback.print_exc()

            return request.send_error('ERR_002')

        return request.send_info('INF_001')

    def put(self, request, pk=None):
        data = request.data
        qs = self.queryset.get(id=pk)

        user = request.user

        data['updated_by'] = user.id

        serializer = self.get_serializer(qs, data=data, partial=True)

        if serializer.is_valid(raise_exception=False):
            self.perform_update(serializer)
            return request.send_info("INF_002")
        else:
            error_fields = []
            for key in serializer.errors:
                return_error = {
                    "field": key,
                    "msg": serializer.errors[key][0]
                }
                error_fields.append(return_error)
            return request.send_error_valid(error_fields)


@permission_classes([IsAuthenticated])
class ChallengeSearchStudentAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer

    filter_backends = [SearchFilter]
    search_fields = ['code', 'first_name', 'register_num']

    def get(self, request):
        datas = self.list(request).data
        return request.send_data(datas)


@permission_classes([IsAuthenticated])
class ChallengeAddKindAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
):

    queryset = Challenge.objects.all()
    serializer_class = ChallengeSerializer

    def put(self, request, pk=None):
        user = request.user

        lesson_year = request.query_params.get('year')
        lesson_season = request.query_params.get('season')
        datas = request.data

        scope = datas.get('scope')
        lesson_id = datas.get('lesson')
        selected = datas.get('groups')

        if selected:
            select_ids = list()
            for group in selected:
                select_ids.append(group.get('id'))

            if len(select_ids) == 0:
                return request.send_error("ERR_003", 'Ангиа сонгоно уу!')

        qs_student = Student.objects.all()
        teacher = Teachers.objects.filter(user_id=user).first()

        timetable_qs = TimeTable.objects.filter(lesson_year=lesson_year, lesson_season=lesson_season, lesson=lesson_id, teacher=teacher)
        timetable_ids = timetable_qs.values_list('id', flat=True)

        exclude_student_ids = TimeTable_to_student.objects.filter(timetable_id__in=timetable_ids, add_flag=False).values_list('student', flat=True)

        if scope == 'all':
            all_lesson_students = get_lesson_choice_student(lesson_id, teacher.id, '', lesson_year, lesson_season)
            student_ids = qs_student.filter(id__in=all_lesson_students).values_list('id', flat=True)

        elif scope == 'group':
            all_student_ids = qs_student.filter(group_id__in=select_ids).values_list('id', flat=True)
            student_ids = set(all_student_ids) - set(list(exclude_student_ids))

        students = Student.objects.filter(id__in=student_ids)

        try:
            challenge = Challenge.objects.get(id=pk)
            for student in students:
                challenge.student.add(student)
            challenge.save()

        except Exception as e:
            print(e)
            return request.send_error("ERR_002")

        return request.send_info("INF_002")


@permission_classes([IsAuthenticated])
class LessonChallengeApiView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    """ Хичээлийн жагсаалт """

    queryset = LessonStandart.objects.all()

    def get(self, request):
        lesson_year, lesson_season = get_active_year_season()

        # Тухайн жилийн дүн оруулсан хичээлийн мэдээлэл
        score_type_ids = TeacherScore.objects.filter(lesson_year=lesson_year, lesson_season=lesson_season).values_list('score_type', flat=True)
        lesson_ids = Lesson_teacher_scoretype.objects.filter(id__in=score_type_ids).values_list('lesson_teacher__lesson', flat=True)

        data = self.queryset.filter(id__in=lesson_ids).values('id', 'code', 'name')

        return request.send_data(list(data))


@permission_classes([IsAuthenticated])
class LessonsApiView(
    generics.GenericAPIView
):

    def get(self, request):

        dep = request.GET.get('dep')
        teacher = request.GET.get('teacher')

        extra_filter = {}
        if dep:
            extra_filter = {
                "department": dep,
            }

        if teacher:
            teacher_of_lessons_qs = Lesson_to_teacher.objects.filter(teacher=teacher).values_list("lesson", flat=True)
            extra_filter['id__in'] = teacher_of_lessons_qs

        data = (
            LessonStandart
                .objects
                .filter(
                    **extra_filter
                )
                .values('id', fname=Concat(F("code"), Value(" "), F("name"), output_field=CharField()))
        )
        data_list = list(data)

        return request.send_data(data_list)


@permission_classes([IsAuthenticated])
class AnalysisApiView(
    generics.GenericAPIView
):

    def get(self, request):

        extra_filters = {}

        year = request.GET.get("year")
        season = request.GET.get("season")
        lesson = request.GET.get("lesson")
        teacher = request.GET.get("teacher")

        assesments = list(
            Score
                .objects
                .values("assesment")
                .annotate(count=Count("assesment"))
                .order_by("assesment")
                .values_list("assesment", flat=True)
        )

        if year:
            extra_filters['lesson_year'] = year

        if year:
            extra_filters['lesson_season'] = season

        if lesson:
            extra_filters['score_type__lesson_teacher__lesson'] = lesson

        # if teacher:
        #     extra_filters['score_type__lesson_teacher__teacher'] = teacher

        chart_data = (
            TeacherScore
                .objects
                .filter(**extra_filters)
                .annotate(
                    total_score=Subquery(
                        TeacherScore.objects.filter(**extra_filters)
                        .filter(student=OuterRef('student'))
                        .values('student')
                        .annotate(total=Sum('score'))
                        .values('total')
                    )
                )
                .annotate(
                    assessment=Subquery(
                        Score.objects.filter(score_max__gte=OuterRef('total_score'), score_min__lte=OuterRef('total_score')).values('assesment')[:1]
                    )
                )
                .values("assessment", 'total_score')
                # .annotate(
                #     count=Count("assessment"),
                # )
                # .values('assessment', 'count')
        )
        print(chart_data)

        data = {
            "data": [
                {
                    "name": "Багшийн дүн",
                    "data": list(chart_data)
                },
                {
                    "name": "Дээд дүн",
                    "data": settings.GPA_ANALYSIS_1_MAX,
                },
                {
                    "name": "Доод дүн",
                    "data": settings.GPA_ANALYSIS_1_MIN,
                }
            ],
            "names": assesments,
        }

        return request.send_data(data)

@permission_classes([IsAuthenticated])
class Analysis2ApiView(
    generics.GenericAPIView
):


    def get(self, request):

        extra_filters = {}

        dep = request.GET.get('dep')
        year = request.GET.get("year")
        season = request.GET.get("season")
        lesson = request.GET.get("lesson")
        teacher = request.GET.get("teacher")

        if year:
            extra_filters['lesson_year'] = year

        if year:
            extra_filters['lesson_season'] = season

        if lesson:
            extra_filters['score_type__lesson_teacher__lesson'] = lesson

        if dep:
            extra_filters['student__group__department'] = dep

        # if teacher:
        #     extra_filters['score_type__lesson_teacher__teacher'] = teacher


        chart_data = list(
            TeacherScore
                .objects
                .filter(**extra_filters)
                .annotate(
                    is_exam=Exists(
                        TeacherScore.objects.filter(score_type__lesson_teacher__lesson=lesson, student=OuterRef('student'), score_type__score_type=Lesson_teacher_scoretype.SHALGALT_ONOO)
                    ),
                    exam_score=(
                        Case(
                            When(
                                is_exam=True,
                                then='score',
                            ),
                            When(
                                is_exam=False,
                                then=Value(0),
                            ),
                            default=Value(1001), output_field=IntegerField(),
                        )
                    ),
                    teach_score=Subquery(
                        TeacherScore.objects.filter(**extra_filters)
                        .filter(student=OuterRef('student'))
                        .exclude(score_type__score_type=Lesson_teacher_scoretype.SHALGALT_ONOO)
                        .values('student')
                        .annotate(total=Sum('score'))
                        .values('total')
                    )
                )
                .values("teach_score", "exam_score")
                .annotate(
                    count=Count("teach_score"),
                )
                .values('count', "teach_score", "exam_score")
        )

        # Filter out None values
        x = [item['teach_score'] for item in chart_data if item['teach_score'] is not None]
        y = [item['exam_score'] for item in chart_data if item['exam_score'] is not None]

        r = 0
        line = []
        if len(y) > 0 and len(x) > 0:
            r, line = pearson_corel(x, y)

        data = {
            "data": chart_data,
            "r": r,
            "line": line
        }

        return request.send_data(data)

@permission_classes([IsAuthenticated])
class ChallengeDetailTableApiView(
    generics.GenericAPIView,
    mixins.ListModelMixin,
):
    queryset = ChallengeStudents.objects.all().order_by('-score')
    serializer_class = ChallengeDetailTableStudentsSerializer

    pagination_class = CustomPagination

    filter_backends = [SearchFilter]
    search_fields = ['student__code', 'student__first_name', 'student__register_num']

    def get(self, request):
        self.queryset = self.queryset.filter(challenge__challenge_type=Challenge.SEMESTR_EXAM)

        #Тухайн шалгалтын id
        test_id = request.query_params.get('test_id')
        department_id = request.query_params.get('department')
        group_id =  request.query_params.get('group')

        if test_id:
            self.queryset= self.queryset.filter(challenge=test_id)

        if department_id:
            self.queryset = self.queryset.filter(student__department=department_id)

        if group_id:
            self.queryset = self.queryset.filter(student__group=group_id)

        datas = self.list(request).data

        return request.send_data(datas)

@permission_classes([IsAuthenticated])
class ChallengeStudentReportAPI(
    mixins.ListModelMixin,
    generics.GenericAPIView,
):

    ''' Оюутны бүртгэл тайлан '''
    def get(self, request):
        school = request.query_params.get('school')
        challenge_id = request.query_params.get('test')
        department = request.query_params.get('department')
        group = request.query_params.get('group')

        # Initialize the extra filter dictionary
        extra_filter = {}

        # Add the school filter if provided
        if school:
            extra_filter.update({'student__group__school': school})

        if department:
            extra_filter.update({'student__group__department': department})

        if group:
            extra_filter.update({'student__group': group})

        # Define grade thresholds
        GRADE_THRESHOLDS = {
            'A': 90,
            'B': 80,
            'C': 70,
            'D': 60,
            'F': 0,
        }

        def get_grade(score, take_score):
            """Map score to grade based on percentage."""
            if take_score and score is not None:
                percentage = (score / take_score) * 100
                for grade, threshold in GRADE_THRESHOLDS.items():
                    if percentage >= threshold:
                        return grade
            return 'F'

        # Filter students based on challenge type and other filters
        challenge_filter = {
            'challenge__challenge_type': Challenge.SEMESTR_EXAM,
        }
        if challenge_id:
            challenge_filter['challenge_id'] = challenge_id

        students = ChallengeStudents.objects.filter(**extra_filter, **challenge_filter)

        # Initialize grade counts by gender
        grade_counts_by_gender = {
            'male': {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0},
            'female': {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0},
        }

        # Count grades for each student
        for student in students:
            grade = get_grade(student.score, student.take_score)
            gender = None

            if student.student.gender == Student.GENDER_MALE:
                gender = 'male'
            elif student.student.gender == Student.GENDER_FEMALE:
                gender = 'female'

            if gender:
                grade_counts_by_gender[gender][grade] += 1

        # Prepare data for response
        data = {
            "male": grade_counts_by_gender['male'],
            "female": grade_counts_by_gender['female'],
        }

        return request.send_data(data)


@permission_classes([IsAuthenticated])
class LessonStandartExamListAPIView(
    generics.GenericAPIView
):
    """ Хичээлийн стандарт """

    queryset = LessonStandart.objects.all()
    def get(self, request):
        lesson_year = request.query_params.get('lesson_year')
        lesson_season = request.query_params.get('lesson_season')
        qs = TeacherScore.objects.all()
        if lesson_year:
            qs = qs.filter(lesson_year=lesson_year)

        if lesson_season:
            qs = qs.filter(lesson_season=lesson_season)

        lesson_standarts = qs.values_list('score_type__lesson_teacher__lesson', flat=True)

        all_data = self.queryset.filter(id__in=lesson_standarts).values('id', 'code', 'name')
        return request.send_data(list(all_data))



@permission_classes([IsAuthenticated])
class LessonStandartGroupScoreAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    """ Дүн гарсан ангийн хичээлийг шинэчлэх """

    queryset = LessonStandart.objects.all()
    serializer_class = LessonStandartSerializer

    def get(self, request, pk=None):

        filters = dict()

        level = request.query_params.get('level')

        if level and level != '7':
            filters['level'] = level

        score_req = ScoreRegister.objects.filter(lesson=pk)

        group_ids = score_req.values_list('student__group_id', flat=True)

        count_qs = (
            score_req
            .filter(
                Q(student__group=OuterRef('pk')),
                Q(
                    Q(student__status__name__icontains='Суралцаж буй') |
                    Q(student__status__name__icontains='Суралцаж байгаа')
                ),
            )
            .values('student__group')
            .order_by('student__group')
            .annotate(
                totals=Count('student__group')
            )
            .values('totals')
        )

        group_datas = (
            Group
            .objects
            .filter(pk__in=group_ids, **filters)
            .annotate(
                active_student_count=Subquery(count_qs),
            )
            .filter(active_student_count__gt=0)
        )

        return_datas = {
            "groups": list(group_datas.values('id', 'name', 'active_student_count', 'level')) if group_datas.exists() else [],
            "levels": list(group_datas.values_list('level', flat=True).distinct()) if group_datas.exists() else [],
        }

        return request.send_data(return_datas)

    @transaction.atomic()
    def put(self, request, pk=None):

        data = request.data
        groups = data.get('group')
        cources = data.get('cources')
        new_lesson = data.get('lesson')

        filters = dict()
        level_filters = Q()

        print(groups)

        if groups:
            filters['student__group__id__in'] = groups

        if cources:
            for cource in cources:
                level_filters |= Q(student__group__level=cource)

        try:
            score_register_qs = (
                ScoreRegister
                .objects
                .filter(
                    Q(
                        lesson=pk,
                        **filters
                    ),
                    level_filters
                )
            )

            if score_register_qs.exists():
                user = request.user
                score_register_qs.update(
                    lesson=new_lesson,
                    updated_user=user
                )

            return request.send_info("INF_002", score_register_qs.count())

        except Exception as e:
            return request.send_error("ERR_002", e.__str__())


@permission_classes([IsAuthenticated])
class LessonSearchAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    queryset = LessonStandart.objects
    serializer_class = LessonStandartSerialzier

    def get(self, request):
        """ Оюутны жагсаалт
        """
        school = request.query_params.get('school')
        state = request.query_params.get('state')

        if state == '2':
            qs_start = (int(state) - 2) * 10
            qs_filter = int(state) * 10
        else:
            qs_start = (int(state) - 1) * 10
            qs_filter = int(state) * 10

        if school:
            self.queryset = self.queryset.filter(school=school)

        self.queryset = self.queryset.order_by('id')[qs_start:qs_filter]
        return_datas = self.list(request).data

        return request.send_data(return_datas)


@permission_classes([IsAuthenticated])
class LessonSearchStudentAPIView(
    generics.GenericAPIView,
    mixins.ListModelMixin
):
    queryset = LessonStandart.objects
    serializer_class = LessonStandartSerialzier
    filter_backends = [SearchFilter]
    search_fields = ['code', 'name']
    def get(self, request):
        """ Оюутны жагсаалт
        """
        school = request.query_params.get('school')
        search = request.query_params.get('search')

        if school:
            self.queryset = self.queryset.filter(Q(school=school))

        return_datas = self.list(request).data

        return request.send_data(return_datas)


@permission_classes([IsAuthenticated])
class GraduateTitleApiView(
    generics.GenericAPIView,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin
):
    """ Төгсөлтийн бүлэг """

    queryset = QuestionMainTitle.objects.all()
    serializer_class = dynamic_serializer(QuestionMainTitle, "__all__", 1)

    def get(self, request):
        user = request.user
        teacher = get_object_or_404(Teachers, user_id=user, action_status=Teachers.APPROVED)
        if not user.is_superuser:
            self.queryset = self.queryset.filter(created_by=teacher)
        datas = self.queryset.values('id', 'name')
        return request.send_data(list(datas))

    def post(self, request):
        ser = dynamic_serializer(QuestionMainTitle, "__all__")
        user_id = request.user
        teacher = get_object_or_404(Teachers, user_id=user_id, action_status=Teachers.APPROVED)
        datas = request.data
        datas['created_by'] = teacher.id

        serializer = ser(data=datas)
        if serializer.is_valid(raise_exception=False):
            self.perform_create(serializer)
            return request.send_info("INF_001")
        else:
            return request.send_error_valid(serializer.errors)

    def put(self, request, pk=None):
        ""
        datas = request.data
        instance = self.queryset.filter(id=pk).first()
        user = request.user
        teacher = get_object_or_404(Teachers, user_id=user, action_status=Teachers.APPROVED)
        datas['created_by'] = teacher.id

        ser = dynamic_serializer(QuestionMainTitle, "__all__", 1)
        serializer = ser(instance, data=datas, partial=True)
        if serializer.is_valid(raise_exception=False):
            self.perform_update(serializer)
            return request.send_info("INF_002")

        else:
            return request.send_error_valid(serializer.errors)

    def delete(self, request, pk=None):
        self.destroy(request, pk)
        return request.send_info("INF_003")


@permission_classes([IsAuthenticated])
class GraduateSubTitleApiView(
    generics.GenericAPIView,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin
):
    """ Төгсөлтийн дэд бүлэг """

    queryset = QuestioSubTitle.objects.all()
    serializer_class = dynamic_serializer(QuestioSubTitle, "__all__", 1)

    def get(self, request):
        main = request.query_params.get('main')
        if main:
            self.queryset = self.queryset.filter(main=main)

        question_sub = ChallengeQuestions.objects.filter(graduate_title=OuterRef('id')).values('graduate_title').annotate(count=Count('id')).values('count')

        user = request.user
        teacher = get_object_or_404(Teachers, user_id=user, action_status=Teachers.APPROVED)
        if not user.is_superuser:
            self.queryset = self.queryset.filter(created_by=teacher)

        datas = self.queryset.annotate(main_name=F('main__name'), question_count=Subquery(question_sub)).values('id', 'name', 'main_name', 'question_count')
        return request.send_data(list(datas))

    def post(self, request):
        user_id = request.user
        teacher = get_object_or_404(Teachers, user_id=user_id, action_status=Teachers.APPROVED)
        datas = request.data
        datas['created_by'] = teacher.id

        ser = dynamic_serializer(QuestioSubTitle, "__all__")
        serializer = ser(data=datas)
        if serializer.is_valid(raise_exception=False):
            self.perform_create(serializer)
            return request.send_info("INF_001")
        else:
            return request.send_error_valid(serializer.errors)

    def put(self, request, pk=None):
        ""
        datas = request.data
        instance = self.queryset.filter(id=pk).first()
        user_id = request.user
        teacher = get_object_or_404(Teachers, user_id=user_id, action_status=Teachers.APPROVED)
        datas['created_by'] = teacher.id

        ser = dynamic_serializer(QuestioSubTitle, "__all__")
        serializer = ser(instance, data=datas, partial=True)
        if serializer.is_valid(raise_exception=False):
            self.perform_update(serializer)
            return request.send_info("INF_002")

        else:
            return request.send_error_valid(serializer.errors)

    def delete(self, request, pk=None):
        self.destroy(request, pk)
        return request.send_info("INF_003")


class GraduateQuestionApiView(
    generics.GenericAPIView
):
    """ Төгсөлтийн шалгалтын асуулт"""

    queryset = QuestioSubTitle.objects.all()
    serializer_class = dynamic_serializer(QuestioSubTitle, "__all__")
    def get(self, request):
        title = request.query_params.get('title')
        if title:
            self.queryset = self.queryset.filter(id=title)

        question_sub = ChallengeQuestions.objects.filter(graduate_title=OuterRef('id')).values('graduate_title').annotate(count=Count('id')).values('count')
        data = self.queryset.annotate(question_count=Subquery(question_sub)).values("id", "name", 'question_count')

        return request.send_data(list(data))
